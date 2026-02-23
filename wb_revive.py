#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
WB Revival V1.5 (dynamic scope + deep card.json) — staged pipeline A..K

Canonical storage:
- JSONL for pipeline data
- .wb_cache for raw WB payloads
- XLSX + HTML as human-facing outputs

Network split:
- WB stages: can run without VPN
- LLM stages: can run with VPN
Use --pause-between-stages if you want interactive switching.

Input scope:
- SKU list is taken from the input sheet (any count >= 1).
- The pipeline NEVER goes beyond that list.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import random
import re
import sys
import shlex
import time
from collections import Counter
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_NAME = Path(__file__).name
SCRIPT_VERSION = "1.5.3.13"
SCHEMA_VERSION = "1.0"

DEFAULT_INPUT_XLSX = "WB_INPUT_64_FROM_POCKETS_POD.xlsx"
DEFAULT_INPUT_SHEET = "INPUT_64"

DEFAULT_DESTS = [-1257786, -1216601, -115136, -421732, 123585595]
DEFAULT_SEARCH_HOSTS = ["u-search.wb.ru", "search.wb.ru"]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "ru,en-US;q=0.9,en;q=0.8",
}


# --- network sessions (WB vs LLM) ---
WB_SESSION = requests.Session()
WB_SESSION.headers.update(HEADERS)
WB_SESSION.trust_env = True  # can be disabled via --no-env-proxy
WB_SESSION.proxies = {}

LLM_SESSION = requests.Session()
LLM_SESSION.trust_env = True  # let users route LLM via env proxy if they want

def configure_wb_network(proxy: str = "", *, no_env_proxy: bool = False) -> None:
    """Configure WB HTTP stack (proxy + trust_env)."""
    WB_SESSION.trust_env = (not no_env_proxy)
    WB_SESSION.proxies = {}
    p = safe_str(proxy).strip()
    if p:
        WB_SESSION.proxies.update({"http": p, "https": p})


STAGE_ORDER = ["A","B","C","D","E","F","G","H","I","J","K"]


STAGE_META = {
    "A": {"title": "Input + Manifest", "network": "LOCAL", "llm_flag": None},
    "B": {"title": "Collect OWN cards", "network": "WB", "llm_flag": None},
    "C": {"title": "Generate Queries", "network": "LLM", "llm_flag": "use_llm"},
    "D": {"title": "SERP Validate Queries", "network": "WB", "llm_flag": None},
    "E": {"title": "Build Competitor Pool (hard filters)", "network": "LOCAL", "llm_flag": None},
    "F": {"title": "Collect Competitor Cards", "network": "WB", "llm_flag": None},
    "G": {"title": "LLM Relevance Filter", "network": "LLM", "llm_flag": "use_llm_relevance"},
    "H": {"title": "Select Final Competitors (roles)", "network": "LOCAL", "llm_flag": None},
    "I": {"title": "Compare + Score", "network": "LOCAL", "llm_flag": None},
    "J": {"title": "Verdict + Backlog", "network": "LLM", "llm_flag": "use_llm_verdict"},
    "K": {"title": "Render Reports + Executive Summary", "network": "LOCAL", "llm_flag": "use_llm_exec_summary"},
}

def _stage_title(code: str) -> str:
    return STAGE_META.get(code, {}).get("title") or code

def _stage_network(code: str) -> str:
    return STAGE_META.get(code, {}).get("network") or "LOCAL"

def _stage_llm_enabled(code: str, args: argparse.Namespace) -> bool:
    meta = STAGE_META.get(code, {})
    flag = meta.get("llm_flag")
    if not flag:
        return False
    return bool(getattr(args, flag, False))

def _stage_vpn_hint(code: str, args: argparse.Namespace) -> str:
    net = _stage_network(code)
    if net == "LLM" and _stage_llm_enabled(code, args):
        return "возможно нужен (LLM API)"
    return "не нужен"

def _next_stage_in_run(stages: List[str], current: str) -> Optional[str]:
    try:
        i = stages.index(current)
    except ValueError:
        return None
    if i + 1 >= len(stages):
        return None
    return stages[i + 1]

def resolve_llm_caps(args: argparse.Namespace) -> Dict[str, int]:
    """Resolve per-stage LLM max_tokens caps with a legacy global override."""
    caps = {
        "C": int(getattr(args, "llm_max_tokens_c", 800) or 800),
        "G": int(getattr(args, "llm_max_tokens_g", 3000) or 3000),
        "J": int(getattr(args, "llm_max_tokens_j", 6000) or 6000),
        "K": int(getattr(args, "llm_max_tokens_k", 15000) or 15000),
    }
    g = int(getattr(args, "llm_max_tokens", 0) or 0)
    if g > 0:
        for k in caps:
            caps[k] = g
    # sanity floor to avoid 0/negative
    for k, v in list(caps.items()):
        if v < 256:
            caps[k] = 256
    return caps


# =========================
# Small utils
# =========================

def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def safe_str(x: Any, default: str = "") -> str:
    if x is None:
        return default
    try:
        return str(x)
    except Exception:
        return default

def safe_int(x: Any, default: Optional[int] = None) -> Optional[int]:
    try:
        s = safe_str(x, "").strip()
        if not s:
            return default
        if re.fullmatch(r"-?\d+(\.0+)?", s):
            return int(float(s))
        return int(s)
    except Exception:
        return default

def safe_float(x: Any, default: Optional[float] = None) -> Optional[float]:
    try:
        s = safe_str(x, "").strip()
        if not s:
            return default
        return float(s)
    except Exception:
        return default

def nm_id_to_str(x: Any) -> str:
    """Coerce nm_id to a stable decimal string (avoid Excel scientific notation issues)."""
    if x is None:
        return ""
    if isinstance(x, bool):
        return ""
    if isinstance(x, int):
        return str(x)
    if isinstance(x, float):
        if math.isfinite(x):
            # Excel often gives float for numeric cells; nm_id is integer
            return str(int(x))
        return safe_str(x, "").strip()

    s = safe_str(x, "").strip()
    if not s:
        return ""

    # Plain integer or integer-like float string
    if re.fullmatch(r"\d+(\.0+)?", s):
        try:
            return str(int(Decimal(s)))
        except Exception:
            return str(int(float(s)))

    # Scientific notation (Excel)
    if re.fullmatch(r"\d+(\.\d+)?[eE][\+\-]?\d+", s):
        try:
            return str(int(Decimal(s)))
        except (InvalidOperation, ValueError, OverflowError):
            try:
                return str(int(float(s)))
            except Exception:
                return s

    # Anything else: keep as-is (already string)
    return s

def sha1_short(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()[:12]

def write_text_atomic(path: Path, text: str, encoding: str = "utf-8") -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text, encoding=encoding)
    tmp.replace(path)

def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))

def write_json(path: Path, obj: Any, indent: int = 2) -> None:
    write_text_atomic(path, json.dumps(obj, ensure_ascii=False, indent=indent), encoding="utf-8")

def read_jsonl(path: Path) -> List[dict]:
    """
    Robust JSONL reader:
    - skips empty lines
    - skips/ignores broken JSON lines (e.g., interrupted writes)
    """
    out: List[dict] = []
    if not path.exists():
        return out
    try:
        with path.open("r", encoding="utf-8") as f:
            for ln, line in enumerate(f, start=1):
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                    if isinstance(obj, dict):
                        out.append(obj)
                except Exception:
                    # best effort: ignore broken line to keep resume usable
                    continue
    except Exception:
        return out
    return out

def append_jsonl(path: Path, rec: dict) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")

def _pool_norm_lc(s: str) -> str:
    s = safe_str(s, "").lower()
    s = re.sub(r"[^a-zа-я0-9]+", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _canon_nospace(s: str) -> str:
    """
    Canonicalizes a string for fuzzy-ish matching:
    - lowercase, keep alnum (via _pool_norm_lc)
    - remove spaces
    - normalize common marketplace variants like "pro max" vs "promax"
    """
    t = _pool_norm_lc(s).replace(" ", "")
    # common variants
    t = t.replace("pro-max", "promax")
    # after _pool_norm_lc, hyphens are already spaces; keep just in case
    t = t.replace("promax", "promax")
    t = t.replace("proplus", "proplus")
    return t

# --- token synonyms for cross-script matching (latin <-> cyrillic + common suffixes) ---
# Keep this SMALL and PURPOSEFUL. If you put colors/materials here, you'll inflate rel50 with garbage.
TOKEN_SYNONYMS: Dict[str, List[str]] = {
    # brands
    "iphone": ["айфон"],
    "айфон": ["iphone"],
    "samsung": ["самсунг"],
    "самсунг": ["samsung"],
    "galaxy": ["галакси"],
    "галакси": ["galaxy"],
    "xiaomi": ["сяоми", "ксиаоми"],
    "сяоми": ["xiaomi"],
    "redmi": ["редми"],
    "редми": ["redmi"],
    "poco": ["поко"],
    "поко": ["poco"],
    "honor": ["хонор"],
    "хонор": ["honor"],
    "huawei": ["хуавей"],
    "хуавей": ["huawei"],
    "realme": ["реалми"],
    "реалми": ["realme"],
    "tecno": ["техно"],
    "техно": ["tecno"],
    "infinix": ["инфиникс"],
    "инфиникс": ["infinix"],
    # common model suffixes
    "pro": ["про"],
    "про": ["pro"],
    "max": ["макс"],
    "макс": ["max"],
    "mini": ["мини"],
    "мини": ["mini"],
    "plus": ["плюс"],
    "плюс": ["plus"],
    "ultra": ["ультра"],
    "ультра": ["ultra"],
    "lite": ["лайт"],
    "лайт": ["lite"],
    "note": ["ноут"],
    "ноут": ["note"],
}

def _term_variants(tlc: str, max_variants: int = 12) -> List[str]:
    """
    Generate a small set of term variants by swapping known tokens (iphone<->айфон, pro<->про, ...).
    We allow up to 2 swaps (depth=2). This is enough for "iphone 13 pro" vs "айфон 13 про".
    """
    tlc = safe_str(tlc, "").strip()
    if not tlc:
        return []
    variants: List[str] = [tlc]
    seen = {tlc}

    for _ in range(2):  # depth
        cur = list(variants)
        for v in cur:
            toks = v.split()
            for i, tok in enumerate(toks):
                syns = TOKEN_SYNONYMS.get(tok)
                if not syns:
                    continue
                for syn in syns:
                    nv = " ".join(toks[:i] + [syn] + toks[i+1:]).strip()
                    if nv and nv not in seen:
                        seen.add(nv)
                        variants.append(nv)
                        if len(variants) >= max_variants:
                            return variants
    return variants

def _term_in_blob(term: str, blob_lc: str, blob_ns: str) -> bool:
    """Return True if term is found in blob (case-insensitive), supporting:
    - space-insensitive matches
    - simple OR lists like "iPhone 15 Pro; Айфон 15 Про"
    - a small set of token swaps (iphone<->айфон, pro<->про, ...)
    """
    t = safe_str(term)
    if not t.strip():
        return False

    # Support "alt1; alt2; alt3" or "alt1|alt2" as OR
    if any(sep in t for sep in [";", "|"]):
        parts = [p.strip() for p in re.split(r"[;|]+", t) if safe_str(p).strip()]
        for p in parts:
            if _term_in_blob(p, blob_lc, blob_ns):
                return True
        return False

    tlc = _pool_norm_lc(t)
    if not tlc:
        return False

    # direct + variants
    for vv in _term_variants(tlc):
        if vv in blob_lc:
            return True
        vns = _canon_nospace(vv)
        if vns and (vns in blob_ns):
            return True

    return False




def _median_int(vals: List[int]) -> Optional[int]:
    xs = [int(v) for v in vals if v is not None]
    if not xs:
        return None
    xs.sort()
    n = len(xs)
    mid = n // 2
    if n % 2 == 1:
        return xs[mid]
    return int(round((xs[mid-1] + xs[mid]) / 2))

def _percentile_int(vals: List[int], p: int) -> Optional[int]:
    xs = [int(v) for v in vals if v is not None]
    if not xs:
        return None
    xs.sort()
    if p <= 0:
        return xs[0]
    if p >= 100:
        return xs[-1]
    k = (len(xs) - 1) * (p / 100.0)
    f = math.floor(k)
    c = math.ceil(k)
    if f == c:
        return xs[int(k)]
    d0 = xs[f] * (c - k)
    d1 = xs[c] * (k - f)
    return int(round(d0 + d1))

# =========================
# Meta + validation
# =========================

def make_meta(run_id: str, stage: str, nm_id: str, vendor_code: str, *, ts: Optional[str] = None, name: str = "", source: str = "script") -> dict:
    return {
        "schema_version": SCHEMA_VERSION,
        "run_id": safe_str(run_id),
        "stage": safe_str(stage),
        "ts": ts or utc_now_iso(),
        "nm_id": nm_id_to_str(nm_id),
        "vendor_code": safe_str(vendor_code),
        "name": safe_str(name),
        "source": safe_str(source),
    }

def validate_meta(meta: dict) -> None:
    if not isinstance(meta, dict):
        raise ValueError("meta must be dict")
    if safe_str(meta.get("schema_version")) != SCHEMA_VERSION:
        raise ValueError("meta.schema_version mismatch")
    if safe_str(meta.get("run_id")).strip() == "":
        raise ValueError("meta.run_id empty")
    st = safe_str(meta.get("stage")).strip()
    if st not in STAGE_ORDER:
        raise ValueError(f"meta.stage invalid: {st}")
    nm = nm_id_to_str(meta.get("nm_id"))
    if not re.fullmatch(r"\d+", nm):
        raise ValueError(f"meta.nm_id invalid: {nm!r}")

# =========================
# Minimal .env loader
# =========================

def load_env_file(env_path: Path, override: bool = False) -> None:
    try:
        if not env_path.exists():
            return
        for raw in env_path.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            k, v = line.split("=", 1)
            k = k.strip()
            v = v.strip().strip('"').strip("'")
            if not k:
                continue
            if (k in os.environ) and not override:
                continue
            os.environ[k] = v
    except Exception:
        return

# =========================
# WB endpoints + HTTP helpers
# =========================

def wb_product_url(nm_id: str) -> str:
    return f"https://www.wildberries.ru/catalog/{nm_id_to_str(nm_id)}/detail.aspx"

def wb_api_v4_url(nm_id: str, dest: int) -> str:
    return f"https://card.wb.ru/cards/v4/detail?appType=1&curr=rub&dest={dest}&nm={nm_id_to_str(nm_id)}"

def wb_api_v1_url(nm_id: str, dest: int) -> str:
    return f"https://card.wb.ru/cards/v1/detail?appType=1&curr=rub&dest={dest}&nm={nm_id_to_str(nm_id)}"



# --- Deep card payload (wbbasket) ---
# WB "card.wb.ru/cards/*/detail" is intentionally skinny for many products.
# The richer payload (description + characteristics/options) lives here:
#   https://basket-XX.wbbasket.ru/vol{vol}/part{part}/{nm}/info/ru/card.json
#
# Basket host is derived from vol = nm_id // 100000 using empiric ranges.
# Ranges were widely published by the parsing community and periodically updated.
# We implement a conservative mapping (01..18) plus a tiny fallback probe (±1 host).

def wb_basket_host_by_vol(vol: int) -> str:
    """Return basket host number as zero-padded string: '01'..'18'."""
    try:
        v = int(vol)
    except Exception:
        return "18"
    if 0 <= v <= 143:   return "01"
    if 144 <= v <= 287: return "02"
    if 288 <= v <= 431: return "03"
    if 432 <= v <= 719: return "04"
    if 720 <= v <= 1007:return "05"
    if 1008 <= v <= 1061:return "06"
    if 1062 <= v <= 1115:return "07"
    if 1116 <= v <= 1169:return "08"
    if 1170 <= v <= 1313:return "09"
    if 1314 <= v <= 1601:return "10"
    if 1602 <= v <= 1655:return "11"
    if 1656 <= v <= 1919:return "12"
    if 1920 <= v <= 2045:return "13"
    if 2046 <= v <= 2189:return "14"
    if 2190 <= v <= 2405:return "15"
    if 2406 <= v <= 2621:return "16"
    if 2622 <= v <= 2837:return "17"
    return "18"

def wb_basket_card_json_url(nm_id: str, *, host_num: Optional[str] = None, lang: str = "ru") -> str:
    nm = safe_int(nm_id, None)
    if nm is None:
        nm = int(nm_id_to_str(nm_id) or "0")
    vol = nm // 100000
    part = nm // 1000
    host = (host_num or wb_basket_host_by_vol(vol)).zfill(2)
    return f"https://basket-{host}.wbbasket.ru/vol{vol}/part{part}/{nm}/info/{lang}/card.json"

def _is_valid_deep_card_json(nm_id: str, js: dict) -> bool:
    if not isinstance(js, dict) or not js:
        return False
    # Sometimes it contains explicit ids, sometimes not. Be permissive but not blind.
    for k in ("nm_id", "nmId", "nm", "id"):
        v = js.get(k)
        if v is not None and nm_id_to_str(v) == nm_id_to_str(nm_id):
            return True
    # Heuristic: should have at least some content fields
    if safe_str(js.get("description") or js.get("selling") or js.get("shortDescription") or js.get("short_description")):
        return True
    opts = js.get("options")
    if isinstance(opts, list) and len(opts) > 0:
        return True
    props = js.get("properties") or js.get("chars") or js.get("characteristics")
    if isinstance(props, list) and len(props) > 0:
        return True
    return False

def fetch_deep_card_json(nm_id: str, *, timeout: int, verbose: bool = False, lang: str = "ru") -> Tuple[str, int, Optional[dict], str, Optional[str]]:
    """Fetch wbbasket card.json (richer content). Returns (status, http, json, url, host)."""
    nm_s = nm_id_to_str(nm_id)
    if not nm_s.isdigit():
        return "invalid", 0, None, "", None

    nm = int(nm_s)
    vol = nm // 100000
    base_host = wb_basket_host_by_vol(vol)
    # A tiny "safety probe" in case WB moved a range boundary.
    try:
        b = int(base_host)
    except Exception:
        b = 18
    hosts = [f"{b:02d}"]
    if 1 <= b-1 <= 18: hosts.append(f"{b-1:02d}")
    if 1 <= b+1 <= 18: hosts.append(f"{b+1:02d}")
    hosts.append("18")
    # unique preserve order
    seen = set()
    hosts = [h for h in hosts if not (h in seen or seen.add(h))]

    last_code, last_url = 0, ""
    for h in hosts:
        url = wb_basket_card_json_url(nm_s, host_num=h, lang=lang)
        code, js, _ = _get(url, timeout=timeout)
        last_code, last_url = code, url
        if verbose:
            print(f"[WB] deep card nm={nm_s} basket={h} -> {code}")
        if code == 200 and js and _is_valid_deep_card_json(nm_s, js):
            return "ok", code, js, url, h

    if last_code == 200:
        return "invalid", last_code, None, last_url, None
    return "not_found", last_code or 404, None, last_url, None


def wb_search_v18_url(query: str, dest: int, *, page: int = 1, limit: int = 100, sort: str = "popular", host: str = "u-search.wb.ru") -> str:
    return (f"https://{host}/exactmatch/ru/common/v18/search"
            f"?appType=1&curr=rub&dest={dest}"
            f"&lang=ru&inheritFilters=false&suppressSpellcheck=false"
            f"&query={requests.utils.quote(query)}"
            f"&page={page}&resultset=catalog&sort={sort}&spp=30&limit={limit}")

def backoff_sleep(attempt: int, base: float = 0.35, cap: float = 6.0) -> None:
    time.sleep(min(cap, base * (2 ** attempt) + random.random() * 0.2))

def _get(url: str, timeout: int, retries: int = 3) -> Tuple[int, Optional[dict], str]:
    last = ""
    for a in range(retries):
        try:
            r = WB_SESSION.get(url, timeout=timeout)
            last = r.text[:2000]
            if r.status_code == 200:
                try:
                    return 200, r.json(), last
                except Exception:
                    return 200, None, last
            if r.status_code in (429, 500, 502, 503, 504):
                backoff_sleep(a)
                continue
            return r.status_code, None, last
        except requests.RequestException:
            backoff_sleep(a)
    return 0, None, last

def _extract_products(js: dict) -> List[dict]:
    if not isinstance(js, dict):
        return []
    data = js.get("data")
    if isinstance(data, dict):
        for k in ("products", "productsW", "productsV2"):
            v = data.get(k)
            if isinstance(v, list):
                return v
    v = js.get("products")
    if isinstance(v, list):
        return v
    return []

def _is_valid_card_json(nm_id: str, js: dict) -> bool:
    prods = _extract_products(js)
    if not prods:
        return False
    p0 = prods[0] if isinstance(prods[0], dict) else {}
    if not isinstance(p0, dict) or not p0:
        return False
    pid = nm_id_to_str(p0.get("id") or p0.get("nmId") or p0.get("nm"))
    if pid and pid != nm_id_to_str(nm_id):
        return False
    name = safe_str(p0.get("name") or p0.get("goodsName") or "")
    brand = safe_str(p0.get("brand") or p0.get("brandName") or "")
    seller = safe_str(p0.get("supplier") or p0.get("supplierName") or p0.get("seller") or "")
    return bool(name.strip() or brand.strip() or seller.strip())

def fetch_card_any(nm_id: str, dests: List[int], timeout: int, verbose: bool = False) -> Tuple[str, int, Optional[dict], str, Optional[int]]:
    last_code, last_url = 0, ""
    for dest in dests:
        for url_fn in (wb_api_v4_url, wb_api_v1_url):
            url = url_fn(nm_id, dest)
            code, js, _ = _get(url, timeout=timeout)
            last_code, last_url = code, url
            if verbose:
                print(f"[WB] card nm={nm_id} dest={dest} -> {code}")
            if code == 200 and js and _is_valid_card_json(nm_id, js):
                return "ok", code, js, url, dest
    if last_code == 200:
        return "invalid", last_code, None, last_url, None
    return "not_found", last_code or 404, None, last_url, None

def parse_search_items(js: dict) -> List[dict]:
    if not isinstance(js, dict):
        return []
    products = None
    if isinstance(js.get("products"), list):
        products = js.get("products")
    else:
        data = js.get("data")
        if isinstance(data, dict) and isinstance(data.get("products"), list):
            products = data.get("products")
    if not isinstance(products, list):
        return []
    out = []
    for idx, p in enumerate(products, start=1):
        if not isinstance(p, dict):
            continue
        nm = nm_id_to_str(p.get("id") or p.get("nmId"))
        if not nm or not nm.isdigit():
            continue
        price_u = p.get("priceU") or p.get("price")
        sale_u = p.get("salePriceU") or p.get("salePrice")
        if (price_u is None or sale_u is None) and isinstance(p.get("sizes"), list) and p["sizes"]:
            s0 = p["sizes"][0] if isinstance(p["sizes"][0], dict) else {}
            pr = s0.get("price") if isinstance(s0, dict) else None
            if isinstance(pr, dict):
                price_u = price_u or pr.get("basic")
                sale_u = sale_u or pr.get("product")
        out.append({
            "pos": idx,
            "id": nm,
            "nm_id": nm,
            "name": safe_str(p.get("name") or ""),
            "brand": safe_str(p.get("brand") or ""),
            "seller": safe_str(p.get("supplier") or p.get("supplierName") or ""),
            "priceU": safe_int(price_u),
            "salePriceU": safe_int(sale_u),
            "rating": safe_float(p.get("rating") if p.get("rating") is not None else p.get("reviewRating") or p.get("nmReviewRating"), None),
            "feedbacks": safe_int(p.get("feedbacks") if p.get("feedbacks") is not None else p.get("nmFeedbacks"), None),
            "pics": safe_int(p.get("pics") or p.get("picsCount"), None),
            "hasVideo": bool(p.get("hasVideo")) if p.get("hasVideo") is not None else None,
            "subjectId": safe_int(p.get("subjectId") or p.get("subject") or p.get("subjectID"), None),
            "subjectName": safe_str(p.get("subjectName") or p.get("subjName") or ""),
            "url": wb_product_url(nm),
            "raw": p,
        })
    return out

def _extract_search_price_rub(p: dict) -> Optional[int]:
    """Extract price in RUB (integer) from a WB search item.

    WB search endpoints typically return prices in kopecks (priceU/salePriceU).
    Older/other shapes may return rubles. We do a best-effort detection.
    """
    if not isinstance(p, dict):
        return None

    raw = p.get("raw") if isinstance(p.get("raw"), dict) else None

    # prefer explicit *U fields
    price_u = safe_int(p.get("priceU"), None)
    sale_u = safe_int(p.get("salePriceU"), None)

    # fall back to non-U keys if present
    if price_u is None:
        price_u = safe_int(p.get("price"), None)
    if sale_u is None:
        sale_u = safe_int(p.get("salePrice"), None)

    used_sizes_price = False
    if (price_u is None or sale_u is None) and isinstance(p.get("sizes"), list) and p["sizes"]:
        s0 = p["sizes"][0] if isinstance(p["sizes"][0], dict) else {}
        pr = s0.get("price") if isinstance(s0, dict) else None
        if isinstance(pr, dict):
            # basic/product are kopecks in WB payloads
            price_u = price_u if price_u is not None else safe_int(pr.get("basic"), None)
            sale_u = sale_u if sale_u is not None else safe_int(pr.get("product"), None)
            used_sizes_price = True

    use = sale_u if sale_u is not None else price_u
    if use is None:
        return None

    # unit detection:
    # - if raw has explicit priceU/salePriceU, treat as kopecks
    # - if we used sizes.price.basic/product, treat as kopecks
    raw_has_u = False
    if isinstance(raw, dict):
        raw_has_u = ("priceU" in raw) or ("salePriceU" in raw)

    if raw_has_u or used_sizes_price:
        return int(max(0, use) // 100)

    # heuristic fallback:
    # if the number looks like kopecks (too big for rubles), divide by 100
    if use >= 10000:
        return int(use // 100)

    # otherwise assume it's already rubles
    return int(use)
def _monster_share_topN(items: List[dict], topN: int = 20, threshold_feedbacks: int = 1000) -> Optional[float]:
    """
    Returns share (0..1) of 'monster' listings in topN.
    Monster heuristic: feedbacks >= threshold_feedbacks.
    """
    if not items:
        return None
    top = items[:max(1, int(topN))]
    fb = []
    for p in top:
        if not isinstance(p, dict):
            continue
        v = safe_int(p.get("feedbacks"), None)
        if v is not None:
            fb.append(v)
    if not fb:
        return None
    monsters = sum(1 for x in fb if x >= int(threshold_feedbacks))
    return round(monsters / max(1, len(fb)), 3)

def _query_tokens(q: str) -> List[str]:
    q = safe_str(q).lower()
    q = re.sub(r"[^a-zа-я0-9 ]+", " ", q)
    toks = [t for t in q.split() if len(t) >= 4]
    stop = {"для", "в", "на", "и", "с", "по", "как", "или", "без", "под", "все", "авто"}
    toks = [t for t in toks if t not in stop]
    return toks[:10]

def _relevance_score(query: str, items: List[dict]) -> int:
    q = safe_str(query).strip()
    if not items:
        return 0
    if q.isdigit():
        qnm = nm_id_to_str(q)
        for it in items[:30]:
            if nm_id_to_str(it.get("nm_id")) == qnm:
                return 100
        return 0
    toks = _query_tokens(q)
    if not toks:
        return 0
    topnames = " ".join(safe_str(x.get("name")) for x in items[:12]).lower()
    return sum(1 for t in toks if t in topnames)

def fetch_search_best(query: str, dests: List[int], timeout: int, limit: int, verbose: bool = False) -> Tuple[str, int, Optional[dict], str, int, int]:
    best = None  # (score, count, code, js, url, dest)
    for host in DEFAULT_SEARCH_HOSTS:
        for dest in dests:
            url = wb_search_v18_url(query, dest, page=1, limit=limit, host=host)
            code, js, _ = _get(url, timeout=timeout)
            if verbose:
                print(f"[WB] search host={host} q={query!r} dest={dest} -> {code}")
            if code != 200 or not js:
                continue
            items = parse_search_items(js)
            score = _relevance_score(query, items)
            cand = (score, len(items), code, js, url, dest)
            if best is None or cand[:2] > best[:2]:
                best = cand
            if (score >= 1 or (safe_str(query).isdigit() and score >= 100)) and len(items) >= min(10, limit):
                return "ok", code, js, url, dest, score
    if best:
        score, _, code, js, url, dest = best
        return "ok", code, js, url, dest, score
    return "not_found", 404, None, "", 0, 0

# =========================
# Normalization helpers
# =========================

BANNED_TERMS_DEFAULT = [
    "космет", "помада", "тушь", "крем", "духи", "парфюм",
    "грудь", "бюст", "лиф", "белье", "эрот",
    "кабель", "заряд", "адаптер", "блок питания", "провод",
    "стекло", "пленка", "защитн", "попсокет", "держател",
    "наушник", "гарнитур", "чехол для паспорта",
]
CASE_LIKE_TERMS_DEFAULT = [
    # core: phone cases
    "чехол", "case", "кейс",
    # common variants / stems
    "накладк", "книжк", "бампер",
    # wallet / pocket-ish wording
    "кошел", "wallet", "карман", "карт",
]

# Strict "case intent" terms used in relevance checks (avoid overly-broad stems like "карт")
CASE_INTENT_STRICT_TERMS = [
    "чехол", "чехл", "case", "кейс",
    "накладк", "книжк", "бампер",
]

# Pocket/cardholder intent aliases (for matching, not for query generation)
POCKET_INTENT_ALIASES = [
    # explicit
    "карман",
    "картхолдер", "карт холдер", "картхолд",
    "для карт", "держатель карт", "отделение для карт",
    "кошелек", "кошел", "wallet",
    "cardholder", "card holder",
    # looser stems (still gated by pocket_req + case intent)
    "карт", "карточ", "картой", "под карт", "слот", "отсек",
]

def _dedupe(xs: List[str]) -> List[str]:
    out = []
    seen = set()
    for x in xs:
        x = safe_str(x).strip()
        if not x:
            continue
        k = x.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(x)
    return out

def _extract_text_from_card(p: dict) -> Tuple[str, str]:
    name = safe_str(p.get("name") or p.get("goodsName") or "")
    descr = safe_str(p.get("description") or p.get("descr") or "")
    return name, descr

def _extract_brand_seller(p: dict) -> Tuple[str, str]:
    brand = safe_str(p.get("brand") or p.get("brandName") or "")
    seller = safe_str(p.get("supplier") or p.get("supplierName") or p.get("seller") or "")
    return brand, seller

def _extract_media(p: dict) -> dict:
    pics = safe_int(p.get("pics") or p.get("picsCount"), 0) or 0
    has_video = p.get("hasVideo")
    if has_video is None:
        has_video = bool(p.get("video")) if p.get("video") is not None else False
    return {"photos": pics, "video": 1 if has_video else 0}

def _extract_social(p: dict) -> dict:
    return {
        "rating": safe_float(p.get("rating") or p.get("reviewRating") or p.get("nmReviewRating"), None),
        "feedbacks": safe_int(p.get("feedbacks") or p.get("feedbackCount") or p.get("reviews") or p.get("nmFeedbacks"), None),
    }

def _extract_pricing(p: dict) -> dict:
    price_u = safe_int(p.get("priceU") or p.get("price"), None)
    sale_u = safe_int(p.get("salePriceU") or p.get("salePrice"), None)
    if (price_u is None or sale_u is None) and isinstance(p.get("sizes"), list) and p["sizes"]:
        s0 = p["sizes"][0] if isinstance(p["sizes"][0], dict) else {}
        pr = s0.get("price") if isinstance(s0, dict) else None
        if isinstance(pr, dict):
            price_u = price_u if price_u is not None else safe_int(pr.get("basic"), None)
            sale_u = sale_u if sale_u is not None else safe_int(pr.get("product"), None)
    def rub(u: Optional[int]) -> Optional[int]:
        if u is None:
            return None
        return int(u // 100) if u > 1000 else int(u)
    return {"price": rub(price_u), "sale_price": rub(sale_u)}

def _extract_options(p: dict) -> List[dict]:
    opts = p.get("options")
    if isinstance(opts, list):
        out = []
        for o in opts:
            if not isinstance(o, dict):
                continue
            out.append({
                "name": safe_str(o.get("name") or o.get("nm") or ""),
                "value": safe_str(o.get("value") or o.get("val") or ""),
            })
        return out
    return []

def _pick_from_options(options: List[dict], keys: List[str]) -> List[str]:
    keys_lc = [_pool_norm_lc(k) for k in keys]
    out: List[str] = []
    for o in options:
        n = _pool_norm_lc(o.get("name",""))
        v = safe_str(o.get("value","")).strip()
        if not v:
            continue
        for k in keys_lc:
            if k and k in n:
                out.append(v)
                break
    seen = set()
    res = []
    for x in out:
        xl = x.strip()
        if not xl or xl in seen:
            continue
        seen.add(xl)
        res.append(xl)
    return res


# --- phone model normalization (fixes ';' mega-strings from options) ---
MODEL_BRAND_TOKENS = {
    'apple','samsung','xiaomi','redmi','poco','realme','honor','huawei','oneplus','tecno','infinix','oppo','vivo'
}
MODEL_NET_TOKENS = {'4g','5g','lte','4glte','5glte'}

def normalize_phone_models(raw: List[str], limit: int = 12) -> List[str]:
    """Normalize phone model strings into a clean list of variants.

    Handles common listing separators (; | , /) and expands a few safe variants:
    - removes trailing network tokens (4g/5g/lte)
    - adds brand-less variant (e.g., 'camon 20 pro')
    - splits internal code packs (e.g., 'ck6 ck6n ck7n')

    Output is lowercase-ish (via _pool_norm_lc), deduped, order-preserving.
    """
    out: List[str] = []
    seen = set()

    def _push(vv: str) -> None:
        k = vv.lower()
        if k in seen:
            return
        seen.add(k)
        out.append(vv)

    def _add(v: str) -> None:
        v = safe_str(v).strip()
        if not v:
            return
        parts = [p.strip() for p in re.split(r'[;|,/\n]+', v) if safe_str(p).strip()]
        for p in parts:
            tlc = _pool_norm_lc(p)
            if not tlc:
                continue
            # collapse duplicated numbers like '20 20 pro' -> '20 pro'
            tlc = re.sub(r'\b(\d{1,3})\s+\1\b', r'\1', tlc)
            tlc = re.sub(r'\s+', ' ', tlc).strip()

            toks = tlc.split()
            codes = [t for t in toks if re.fullmatch(r'[a-z]{1,4}\d{1,3}[a-z]?$', t)]

            is_modelish = any(ch.isdigit() for ch in tlc) or ('iphone' in tlc) or ('айфон' in tlc)
            if is_modelish and len(tlc) >= 5:
                _push(tlc)
                if len(out) >= limit:
                    return

            # remove trailing net token
            if toks and toks[-1] in MODEL_NET_TOKENS:
                v2 = ' '.join(toks[:-1]).strip()
                if len(v2) >= 5 and (any(ch.isdigit() for ch in v2) or 'iphone' in v2 or 'айфон' in v2):
                    _push(v2)
                    if len(out) >= limit:
                        return

            # brand-less version
            for bt in MODEL_BRAND_TOKENS:
                if bt in toks:
                    v3 = ' '.join([t for t in toks if t != bt]).strip()
                    if len(v3) >= 5 and any(ch.isdigit() for ch in v3):
                        _push(v3)
                        if len(out) >= limit:
                            return

            # code tokens last
            if len(codes) >= 2:
                for c in codes:
                    if len(c) >= 3:
                        _push(c)
                        if len(out) >= limit:
                            return

    for s in raw:
        _add(s)
        if len(out) >= limit:
            break

    return out

PHONE_MODEL_PATTERNS = [
    r"(iphone\s?\d{1,2}\s?(pro\s?max|pro|max|mini)?)",
    r"(iphone\s?se\s?\d?)",
    r"(samsung\s?(galaxy\s?)?(s|a|m)\d{1,2}\s?(ultra|plus|\+|fe)?)",
    r"(xiaomi\s?(redmi\s?)?(note\s?)?\d{1,2}\s?(pro\s?\+?|pro|plus|\+)?)",
    r"(poco\s?[a-z]?\d{1,2}\s?(pro|max|\+)?)",
    r"(realme\s?\d{1,2}\s?(pro\s?\+?|pro|plus|\+)?)",
    r"(honor\s?\d{1,2}\s?(lite|pro|plus)?)",
    r"(huawei\s?(p|mate)\d{1,2}\s?(pro|plus)?)",
    r"(oneplus\s?\d{1,2}\s?(pro|r)?)",
    r"(tecno\s?(camon|pova|spark)\s?\d{1,2}\s?(pro\s?max|pro|plus|neo|premier|5g|4g)?)",
    r"(infinix\s?(hot|note|zero)\s?\d{1,2}\s?(pro|plus|5g|4g)?)",
    r"(oppo\s?(reno|a|find)\s?\d{1,2}\s?(pro|plus|5g|4g)?)",
    r"(vivo\s?(v|y|x)\s?\d{1,2}\s?(pro|plus|5g|4g)?)",
    r"(айфон\s?\d{1,2}\s?(про\s?макс|про|max|мини)?)",
    r"(самсунг\s?(галакси\s?)?(s|a|m)\d{1,2}\s?(ультра|плюс|фе)?)",
    r"(редми\s?(ноут\s?)?\d{1,2}\s?(про\s?плюс|про|плюс)?)",
]

def extract_phone_models(text: str, limit: int = 5) -> List[str]:
    t = _pool_norm_lc(text)
    # help regex catch things like '20pro' / '15pro' written without space
    t = re.sub(r"(\d)([a-zа-я])", r"\1 \2", t, flags=re.IGNORECASE)
    t = re.sub(r"([a-zа-я])(\d)", r"\1 \2", t, flags=re.IGNORECASE)
    t = re.sub(r"\s+", " ", t).strip()
    t = re.sub(r"\b(\d{1,3})\s+\1\b", r"\1", t)
    found: List[str] = []
    for pat in PHONE_MODEL_PATTERNS:
        for m in re.finditer(pat, t, flags=re.IGNORECASE):
            s = m.group(0)
            s = re.sub(r"\s+", " ", s).strip()
            if s and s not in found:
                found.append(s)
            if len(found) >= limit:
                return found
    return found

def extract_features(text: str) -> List[str]:
    t = _pool_norm_lc(text)
    feats = []

    # Pocket / cardholder intent (be specific; 'карт' substring is too noisy and kills rel50)
    pocket_hit = (
        ("карман" in t)
        or ("картхолд" in t) or ("карт холд" in t)
        or ("cardholder" in t) or ("card holder" in t)
        or ("wallet" in t) or ("кошелек" in t) or ("кошел" in t)
        or ("для карт" in t) or ("держатель карт" in t) or ("отделение для карт" in t)
        or bool(re.search(r"\bкарточ\w+", t))
        or bool(re.search(r"\bкарт(ы|а|у|е|ой|ами)?\b", t))
    )
    if pocket_hit:
        feats.append("карман для карт")

    if "magsafe" in t or "магнит" in t:
        feats.append("магнит")
    if "книжк" in t or "flip" in t or "book" in t:
        feats.append("книжка")
    if "противоудар" in t or "armor" in t or "shock" in t:
        feats.append("противоударный")
    return feats

def build_must_ban_terms(own_title: str, own_desc: str, models: List[str], feats: List[str]) -> Tuple[List[str], List[str]]:
    # "must_terms" are used as HARD gates in Stage D/E. If you make them too strict,
    # you get rel50=0 and the whole pipeline dies. So we bias for recall.
    must = ["чехол"]

    if models:
        # Prefer real model phrases over short internal codes (ck6, a52, etc.)
        # and prefer variants that actually appear in WB titles (often without the brand).
        code_pat = re.compile(r"^[a-z]{1,4}\d{1,3}[a-z]?$", re.IGNORECASE)

        def has_brand_token(s: str) -> bool:
            toks = _pool_norm_lc(s).split()
            return any(bt in toks for bt in MODEL_BRAND_TOKENS)

        cleaned: List[str] = []
        for mm in models:
            tlc = _pool_norm_lc(mm)
            if not tlc or len(tlc) < 5:
                continue
            if code_pat.fullmatch(tlc):
                continue
            # If the last token is a network token (4g/5g/lte), skip it because
            # normalize_phone_models already produces a net-less variant.
            toks = tlc.split()
            if toks and toks[-1] in MODEL_NET_TOKENS:
                continue
            cleaned.append(mm)

        if not cleaned:
            cleaned = models[:]

        # Pick 2-3 variants: brandless first (better recall), then brandful (better precision)
        brandless = [mm for mm in cleaned if not has_brand_token(mm)]
        brandful = [mm for mm in cleaned if has_brand_token(mm)]

        picked: List[str] = []
        if brandless:
            picked.append(brandless[0])
        if brandful:
            picked.append(brandful[0])
        # Add one more brandless if available (covers weird WB spellings)
        for mm in brandless[1:]:
            if len(picked) >= 3:
                break
            picked.append(mm)

        if not picked:
            picked = cleaned[:2]

        must.extend(picked[:3])

    # Pocket intent
    if any("карман" in f for f in feats):
        must.append("карман")

    ban = list(BANNED_TERMS_DEFAULT)
    if "стекл" in _pool_norm_lc(own_title):
        ban.append("стекло")
    return _dedupe(must), _dedupe(ban)


def serp_relevance_pass(p: dict, must_terms: List[str], ban_terms: List[str]) -> bool:
    """Hard relevance filter for SERP items (used in Stage D/E).
    Goal: keep same-intent products, but avoid false negatives from wording variance.

    NOTE: This function is the #1 reason you get rel50=0. Treat it like a chainsaw, not a scalpel.
    """
    name = safe_str(p.get("name") or p.get("goodsName") or p.get("title") or "")
    subj = safe_str(p.get("subjectName") or p.get("subjName") or "")
    blob_lc = _pool_norm_lc(" ".join([name, subj]))
    blob_ns = blob_lc.replace(" ", "")

    # ban terms
    for bt in ban_terms:
        if _term_in_blob(bt, blob_lc, blob_ns):
            return False

    if not must_terms:
        return True

    must_norm = [_pool_norm_lc(m) for m in must_terms if safe_str(m).strip()]

    # Case intent (do NOT require literal "чехол"; allow накладка/бампер/книжка)
    if any(("чехол" in m) or (m == "case") or ("кейс" in m) for m in must_norm):
        if not any(_term_in_blob(t, blob_lc, blob_ns) for t in CASE_INTENT_STRICT_TERMS):
            return False

    # Pocket intent: if must terms imply "pocket/cardholder", allow alias matches
    pocket_req = any(
        ("карман" in m) or ("картхолд" in m) or ("кошел" in m) or ("wallet" in m) or ("cardholder" in m) or ("длякарт" in _canon_nospace(m))
        for m in must_norm
    )
    pocket_hit = False
    if pocket_req:
        pocket_hit = any(_term_in_blob(a, blob_lc, blob_ns) for a in POCKET_INTENT_ALIASES)
        if not pocket_hit:
            return False

    # Model match: require at least one *real* model-ish term hit if present.
    # Ignore short internal codes (ck6, a52...) and tiny number tokens.
    code_pat = re.compile(r"^[a-z]{1,4}\d{1,3}[a-z]?$", re.IGNORECASE)

    model_terms: List[str] = []
    for t in must_terms:
        tlc = _pool_norm_lc(t)
        if not tlc or len(tlc) < 5:
            continue
        if code_pat.fullmatch(tlc):
            continue
        # Reject "model terms" that are basically just a number
        if re.fullmatch(r"\d{1,3}", tlc):
            continue
        if any(ch.isdigit() for ch in tlc) or any(x in tlc for x in ["iphone","айфон","galaxy","галакси","poco","redmi","tecno","infinix","honor","huawei","oneplus","oppo","vivo","realme","xiaomi","samsung"]):
            model_terms.append(t)

    if model_terms:
        if not any(_term_in_blob(t, blob_lc, blob_ns) for t in model_terms):
            return False

    # Count must hits, but treat "карман" as satisfied by any pocket alias hit
    hits = 0
    for t in must_terms:
        tlc = _pool_norm_lc(t)
        if not tlc:
            continue
        if pocket_req and (tlc == "карман" or "карман" in tlc):
            if pocket_hit:
                hits += 1
            continue
        if _term_in_blob(t, blob_lc, blob_ns):
            hits += 1

    # After explicit intent checks above, 1 hit is enough (prevents empty pools)
    return hits >= 1


# =========================
# Stage A: Input + Manifest
# =========================
# =========================

def read_input_scope(input_xlsx: Path, sheet: str, *, expect_count: Optional[int] = None, dedupe: bool = False) -> List[dict]:
    """
    Read SKU scope from an Excel sheet.

    Required column: nm_id (case-insensitive).
    Optional columns: vendor_code, name/title, potential_qty.

    By default, duplicates are treated as an error (safer).
    If dedupe=True, keeps the first occurrence of each nm_id and drops the rest.
    If expect_count is provided, enforces exact number of resulting SKU rows.
    """
    if not input_xlsx.exists():
        raise FileNotFoundError(f"Input file not found: {input_xlsx}")
    wb = load_workbook(input_xlsx, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet}. Available: {wb.sheetnames}")
    ws = wb[sheet]

    headers: List[str] = []
    for c in range(1, ws.max_column + 1):
        h = safe_str(ws.cell(row=1, column=c).value).strip()
        headers.append(h)

    def col_idx(name_candidates: List[str]) -> Optional[int]:
        for cand in name_candidates:
            if cand in headers:
                return headers.index(cand) + 1
        low = [h.lower() for h in headers]
        for cand in name_candidates:
            cl = cand.lower()
            if cl in low:
                return low.index(cl) + 1
        return None

    c_nm = col_idx(["nm_id", "nmId", "nm"])
    c_vc = col_idx(["vendor_code", "vendorCode", "vc"])
    c_name = col_idx(["name", "title"])
    c_qty = col_idx(["potential_qty", "potentialQty", "qty", "potential"])

    if c_nm is None:
        raise ValueError("Input sheet must contain an 'nm_id' column")

    items: List[dict] = []
    for r in range(2, ws.max_row + 1):
        nm = nm_id_to_str(ws.cell(row=r, column=c_nm).value)
        if not nm:
            continue
        vc = safe_str(ws.cell(row=r, column=c_vc).value).strip() if c_vc else ""
        nm_name = safe_str(ws.cell(row=r, column=c_name).value).strip() if c_name else ""
        qty = safe_int(ws.cell(row=r, column=c_qty).value, None) if c_qty else None
        items.append({"nm_id": nm, "vendor_code": vc, "name": nm_name, "potential_qty": qty})

    if not items:
        raise ValueError(f"No SKU rows found in sheet '{sheet}' (nm_id column is empty?)")

    # duplicates
    ids = [x["nm_id"] for x in items]
    if len(set(ids)) != len(ids):
        if not dedupe:
            raise ValueError("Duplicate nm_id in input sheet")
        seen = set()
        deduped: List[dict] = []
        dropped = 0
        for it in items:
            nm = it["nm_id"]
            if nm in seen:
                dropped += 1
                continue
            seen.add(nm)
            deduped.append(it)
        items = deduped
        print(f"[A] warning: deduped scope, dropped {dropped} duplicate rows", file=sys.stderr)

    if expect_count is not None and len(items) != int(expect_count):
        raise ValueError(f"Expected exactly {int(expect_count)} SKU rows in {sheet}. Got {len(items)}")

    return items

def stage_A(
    out_dir: Path,
    input_xlsx: Path,
    sheet: str,
    *,
    llm_provider: str,
    model_small: str,
    model_main: str,
    dests: List[int],
    search_limit: int,
    deep_card_enabled: bool,
    expect_count: Optional[int] = None,
    dedupe: bool = False,
    case_like_terms: Optional[List[str]] = None,
) -> None:
    ensure_dir(out_dir)

    ct = case_like_terms if isinstance(case_like_terms, list) else []
    ct = [safe_str(x).strip() for x in ct if safe_str(x).strip()]
    if not ct:
        ct = CASE_LIKE_TERMS_DEFAULT

    scope_list = read_input_scope(input_xlsx, sheet, expect_count=expect_count, dedupe=dedupe)

    manifest = {
        "schema_version": SCHEMA_VERSION,
        "script": {"name": SCRIPT_NAME, "version": SCRIPT_VERSION},
        "created_at": utc_now_iso(),
        "run_id": f"run_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}_{sha1_short(str(out_dir))}",
        "input": {
            "file": str(input_xlsx),
            "sheet": sheet,
            "expect_count": int(expect_count) if expect_count is not None else None,
            "dedupe": bool(dedupe),
        },
        "scope": {"sku_list": scope_list},
        "scope_count": len(scope_list),
        "wb": {"dests": dests, "search_limit": int(search_limit), "deep_card": {"enabled": bool(deep_card_enabled), "lang": "ru"}},
        "llm": {"provider": llm_provider, "model_small": model_small, "model_main": model_main},
        "filters": {
            "case_like_terms": ct,
            "ban_terms_default": BANNED_TERMS_DEFAULT,
        },
    }
    write_json(out_dir / "run_manifest.json", manifest)
def normalize_card_detail(nm_id: str, js: Optional[dict], deep_js: Optional[dict] = None) -> dict:
    """Normalize WB card payload into compact, model-friendly fields.

    Inputs:
      - js: card.wb.ru cards/*/detail payload (can be empty/None)
      - deep_js: basket-XX.wbbasket.ru .../info/ru/card.json (optional, richer)
    """
    js = js if isinstance(js, dict) else {}
    deep_js = deep_js if isinstance(deep_js, dict) else None

    # --- detail payload ---
    prods = _extract_products(js)
    p = prods[0] if prods and isinstance(prods[0], dict) else {}
    title_d, desc_d = _extract_text_from_card(p)
    brand_d, seller_d = _extract_brand_seller(p)
    media = _extract_media(p) if p else {"photos": 0, "video": 0}
    social = _extract_social(p) if p else {"rating": None, "feedbacks": None}
    pricing = _extract_pricing(p) if p else {"price": None, "sale_price": None}
    options_d = _extract_options(p) if p else []

    # --- deep payload (card.json) ---
    title_x = ""
    desc_x = ""
    options_x: List[dict] = []
    brand_x = ""
    seller_x = ""
    subject_x = ""
    subject_id_x: Optional[int] = None

    if deep_js:
        # title/desc
        title_x = safe_str(
            deep_js.get("imt_name") or deep_js.get("imtName") or deep_js.get("name") or deep_js.get("goods_name") or deep_js.get("goodsName") or ""
        ).strip()
        desc_x = safe_str(
            deep_js.get("description") or deep_js.get("descr") or deep_js.get("selling") or deep_js.get("shortDescription") or deep_js.get("short_description") or ""
        ).strip()

        # identity (best-effort)
        brand_x = safe_str(deep_js.get("brand") or deep_js.get("brandName") or deep_js.get("brand_name") or "").strip()
        seller_x = safe_str(deep_js.get("supplier") or deep_js.get("supplierName") or deep_js.get("seller") or "").strip()
        subject_x = safe_str(deep_js.get("subjectName") or deep_js.get("subjName") or deep_js.get("subject") or "").strip()
        subject_id_x = safe_int(deep_js.get("subjectId") or deep_js.get("subjectID") or deep_js.get("subject_id"), None)

        # options/characteristics
        # Most often: deep_js["options"] = [{name,value}, ...]
        options_x = _extract_options(deep_js)

        # fallback structures
        if not options_x:
            for k in ("properties", "chars", "characteristics"):
                v = deep_js.get(k)
                if isinstance(v, list):
                    tmp = []
                    for o in v:
                        if not isinstance(o, dict):
                            continue
                        tmp.append({
                            "name": safe_str(o.get("name") or o.get("nm") or o.get("key") or o.get("title") or ""),
                            "value": safe_str(o.get("value") or o.get("val") or o.get("text") or o.get("v") or ""),
                        })
                    options_x = tmp
                    if options_x:
                        break

    # --- merge ---
    title = title_d.strip() if safe_str(title_d).strip() else title_x
    # prefer the longer description (deep usually wins)
    desc_d = safe_str(desc_d).strip()
    desc = desc_x if len(desc_x) >= len(desc_d) else desc_d
    if desc:
        desc = desc[:1500]  # keep JSONL sane

    brand = brand_d.strip() if brand_d.strip() else brand_x
    seller = seller_d.strip() if seller_d.strip() else seller_x

    # merge options and dedupe
    options: List[dict] = []
    seen = set()
    for o in (options_d or []) + (options_x or []):
        if not isinstance(o, dict):
            continue
        n = safe_str(o.get("name","")).strip()
        v = safe_str(o.get("value","")).strip()
        if not n or not v:
            continue
        key = (n.lower(), v.lower())
        if key in seen:
            continue
        seen.add(key)
        # trim extreme values
        options.append({"name": n[:80], "value": v[:200]})
        if len(options) >= 120:
            break

    subject_name = safe_str(p.get("subjectName") or p.get("subjName") or "") if p else ""
    subject_id = safe_int(p.get("subjectId") or p.get("subject") or p.get("subjectID"), None) if p else None
    if not subject_name:
        subject_name = subject_x
    if subject_id is None:
        subject_id = subject_id_x

    identity = {"brand": brand, "seller": seller, "subject": subject_name, "subject_id": subject_id}
    content = {"title": title, "description": desc}
    attributes = {"attributes_raw": options}

    # derived fields from merged content/options
    models_opt = _pick_from_options(options, ["модель", "совместим", "телефон", "смартфон"])
    models_re = extract_phone_models(" ".join([title, desc or ""]), limit=8)
    # Prefer regex-found (usually clean) models, but salvage option strings too
    models = normalize_phone_models([*models_re, *models_opt], limit=12)

    feats = _pick_from_options(options, ["особенности", "функц", "тип", "карман", "карт"])
    feats2 = extract_features(" ".join([title, desc or ""]))
    feats = _dedupe([*feats, *feats2])

    mats = _pick_from_options(options, ["материал", "состав", "кожа", "силикон", "тпу", "пластик", "poly", "pu", "tpU"])

    attributes.update({
        "phone_models": models[:10],
        "materials": mats[:10],
        "features": feats[:20],
    })

    must, ban = build_must_ban_terms(title, desc or "", models, feats)
    extracted = {
        "must_terms_seed": must,
        "ban_terms_seed": ban,
        "content_source": {
            "detail_desc_len": len(desc_d),
            "deep_desc_len": len(desc_x),
            "deep_used": bool(deep_js),
        },
    }

    return {
        "identity": identity,
        "content": content,
        "attributes": attributes,
        "media": media,
        "social": social,
        "pricing": pricing,
        "extracted": extracted,
    }

def stage_B(out_dir: Path, *, timeout: int, sleep_sec: float, resume: bool, verbose: bool, deep_card_enabled: Optional[bool] = None) -> None:
    manifest = read_json(out_dir / "run_manifest.json")
    run_id = manifest["run_id"]
    scope = manifest["scope"]["sku_list"]
    total = len(scope)
    dests = manifest.get("wb", {}).get("dests", DEFAULT_DESTS)

    out_path = out_dir / "own_norm.jsonl"
    already = set()
    if resume and out_path.exists():
        for r in read_jsonl(out_path):
            already.add(nm_id_to_str(r.get("meta", {}).get("nm_id")))

    cache_dir = out_dir / ".wb_cache"
    ensure_dir(cache_dir)

    for i, sku in enumerate(scope, start=1):
        nm_id = nm_id_to_str(sku["nm_id"])
        if resume and nm_id in already:
            if verbose:
                print(f"[B] skip nm={nm_id}")
            continue

        vendor_code = safe_str(sku.get("vendor_code",""))
        name = safe_str(sku.get("name",""))

        status, code, js, url, dest_used = fetch_card_any(nm_id, dests, timeout=timeout, verbose=verbose)
        # deep (wbbasket) fetch: richer description + characteristics
        deep_cfg = manifest.get("wb", {}).get("deep_card", {}) if isinstance(manifest.get("wb", {}), dict) else {}
        deep_enabled = bool(deep_cfg.get("enabled", True)) if deep_card_enabled is None else bool(deep_card_enabled)
        deep_js = None
        deep_cache_rel = ""
        deep_status, deep_code, deep_url = "skip", 0, ""
        if deep_enabled:
            deep_status, deep_code, deep_js, deep_url, deep_host = fetch_deep_card_json(nm_id, timeout=timeout, verbose=verbose, lang=safe_str(deep_cfg.get("lang","ru")) or "ru")
            deep_cache_path = cache_dir / f"nm_{nm_id}" / "own_card_deep.json"
            ensure_dir(deep_cache_path.parent)
            if deep_status == "ok" and deep_js:
                write_json(deep_cache_path, {"nm_id": nm_id, "url": deep_url, "payload": deep_js})
                deep_cache_rel = str(deep_cache_path.relative_to(out_dir))
        cache_path = cache_dir / f"nm_{nm_id}" / "own_card.json"
        ensure_dir(cache_path.parent)
        if status == "ok" and js:
            write_json(cache_path, {"nm_id": nm_id, "dest": dest_used, "url": url, "payload": js})
            norm = normalize_card_detail(nm_id, js, deep_js)
            rec = {
                "meta": make_meta(run_id, "B", nm_id, vendor_code, name=name, source="wb"),
                "cache_path": str(cache_path.relative_to(out_dir)),
                "cache_path_deep": deep_cache_rel,
                "own_card": norm,
                "errors": [],
                "warnings": [],
            }
        else:
            rec = {
                "meta": make_meta(run_id, "B", nm_id, vendor_code, name=name, source="wb"),
                "cache_path": str(cache_path.relative_to(out_dir)),
                "cache_path_deep": deep_cache_rel,
                "own_card": {},
                "errors": [{"code": "WB_CARD_FAIL", "msg": f"{status}:{code}", "where": "stage_B"}],
                "warnings": [],
            }
        # Harmonize errors/warnings when deep payload salvages skinny/failed detail
        if deep_enabled:
            if deep_status != "ok":
                rec["warnings"].append({"code": "WB_DEEP_FAIL", "msg": f"{deep_status}:{deep_code}", "where": "stage_B"})
            if status != "ok" and deep_status == "ok":
                # Detail failed but deep worked: keep going (content is still usable for queries)
                rec["warnings"].append({"code": "WB_DETAIL_FAIL_BUT_DEEP_OK", "msg": f"{status}:{code}", "where": "stage_B"})
                # If we have normalized content, don't mark as hard error
                if rec.get("errors"):
                    rec["errors"] = []
        else:
            pass
        validate_meta(rec["meta"])
        append_jsonl(out_path, rec)
        if verbose:
            print(f"[B] ({i}/{total}) nm={nm_id} status={status} http={code}")
        time.sleep(max(0.0, float(sleep_sec)))

# =========================
# Stage C: Generate Queries -> queries_raw.jsonl
# =========================

def _rules_queries_from_own(own_norm: dict) -> List[str]:
    own = own_norm.get("own_card", {})
    content = own.get("content", {}) if isinstance(own.get("content"), dict) else {}
    attrs = own.get("attributes", {}) if isinstance(own.get("attributes"), dict) else {}
    title = safe_str(content.get("title") or own_norm.get("meta", {}).get("name") or "")
    desc = safe_str(content.get("description") or "")
    models_raw = attrs.get("phone_models") if isinstance(attrs.get("phone_models"), list) else []
    models_re = extract_phone_models(" ".join([title, desc]), limit=8)
    models = normalize_phone_models([*models_re, *models_raw], limit=12)
    feats = attrs.get("features") if isinstance(attrs.get("features"), list) else []
    blob = _pool_norm_lc(" ".join([title, desc, " ".join(map(str, feats))]))

    pocket = ("карман" in blob) or ("карт" in blob) or ("card" in blob)
    book = ("книжк" in blob) or ("book" in blob) or ("flip" in blob)

    qs: List[str] = []
    if models:
        m = models[0]
        if pocket:
            qs.append(f"чехол {m} с карманом для карт")
            qs.append(f"чехол {m} картхолдер")
            qs.append(f"{m} чехол кошелек")
            if book:
                qs.append(f"чехол книжка {m} с карманом")
        else:
            qs.append(f"чехол {m}")
            qs.append(f"{m} чехол")
            if book:
                qs.append(f"чехол книжка {m}")
    if pocket:
        qs.append("чехол с карманом для карт")
        qs.append("чехол картхолдер")
    qs.append("чехол для телефона")
    out = _dedupe(qs)
    return out[:8]

def validate_queries_raw_record(rec: dict) -> None:
    validate_meta(rec.get("meta", {}))
    qp = rec.get("query_pack")
    if not isinstance(qp, dict):
        raise ValueError("query_pack missing")
    qs = qp.get("queries")
    if not isinstance(qs, list) or not (3 <= len(qs) <= 10):
        raise ValueError("query_pack.queries must be list 3..10")
    for q in qs:
        if not isinstance(q, dict):
            raise ValueError("query entry not dict")
        if not safe_str(q.get("query_id")).startswith("q_"):
            raise ValueError("query_id invalid")
        if not safe_str(q.get("query_text")).strip():
            raise ValueError("query_text empty")
        if safe_str(q.get("source")).strip() == "":
            raise ValueError("query source empty")
    if not isinstance(qp.get("must_terms"), list) or not qp.get("must_terms"):
        raise ValueError("must_terms missing/empty")
    if not isinstance(qp.get("ban_terms"), list):
        raise ValueError("ban_terms missing")

def build_query_enrich_prompt(own_norm: dict, rule_queries: List[str]) -> List[dict]:
    own = own_norm.get("own_card", {})
    content = own.get("content", {}) if isinstance(own.get("content"), dict) else {}
    attrs = own.get("attributes", {}) if isinstance(own.get("attributes"), dict) else {}
    title = safe_str(content.get("title") or "")
    models_raw = attrs.get("phone_models") if isinstance(attrs.get("phone_models"), list) else []
    models_re = extract_phone_models(title, limit=8)
    models = normalize_phone_models([*models_re, *models_raw], limit=12)
    feats = attrs.get("features") if isinstance(attrs.get("features"), list) else []

    sys = (
        "You generate search queries for Wildberries.\n"
        "Return ONLY JSON object.\n"
        "Task: propose 2-4 additional queries (Russian) that match the same product intent.\n"
        "Rules:\n"
        "- Do not add brands.\n"
        "- Prefer concise queries.\n"
        "- Queries must be for the same phone model(s) if provided.\n"
        "Output schema: {\"queries\":[\"...\",\"...\"]}\n"
    )
    payload = {
        "title": title,
        "phone_models": models[:3],
        "features": feats[:10],
        "rule_queries": rule_queries[:8],
    }
    return [{"role":"system","content":sys},{"role":"user","content":json.dumps(payload, ensure_ascii=False)}]

def stage_C(
    out_dir: Path,
    *,
    provider: str,
    model: str,
    api_key: str,
    base_url: str = "",
    use_llm: bool,
    llm_timeout: int,
    max_tokens: int,
    temperature: float,
    resume: bool,
    verbose: bool,
) -> None:
    manifest = read_json(out_dir / "run_manifest.json")
    run_id = manifest["run_id"]
    scope = manifest["scope"]["sku_list"]
    total = len(scope)

    own_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "own_norm.jsonl")}

    out_path = out_dir / "queries_raw.jsonl"
    already = set()
    if resume and out_path.exists():
        for r in read_jsonl(out_path):
            already.add(nm_id_to_str(r.get("meta", {}).get("nm_id")))

    for i, sku in enumerate(scope, start=1):
        nm_id = nm_id_to_str(sku["nm_id"])
        if resume and nm_id in already:
            if verbose:
                print(f"[C] skip nm={nm_id}")
            continue

        vendor_code = safe_str(sku.get("vendor_code",""))
        name = safe_str(sku.get("name",""))
        own = own_map.get(nm_id)
        if not own:
            raise ValueError(f"missing own_norm for nm_id={nm_id}")

        rule_qs = _rules_queries_from_own(own)
        extra_qs: List[str] = []

        warnings: List[dict] = []
        errors: List[dict] = []
        llm_dbg = {"used": False, "provider": provider, "model": model, "http_code": None, "usage": {}}

        if use_llm:
            try:
                messages = build_query_enrich_prompt(own, rule_qs)
                parsed, dbg = call_llm_json(
                    provider=provider,
                    model=model,
                    api_key=api_key,
                    base_url=base_url,
                    messages=messages,
                    timeout_sec=llm_timeout,
                    max_tokens=max_tokens,
                    temperature=temperature,
                    force_json=True,
                )
                llm_dbg["used"] = True
                llm_dbg.update(dbg)
                qlist = parsed.get("queries", [])
                if not isinstance(qlist, list):
                    raise ValueError("LLM missing queries list")
                extra_qs = [safe_str(x).strip() for x in qlist if safe_str(x).strip()]
                extra_qs = _dedupe(extra_qs)[:4]
            except Exception as e:
                warnings.append({"code":"LLM_FAIL", "msg": str(e)[:200], "where":"stage_C"})
                extra_qs = []

        all_qs = _dedupe(rule_qs + extra_qs)
        if len(all_qs) < 3:
            all_qs = _dedupe(all_qs + ["чехол для телефона", "чехол", "чехол с карманом"])
        all_qs = all_qs[:10]

        # Recompute must/ban from current normalized content (do NOT trust historical seeds)
        own_card = own.get("own_card", {}) if isinstance(own.get("own_card"), dict) else {}
        content = own_card.get("content", {}) if isinstance(own_card.get("content"), dict) else {}
        attrs = own_card.get("attributes", {}) if isinstance(own_card.get("attributes"), dict) else {}
        title = safe_str(content.get("title") or own.get("meta", {}).get("name") or "")
        desc = safe_str(content.get("description") or "")
        feats = attrs.get("features") if isinstance(attrs.get("features"), list) else []
        models_raw = attrs.get("phone_models") if isinstance(attrs.get("phone_models"), list) else []
        models_re = extract_phone_models(" ".join([title, desc]), limit=8)
        models = normalize_phone_models([*models_re, *models_raw], limit=12)
        must_terms, ban_terms = build_must_ban_terms(title, desc, models, feats)

        queries = []
        for idxq, q in enumerate(all_qs, start=1):
            queries.append({"query_id": f"q_{idxq:02d}", "query_text": q, "source": "rules" if q in rule_qs else "llm"})

        rec = {
            "meta": make_meta(run_id, "C", nm_id, vendor_code, name=name, source="script"),
            "query_pack": {"queries": queries, "must_terms": must_terms, "ban_terms": ban_terms},
            "llm_debug": llm_dbg,
            "errors": errors,
            "warnings": warnings,
        }
        validate_queries_raw_record(rec)
        append_jsonl(out_path, rec)
        if verbose:
            print(f"[C] ({i}/{total}) nm={nm_id} queries={len(queries)} llm_used={llm_dbg['used']}")

# =========================
# Stage D: SERP validate queries -> queries_valid.jsonl + cache
# =========================

def validate_queries_valid_record(rec: dict) -> None:
    validate_meta(rec.get("meta", {}))
    vq = rec.get("valid_query_ids")
    if not isinstance(vq, list) or not vq:
        raise ValueError("valid_query_ids missing/empty")
    for qid in vq:
        if not safe_str(qid).startswith("q_"):
            raise ValueError("invalid query id in valid_query_ids")
    val = rec.get("validation")
    if not isinstance(val, dict):
        raise ValueError("validation missing")
    qs = val.get("queries")
    if not isinstance(qs, list) or not qs:
        raise ValueError("validation.queries missing/empty")
    for it in qs:
        if not isinstance(it, dict):
            raise ValueError("validation query item not dict")
        if not safe_str(it.get("query_id")).startswith("q_"):
            raise ValueError("validation query_id invalid")
        if safe_str(it.get("serp_cache_path")).strip() == "":
            raise ValueError("serp_cache_path missing")

def stage_D(out_dir: Path, *, timeout: int, sleep_sec: float, search_limit: int, resume: bool, verbose: bool) -> None:
    manifest = read_json(out_dir / "run_manifest.json")
    run_id = manifest["run_id"]
    scope = manifest["scope"]["sku_list"]
    scope_total = len(scope)
    dests = manifest.get("wb", {}).get("dests", DEFAULT_DESTS)
    search_limit = int(search_limit or manifest.get("wb", {}).get("search_limit", 100))

    qraw_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "queries_raw.jsonl")}

    out_path = out_dir / "queries_valid.jsonl"
    already = set()
    if resume and out_path.exists():
        for r in read_jsonl(out_path):
            already.add(nm_id_to_str(r.get("meta", {}).get("nm_id")))

    cache_dir = out_dir / ".wb_cache"
    ensure_dir(cache_dir)

    for i, sku in enumerate(scope, start=1):
        nm_id = nm_id_to_str(sku["nm_id"])
        if resume and nm_id in already:
            if verbose:
                print(f"[D] skip nm={nm_id}")
            continue

        vendor_code = safe_str(sku.get("vendor_code",""))
        name = safe_str(sku.get("name",""))

        qraw = qraw_map.get(nm_id)
        if not qraw:
            raise ValueError(f"missing queries_raw for nm_id={nm_id}")

        qp = qraw["query_pack"]
        must_terms = qp.get("must_terms", [])
        ban_terms = qp.get("ban_terms", [])
        queries = qp.get("queries", [])

        results = []
        for q in queries:
            qid = safe_str(q.get("query_id"))
            qtext = safe_str(q.get("query_text"))
            if not qid or not qtext:
                continue

            status, code, js, url, dest_used, rel_score = fetch_search_best(qtext, dests, timeout=timeout, limit=search_limit, verbose=False)

            cache_subdir = cache_dir / f"nm_{nm_id}"
            ensure_dir(cache_subdir)
            cache_name = f"serp_{sha1_short(qtext)}.json"
            cache_path = cache_subdir / cache_name

            items = []
            total = None
            if status == "ok" and js:
                items = parse_search_items(js)
                serp_total = safe_int(js.get("data", {}).get("total") if isinstance(js.get("data"), dict) else js.get("total"), None)
                write_json(cache_path, {"query": qtext, "url": url, "dest": dest_used, "total": serp_total, "products": items})
            else:
                write_json(cache_path, {"query": qtext, "url": url, "dest": dest_used, "total": serp_total, "products": []})

            top50 = items[:50]
            rel_cnt = sum(1 for p in top50 if serp_relevance_pass(p, must_terms, ban_terms))
            prices = [p for p in (_extract_search_price_rub(pp) for pp in items[:100]) if p]
            monster_share = _monster_share_topN(items, topN=20, threshold_feedbacks=1000)
            m = {
                "relevance_score": rel_score,
                "relevant_count_top50": int(rel_cnt),
                "price_p10": _percentile_int(prices, 10),
                "price_median": _percentile_int(prices, 50),
                "price_p90": _percentile_int(prices, 90),
                "total_estimate": serp_total,
                    "monster_share_top20": monster_share,
                }

            results.append({
                "query_id": qid,
                "query_text": qtext,
                "status": status,
                "http_code": code,
                "serp_cache_path": str(cache_path.relative_to(out_dir)),
                "metrics": m,
            })

            if verbose:
                print(f"[D] nm={nm_id} {qid} rel50={rel_cnt} score={rel_score} status={status}")
            time.sleep(max(0.0, float(sleep_sec)))

        scored = sorted(results, key=lambda x: (safe_int(x["metrics"].get("relevant_count_top50"),0) or 0,
                                               safe_int(x["metrics"].get("relevance_score"),0) or 0), reverse=True)
        # Choose valid queries: keep at most 5. We prefer strong signals but top-up
        # with medium/soft ones so Stage E can build a richer competitor pool.
        REL_HARD = 8   # strong signal
        REL_MED  = 3   # usable
        REL_SOFT = 1   # weak but >0
        MAX_KEEP = 5

        def _rel(r):
            return safe_int(r["metrics"].get("relevant_count_top50"), 0) or 0

        valid = [r for r in scored if _rel(r) >= REL_HARD]

        def _topup(thresh: int) -> None:
            for r in scored:
                if len(valid) >= MAX_KEEP:
                    break
                if r in valid:
                    continue
                if _rel(r) >= thresh:
                    valid.append(r)

        _topup(REL_MED)
        _topup(REL_SOFT)

        valid_ids = [r["query_id"] for r in valid[:MAX_KEEP]] if valid else []

        no_valid_queries = (len(valid_ids) == 0)

        rec = {
            "meta": make_meta(run_id, "D", nm_id, vendor_code, name=name, source="wb"),
            "valid_query_ids": valid_ids,
            "validation": {"queries": results},
            "errors": [],
            "warnings": ([{"code":"NO_VALID_QUERIES","msg":"No SERP query reached relevance threshold","where":"stage_D"}] if no_valid_queries else []),
        }
        validate_queries_valid_record(rec)
        append_jsonl(out_path, rec)
        if verbose:
            print(f"[D] ({i}/{scope_total}) nm={nm_id} valid={len(valid_ids)}")

# =========================
# Stage E: Build Competitor Pool -> competitor_pool.jsonl
# =========================

def validate_competitor_pool_record(rec: dict) -> None:
    validate_meta(rec.get("meta", {}))
    pool = rec.get("pool")
    if not isinstance(pool, dict):
        raise ValueError("pool missing")
    cands = pool.get("candidates")
    if not isinstance(cands, list):
        raise ValueError("pool.candidates missing")
    sel_ids = pool.get("selected_ids")
    if not isinstance(sel_ids, list):
        raise ValueError("pool.selected_ids missing")
    sel2 = [c.get("cand_nm_id") for c in cands if isinstance(c, dict) and c.get("selected")]
    if set(map(nm_id_to_str, sel_ids)) != set(map(nm_id_to_str, sel2)):
        raise ValueError("selected_ids mismatch selected flags")

def _hard_filter_candidate(
    c: dict,
    own_nm_id: str,
    must_terms: List[str],
    ban_terms: List[str],
    own_models: List[str],
    case_terms: Optional[List[str]] = None,
) -> Tuple[bool, List[str]]:
    reasons: List[str] = []
    cid = nm_id_to_str(c.get("cand_nm_id"))
    if cid == own_nm_id:
        reasons.append("is_owner")
        return False, reasons

    title = safe_str(c.get("title") or "")
    subj = safe_str(c.get("subject") or "")
    blob_lc = _pool_norm_lc(" ".join([title, subj]))
    blob_ns = blob_lc.replace(" ", "")

    # ban terms: hard reject
    for bt in ban_terms:
        if _term_in_blob(bt, blob_lc, blob_ns):
            reasons.append(f"ban:{bt}")
            return False, reasons

    # category-ish filter (project-specific by default, but configurable)
    ct = case_terms if isinstance(case_terms, list) and case_terms else CASE_LIKE_TERMS_DEFAULT
    ct = [safe_str(x).strip().lower() for x in ct if safe_str(x).strip()]
    if ct and not any(_term_in_blob(t, blob_lc, blob_ns) for t in ct):
        reasons.append("not_case_like")
        return False, reasons

    # model match (soft but important)
    if own_models:
        om = [m for m in own_models[:3] if safe_str(m).strip()]
        if om and not any(_term_in_blob(m, blob_lc, blob_ns) for m in om):
            reasons.append("no_model_match")
            return False, reasons

    # final relevance pass using must/ban terms
    if not serp_relevance_pass({"name": title, "subjectName": subj}, must_terms, ban_terms):
        reasons.append("serp_relevance_fail")
        return False, reasons

    return True, reasons
def stage_E(out_dir: Path, *, pool_limit: int = 30, resume: bool = False, verbose: bool = False) -> None:
    manifest = read_json(out_dir / "run_manifest.json")
    run_id = manifest["run_id"]
    scope = manifest["scope"]["sku_list"]
    total = len(scope)

    case_terms = manifest.get("filters", {}).get("case_like_terms") if isinstance(manifest.get("filters", {}), dict) else None
    if not isinstance(case_terms, list) or not case_terms:
        case_terms = CASE_LIKE_TERMS_DEFAULT

    own_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "own_norm.jsonl")}
    qraw_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "queries_raw.jsonl")}
    qv_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "queries_valid.jsonl")}

    out_path = out_dir / "competitor_pool.jsonl"
    already = set()
    if resume and out_path.exists():
        for r in read_jsonl(out_path):
            already.add(nm_id_to_str(r.get("meta", {}).get("nm_id")))

    for i, sku in enumerate(scope, start=1):
        nm_id = nm_id_to_str(sku["nm_id"])
        if resume and nm_id in already:
            if verbose:
                print(f"[E] skip nm={nm_id}")
            continue
        vendor_code = safe_str(sku.get("vendor_code",""))
        name = safe_str(sku.get("name",""))

        own = own_map.get(nm_id)
        qraw = qraw_map.get(nm_id)
        qv = qv_map.get(nm_id)
        if not own or not qraw or not qv:
            raise ValueError(f"missing deps for nm_id={nm_id} (need own+qraw+qv)")

        must_terms = qraw.get("query_pack", {}).get("must_terms", [])
        ban_terms = qraw.get("query_pack", {}).get("ban_terms", [])
        own_models = own.get("own_card", {}).get("attributes", {}).get("phone_models", [])
        if not isinstance(own_models, list):
            own_models = []

        candidates: Dict[str, dict] = {}
        best_query_id = None
        best_rel = -1

        val_qs = qv.get("validation", {}).get("queries", [])
        val_map = {safe_str(it.get("query_id")): it for it in val_qs if isinstance(it, dict)}
        valid_ids = qv.get("valid_query_ids", [])
        for qid in valid_ids:
            it = val_map.get(qid)
            if not it:
                continue
            cache_rel = safe_str(it.get("serp_cache_path"))
            pth = out_dir / cache_rel
            if not pth.exists():
                continue
            payload = read_json(pth)
            prods = payload.get("products", [])
            if not isinstance(prods, list):
                continue
            rel_cnt = sum(1 for p in prods[:50] if serp_relevance_pass(p, must_terms, ban_terms))
            if rel_cnt > best_rel:
                best_rel = rel_cnt
                best_query_id = qid

            for pos, p in enumerate(prods[:50], start=1):
                cid = nm_id_to_str(p.get("nm_id") or p.get("id"))
                if not cid or not cid.isdigit():
                    continue
                title = safe_str(p.get("name") or "")
                brand = safe_str(p.get("brand") or "")
                subj = safe_str(p.get("subjectName") or "")
                price = _extract_search_price_rub(p)
                rating = safe_float(p.get("rating"), None)
                feedbacks = safe_int(p.get("feedbacks"), None)
                prev = candidates.get(cid)
                if prev is None:
                    candidates[cid] = {
                        "cand_nm_id": cid,
                        "title": title,
                        "brand": brand,
                        "subject": subj,
                        "price": price,
                        "rating": rating,
                        "feedbacks": feedbacks,
                        "rank": pos,
                        "sources": [{"query_id": qid, "pos": pos}],
                    }
                else:
                    if pos < int(prev.get("rank") or 10**9):
                        prev["rank"] = pos
                    prev["sources"].append({"query_id": qid, "pos": pos})

        cand_list = sorted(list(candidates.values()), key=lambda c: (int(c.get("rank") or 10**9), -(safe_int(c.get("feedbacks"),0) or 0)))

        for c in cand_list:
            passed, reasons = _hard_filter_candidate(c, nm_id, must_terms, ban_terms, own_models, case_terms)
            c["hard_filter"] = {"passed": bool(passed), "reasons": reasons}
            c["selected"] = False

        passed = [c for c in cand_list if c.get("hard_filter", {}).get("passed")]
        selected = passed[:max(1, int(pool_limit))]
        selected_ids = [nm_id_to_str(c.get("cand_nm_id")) for c in selected]
        sel_set = set(selected_ids)
        for c in cand_list:
            if nm_id_to_str(c.get("cand_nm_id")) in sel_set:
                c["selected"] = True

        rec = {
            "meta": make_meta(run_id, "E", nm_id, vendor_code, name=name, source="script"),
            "pool": {
                "best_query_id": best_query_id or (valid_ids[0] if valid_ids else "q_01"),
                "candidates": cand_list,
                "selected_ids": selected_ids,
                "pool_limit": int(pool_limit),
            },
            "errors": [],
            "warnings": [] if selected_ids else [{"code":"EMPTY_POOL", "msg":"No candidates after filters", "where":"stage_E"}],
        }
        validate_competitor_pool_record(rec)
        append_jsonl(out_path, rec)
        if verbose:
            print(f"[E] ({i}/{total}) nm={nm_id} candidates={len(cand_list)} selected={len(selected_ids)}")

# =========================
# Stage F: Collect Competitor Cards -> competitor_norm.jsonl
# =========================

def validate_competitor_norm_record(rec: dict) -> None:
    validate_meta(rec.get("meta", {}))
    cid = nm_id_to_str(rec.get("competitor_nm_id"))
    if not re.fullmatch(r"\d+", cid):
        raise ValueError("competitor_nm_id invalid")
    if not isinstance(rec.get("competitor_card"), dict):
        raise ValueError("competitor_card missing or not dict")

def stage_F(out_dir: Path, *, timeout: int, sleep_sec: float, resume: bool, verbose: bool, deep_card_enabled: Optional[bool] = None) -> None:
    manifest = read_json(out_dir / "run_manifest.json")
    run_id = manifest["run_id"]
    scope = manifest["scope"]["sku_list"]
    total = len(scope)
    dests = manifest.get("wb", {}).get("dests", DEFAULT_DESTS)

    pool_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "competitor_pool.jsonl")}

    out_path = out_dir / "competitor_norm.jsonl"
    out_path.touch(exist_ok=True)
    already: set[Tuple[str,str]] = set()
    if resume and out_path.exists():
        for r in read_jsonl(out_path):
            owner = nm_id_to_str(r.get("meta", {}).get("nm_id"))
            cid = nm_id_to_str(r.get("competitor_nm_id"))
            if owner and cid:
                already.add((owner, cid))

    cache_dir = out_dir / ".wb_cache"
    ensure_dir(cache_dir)

    for i, sku in enumerate(scope, start=1):
        owner_nm = nm_id_to_str(sku["nm_id"])
        vendor_code = safe_str(sku.get("vendor_code",""))
        name = safe_str(sku.get("name",""))

        pool_rec = pool_map.get(owner_nm)
        if not pool_rec:
            raise ValueError(f"missing competitor_pool for owner={owner_nm}")

        sel_ids = pool_rec.get("pool", {}).get("selected_ids", [])
        if not isinstance(sel_ids, list):
            sel_ids = []

        for cid in sel_ids:
            cid = nm_id_to_str(cid)
            if resume and (owner_nm, cid) in already:
                continue

            status, code, js, url, dest_used = fetch_card_any(cid, dests, timeout=timeout, verbose=False)
            # deep (wbbasket) fetch for competitor
            deep_cfg = manifest.get("wb", {}).get("deep_card", {}) if isinstance(manifest.get("wb", {}), dict) else {}
            deep_enabled = bool(deep_cfg.get("enabled", True)) if deep_card_enabled is None else bool(deep_card_enabled)
            deep_js = None
            deep_cache_rel = ""
            deep_status, deep_code, deep_url = "skip", 0, ""
            if deep_enabled:
                deep_status, deep_code, deep_js, deep_url, deep_host = fetch_deep_card_json(cid, timeout=timeout, verbose=False, lang=safe_str(deep_cfg.get("lang","ru")) or "ru")
                deep_cache_path = cache_dir / f"nm_{owner_nm}" / f"comp_{cid}_deep.json"
                ensure_dir(deep_cache_path.parent)
                if deep_status == "ok" and deep_js:
                    write_json(deep_cache_path, {"owner_nm_id": owner_nm, "competitor_nm_id": cid, "url": deep_url, "payload": deep_js})
                    deep_cache_rel = str(deep_cache_path.relative_to(out_dir))
            cache_path = cache_dir / f"nm_{owner_nm}" / f"comp_{cid}.json"
            ensure_dir(cache_path.parent)
            if status == "ok" and js:
                write_json(cache_path, {"owner_nm_id": owner_nm, "competitor_nm_id": cid, "dest": dest_used, "url": url, "payload": js})
                norm = normalize_card_detail(cid, js, deep_js)
                rec = {
                    "meta": make_meta(run_id, "F", owner_nm, vendor_code, name=name, source="wb"),
                    "competitor_nm_id": cid,
                    "cache_path": str(cache_path.relative_to(out_dir)),
                    "cache_path_deep": deep_cache_rel,
                    "competitor_card": norm,
                    "errors": [],
                    "warnings": [],
                }
            else:
                rec = {
                    "meta": make_meta(run_id, "F", owner_nm, vendor_code, name=name, source="wb"),
                    "competitor_nm_id": cid,
                    "cache_path": str(cache_path.relative_to(out_dir)),
                    "cache_path_deep": deep_cache_rel,
                    "competitor_card": {},
                    "errors": [{"code":"WB_CARD_FAIL", "msg": f"{status}:{code}", "where":"stage_F"}],
                    "warnings": [],
                }
            validate_competitor_norm_record(rec)
            append_jsonl(out_path, rec)
            if verbose:
                print(f"[F] owner={owner_nm} comp={cid} status={status} http={code}")
            time.sleep(max(0.0, float(sleep_sec)))
        if verbose:
            print(f"[F] ({i}/{total}) owner={owner_nm} comps={len(sel_ids)}")

# --- chunk1 end ---


# =========================
# LLM Provider abstraction
# =========================

# --- LLM response parsing helpers ---

def _extract_chat_content_from_response(js: Any) -> Optional[str]:
    """Extract assistant message content from a Chat Completions-like response."""
    if js is None:
        return None

    # If someone passed already-parsed JSON (rare but useful for tests)
    if isinstance(js, dict) and "choices" not in js:
        return None

    if not isinstance(js, dict):
        return None

    choices = js.get("choices")
    if not isinstance(choices, list) or not choices:
        return None

    c0 = choices[0] if isinstance(choices[0], dict) else None
    if not isinstance(c0, dict):
        return None

    msg = c0.get("message")
    if isinstance(msg, dict):
        content = msg.get("content")
        if isinstance(content, str) and content.strip():
            return content

    # Fallbacks (some providers)
    text = c0.get("text")
    if isinstance(text, str) and text.strip():
        return text

    delta = c0.get("delta")
    if isinstance(delta, dict):
        content = delta.get("content")
        if isinstance(content, str) and content.strip():
            return content

    return None

def _strip_code_fences(s: str) -> str:
    s = safe_str(s, "").strip()
    if not s:
        return ""
    if "```" not in s:
        return s
    m = re.search(r"```(?:json)?\s*(.*?)\s*```", s, flags=re.IGNORECASE | re.DOTALL)
    if m:
        return safe_str(m.group(1), "").strip()
    return s

def extract_json_object(js: Any) -> Optional[dict]:
    """Return a dict parsed from LLM response JSON. Best-effort."""
    # If js already looks like parsed object (no choices), accept it
    if isinstance(js, dict) and "choices" not in js:
        return js

    content = _extract_chat_content_from_response(js)
    if not content:
        return None

    raw = _strip_code_fences(content)

    # Direct parse
    try:
        obj = json.loads(raw)
        if isinstance(obj, dict):
            return obj
        if isinstance(obj, list):
            return {"items": obj}
    except Exception:
        pass

    # Try to locate JSON object in the text
    s = raw.strip()
    # Prefer {...}
    i = s.find("{")
    j = s.rfind("}")
    if i != -1 and j != -1 and j > i:
        try:
            obj = json.loads(s[i:j+1])
            if isinstance(obj, dict):
                return obj
        except Exception:
            pass

    # Or [...]
    i = s.find("[")
    j = s.rfind("]")
    if i != -1 and j != -1 and j > i:
        try:
            obj = json.loads(s[i:j+1])
            if isinstance(obj, list):
                return {"items": obj}
        except Exception:
            pass

    return None

def call_llm_json(
    *,
    provider: str,
    model: str,
    api_key: str,
    messages: List[dict],
    base_url: str = "",
    timeout_sec: int = 60,
    max_tokens: int = 800,
    temperature: float = 0.2,
    force_json: bool = True,
) -> Tuple[dict, dict]:
    """
    Returns (parsed_json, debug_dict)

    Supported providers: openai | openrouter

    Notes:
    - Uses Chat Completions for both providers.
    - Tries to enforce JSON via response_format when supported.
    - If response_format is rejected, retries once without it.
    - Robustly extracts a JSON object even if the model adds junk around it.
    """
    provider = safe_str(provider).strip().lower()
    if provider not in {"openai", "openrouter"}:
        raise ValueError(f"Unsupported provider: {provider}")

    if not api_key:
        raise ValueError("Missing LLM API key")

    def _endpoint_for(provider: str, base_url: str) -> str:
        b = safe_str(base_url).strip()
        if not b:
            b = os.environ.get("LLM_BASE_URL", "").strip()
        if not b:
            b = os.environ.get("OPENAI_BASE_URL" if provider == "openai" else "OPENROUTER_BASE_URL", "").strip()
        if not b:
            b = "https://api.openai.com/v1" if provider == "openai" else "https://openrouter.ai/api/v1"

        b = b.rstrip("/")
        if b.endswith("/chat/completions") or b.endswith("/v1/chat/completions"):
            return b
        return b + "/chat/completions"

    url = _endpoint_for(provider, base_url)

    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    if provider == "openrouter":
        if os.environ.get("OPENROUTER_HTTP_REFERER"):
            headers["HTTP-Referer"] = os.environ["OPENROUTER_HTTP_REFERER"]
        if os.environ.get("OPENROUTER_X_TITLE"):
            headers["X-Title"] = os.environ["OPENROUTER_X_TITLE"]

    def _post(payload: Dict[str, Any]) -> Tuple[dict, dict]:
        t0 = time.time()
        r = LLM_SESSION.post(url, headers=headers, json=payload, timeout=timeout_sec)
        dt = time.time() - t0
        dbg = {
            "url": url,
            "http_code": r.status_code,
            "elapsed_sec": round(dt, 3),
            "usage": {},
            "retried_without_response_format": False,
            "retried_token_param": False,
        }
        txt_preview = safe_str(r.text)[:2000]
        dbg["raw_http_preview"] = txt_preview
        if r.status_code != 200:
            raise RuntimeError(f"LLM HTTP {r.status_code}: {txt_preview[:700]}")
        js = r.json()
        usage = js.get("usage") if isinstance(js, dict) else None
        if isinstance(usage, dict):
            dbg["usage"] = usage
        return js, dbg

    payload_base: Dict[str, Any] = {
        "model": model,
        "messages": messages,
        "temperature": float(temperature),
    }
    if force_json:
        payload_base["response_format"] = {"type": "json_object"}

    token_variants: List[Tuple[str, int]] = []
    if provider == "openai":
        token_variants = [("max_completion_tokens", int(max_tokens)), ("max_tokens", int(max_tokens))]
    else:
        token_variants = [("max_tokens", int(max_tokens))]

    last_err: Optional[Exception] = None
    dbg_final: dict = {}

    for tkey, tval in token_variants:
        payload = dict(payload_base)
        payload[tkey] = tval

        try:
            js, dbg = _post(payload)
            dbg_final = dbg
        except Exception as e:
            last_err = e
            msg = safe_str(e)
            if force_json and ("response_format" in msg or "json_object" in msg or "Unsupported" in msg or "400" in msg or "422" in msg):
                try:
                    payload2 = dict(payload)
                    payload2.pop("response_format", None)
                    js, dbg = _post(payload2)
                    dbg["retried_without_response_format"] = True
                    dbg_final = dbg
                except Exception as e2:
                    last_err = e2
                    continue
            else:
                continue

        obj = extract_json_object(js)
        if obj is not None:
            return obj, dbg_final

        last_err = RuntimeError("LLM returned no JSON object")
        continue

    raise RuntimeError(f"LLM request failed: {safe_str(last_err)}")

# =========================
# Stage G: LLM relevance classifier (optional) -> llm_relevance.jsonl
# =========================

def validate_llm_relevance_record(rec: dict) -> None:
    validate_meta(rec.get("meta", {}))
    if not isinstance(rec.get("relevance"), dict):
        raise ValueError("relevance missing")
    items = rec["relevance"].get("items")
    if not isinstance(items, list):
        raise ValueError("relevance.items missing")
    for it in items:
        if not isinstance(it, dict):
            raise ValueError("relevance item not dict")
        if not re.fullmatch(r"\d+", nm_id_to_str(it.get("cand_nm_id"))):
            raise ValueError("cand_nm_id invalid in relevance")
        if safe_str(it.get("label")) not in {"KEEP", "DROP"}:
            raise ValueError("label must be KEEP/DROP")
        if safe_str(it.get("reason")).strip() == "":
            raise ValueError("reason missing")

def build_relevance_prompt(owner_nm: str, owner_title: str, owner_models: List[str], owner_feats: List[str], candidates: List[dict]) -> List[dict]:
    sys = (
        "Ты строгий классификатор маркетплейса.\n"
        "Определи, является ли каждый кандидат настоящим конкурентом владельца SKU.\n"
        "Верни ТОЛЬКО JSON.\n"
        "Метки:\n"
        "- KEEP: тот же интент (например, чехол для той же линейки моделей, близкий тип).\n"
        "- DROP: другая категория или несовместимая модель.\n"
        "Требования:\n"
        "- НЕ переводить метки KEEP/DROP.\n"
        "- Поле reason пиши ТОЛЬКО на русском (кириллица), 5–20 слов.\n"
        "Схема: {\"items\":[{\"cand_nm_id\":\"...\",\"label\":\"KEEP|DROP\",\"reason\":\"...\"}]}\n"
    )
    payload = {
        "owner_nm_id": owner_nm,
        "owner_title": owner_title,
        "owner_models": owner_models[:3],
        "owner_features": owner_feats[:10],
        "candidates": [{"cand_nm_id": nm_id_to_str(c.get("cand_nm_id")),
                        "title": safe_str(c.get("title",""))[:180],
                        "subject": safe_str(c.get("subject",""))[:80]} for c in candidates[:25]],
    }
    return [{"role":"system","content":sys},{"role":"user","content":json.dumps(payload, ensure_ascii=False)}]

def stage_G(
    out_dir: Path,
    *,
    provider: str,
    model: str,
    api_key: str,
    base_url: str = "",
    use_llm: bool,
    llm_timeout: int,
    max_tokens: int,
    temperature: float,
    resume: bool,
    verbose: bool,
) -> None:
    manifest = read_json(out_dir / "run_manifest.json")
    run_id = manifest["run_id"]
    scope = manifest["scope"]["sku_list"]
    total = len(scope)

    own_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "own_norm.jsonl")}
    pool_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "competitor_pool.jsonl")}

    out_path = out_dir / "llm_relevance.jsonl"
    already = set()
    if resume and out_path.exists():
        for r in read_jsonl(out_path):
            already.add(nm_id_to_str(r.get("meta", {}).get("nm_id")))

    for i, sku in enumerate(scope, start=1):
        owner_nm = nm_id_to_str(sku["nm_id"])
        if resume and owner_nm in already:
            if verbose:
                print(f"[G] skip nm={owner_nm}")
            continue

        vendor_code = safe_str(sku.get("vendor_code",""))
        name = safe_str(sku.get("name",""))

        own = own_map.get(owner_nm)
        pool = pool_map.get(owner_nm)
        if not own or not pool:
            raise ValueError(f"missing deps for stage_G owner={owner_nm}")

        cand_list = pool.get("pool", {}).get("candidates", [])
        selected_ids = set(nm_id_to_str(x) for x in pool.get("pool", {}).get("selected_ids", []))
        selected = [c for c in cand_list if nm_id_to_str(c.get("cand_nm_id")) in selected_ids]

        owner_title = safe_str(own.get("own_card", {}).get("content", {}).get("title", ""))
        owner_models = own.get("own_card", {}).get("attributes", {}).get("phone_models", [])
        owner_feats = own.get("own_card", {}).get("attributes", {}).get("features", [])
        if not isinstance(owner_models, list):
            owner_models = []
        if not isinstance(owner_feats, list):
            owner_feats = []

        items = []
        warnings: List[dict] = []
        errors: List[dict] = []
        llm_dbg = {"used": False, "provider": provider, "model": model, "http_code": None, "usage": {}}

        use_llm_this = bool(use_llm)
        if use_llm_this:
            try:
                messages = build_relevance_prompt(owner_nm, owner_title, owner_models, owner_feats, selected)
                parsed, dbg = call_llm_json(
                    provider=provider,
                    model=model,
                    api_key=api_key,
                    base_url=base_url,
                    messages=messages,
                    timeout_sec=llm_timeout,
                    max_tokens=max_tokens,
                    temperature=temperature,
                    force_json=True,
                )
                llm_dbg["used"] = True
                llm_dbg.update(dbg)

                raw_items = parsed.get("items", [])
                if not isinstance(raw_items, list):
                    raise ValueError("LLM items missing")
                for it in raw_items:
                    if not isinstance(it, dict):
                        continue
                    cid = nm_id_to_str(it.get("cand_nm_id"))
                    label = safe_str(it.get("label")).strip().upper()
                    reason = safe_str(it.get("reason")).strip()
                    if cid and label in {"KEEP","DROP"} and reason:
                        items.append({"cand_nm_id": cid, "label": label, "reason": reason[:240]})
                # ensure all selected candidates present (fallback KEEP)
                present = set(nm_id_to_str(x.get("cand_nm_id")) for x in items)
                for c in selected:
                    cid = nm_id_to_str(c.get("cand_nm_id"))
                    if cid and cid not in present:
                        items.append({"cand_nm_id": cid, "label": "KEEP", "reason": "fallback"})
            except Exception as e:
                warnings.append({"code":"LLM_FAIL", "msg": str(e)[:200], "where":"stage_G"})
                use_llm_this = False

        if not use_llm_this:
            # deterministic fallback: KEEP all selected (we already hard-filtered)
            for c in selected:
                cid = nm_id_to_str(c.get("cand_nm_id"))
                if cid:
                    items.append({"cand_nm_id": cid, "label": "KEEP", "reason": "rules_fallback"})

        rec = {
            "meta": make_meta(run_id, "G", owner_nm, vendor_code, name=name, source="llm" if llm_dbg["used"] else "script"),
            "relevance": {"items": items},
            "llm_debug": llm_dbg,
            "errors": errors,
            "warnings": warnings,
        }
        validate_llm_relevance_record(rec)
        append_jsonl(out_path, rec)
        if verbose:
            kept = sum(1 for it in items if it.get("label") == "KEEP")
            print(f"[G] ({i}/{total}) owner={owner_nm} kept={kept}/{len(items)} llm_used={llm_dbg['used']}")

# =========================
# Stage H: Select final competitors -> competitors_selected.jsonl
# =========================

def validate_competitors_selected_record(rec: dict) -> None:
    validate_meta(rec.get("meta", {}))
    sel = rec.get("selection")
    if not isinstance(sel, dict):
        raise ValueError("selection missing")
    ids = sel.get("selected_ids")
    if not isinstance(ids, list):
        raise ValueError("selected_ids must be list (can be empty)")
    for x in ids:
        if not re.fullmatch(r"\d+", nm_id_to_str(x)):
            raise ValueError("invalid selected id")
    src = safe_str(sel.get("source"))
    if src not in {"llm", "rules"}:
        raise ValueError("selection.source invalid")
    roles = sel.get("roles")
    if roles is not None:
        if not isinstance(roles, dict):
            raise ValueError("selection.roles must be dict")
        for k, v in roles.items():
            if safe_str(k).strip() == "":
                raise ValueError("role key empty")
            if v is None:
                continue
            if not re.fullmatch(r"\d+", nm_id_to_str(v)):
                raise ValueError("role value must be nm_id")

def stage_H(out_dir: Path, *, max_competitors: int = 12, resume: bool = False, verbose: bool = False) -> None:
    """
    Select final competitors for each owner SKU.

    V1.1 improvement:
    - allows empty selection (pipeline continues with warnings)
    - selects role-based anchors when possible:
        LEADER, PRICE_FLOOR, PRICE_CEILING, CONTENT_BEAST, CLOSEST_MATCH
      then fills remaining slots by best rank.
    """
    manifest = read_json(out_dir / "run_manifest.json")
    run_id = manifest["run_id"]
    scope = manifest["scope"]["sku_list"]
    total = len(scope)

    own_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "own_norm.jsonl")}
    pool_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "competitor_pool.jsonl")}
    rel_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "llm_relevance.jsonl")}
    comp_norm = read_jsonl(out_dir / "competitor_norm.jsonl")

    comp_by_owner: Dict[str, Dict[str, dict]] = {}
    for r in comp_norm:
        owner = nm_id_to_str(r.get("meta", {}).get("nm_id"))
        cid = nm_id_to_str(r.get("competitor_nm_id"))
        if owner and cid:
            comp_by_owner.setdefault(owner, {})[cid] = r

    out_path = out_dir / "competitors_selected.jsonl"
    already = set()
    if resume and out_path.exists():
        for r in read_jsonl(out_path):
            already.add(nm_id_to_str(r.get("meta", {}).get("nm_id")))

    def _media_score(owner_nm: str, cid: str) -> int:
        r = comp_by_owner.get(owner_nm, {}).get(cid)
        if not r:
            return 0
        cc = r.get("competitor_card", {})
        if not isinstance(cc, dict):
            return 0
        media = cc.get("media", {}) if isinstance(cc.get("media"), dict) else {}
        ph = safe_int(media.get("photos"), 0) or 0
        vd = safe_int(media.get("video"), 0) or 0
        return int(ph + (5 * vd))

    def _closest_score(owner_models: List[str], owner_feats: List[str], cand: dict) -> int:
        title = safe_str(cand.get("title",""))
        subj = safe_str(cand.get("subject",""))
        blob_lc = _pool_norm_lc(title + " " + subj)
        blob_ns = blob_lc.replace(" ", "")
        s = 0
        for m in owner_models[:3]:
            if _term_in_blob(m, blob_lc, blob_ns):
                s += 3
        feats_blob = blob_lc
        if any("карман" in _pool_norm_lc(f) for f in owner_feats):
            if "карман" in feats_blob or "card" in feats_blob:
                s += 1
        if any("magsafe" in _pool_norm_lc(f) or "магнит" in _pool_norm_lc(f) for f in owner_feats):
            if "magsafe" in feats_blob or "магнит" in feats_blob:
                s += 1
        if any("книжк" in _pool_norm_lc(f) for f in owner_feats):
            if "книжк" in feats_blob or "flip" in feats_blob or "book" in feats_blob:
                s += 1
        return s

    for i, sku in enumerate(scope, start=1):
        owner_nm = nm_id_to_str(sku["nm_id"])
        if resume and owner_nm in already:
            if verbose:
                print(f"[H] skip nm={owner_nm}")
            continue

        vendor_code = safe_str(sku.get("vendor_code",""))
        name = safe_str(sku.get("name",""))

        pool = pool_map.get(owner_nm)
        own = own_map.get(owner_nm)
        if not pool or not own:
            raise ValueError(f"missing deps for stage_H owner={owner_nm}")

        base_ids = [nm_id_to_str(x) for x in pool.get("pool", {}).get("selected_ids", []) if nm_id_to_str(x)]
        base_set = set(base_ids)

        # apply LLM relevance KEEP if available
        src = "rules"
        keep_ids = list(base_ids)
        if owner_nm in rel_map:
            rel = rel_map[owner_nm]
            items = rel.get("relevance", {}).get("items", [])
            keep = [nm_id_to_str(it.get("cand_nm_id")) for it in items if isinstance(it, dict) and safe_str(it.get("label")) == "KEEP"]
            keep = [x for x in keep if x in base_set]
            if keep:
                src = "llm"
                keep_ids = keep

        # build candidate lookup from pool candidates
        cand_list = pool.get("pool", {}).get("candidates", [])
        cand_by_id: Dict[str, dict] = {nm_id_to_str(c.get("cand_nm_id")): c for c in cand_list if isinstance(c, dict)}
        kept_cands = [cand_by_id[cid] for cid in keep_ids if cid in cand_by_id]

        # owner context for CLOSEST_MATCH
        owner_models = own.get("own_card", {}).get("attributes", {}).get("phone_models", [])
        owner_feats = own.get("own_card", {}).get("attributes", {}).get("features", [])
        if not isinstance(owner_models, list):
            owner_models = []
        if not isinstance(owner_feats, list):
            owner_feats = []

        # role picks
        picked: List[str] = []
        roles: Dict[str, str] = {}

        def _pick(role: str, candidates: List[dict], key_fn) -> None:
            nonlocal picked, roles
            avail = [c for c in candidates if nm_id_to_str(c.get("cand_nm_id")) and nm_id_to_str(c.get("cand_nm_id")) not in set(picked)]
            if not avail:
                return
            best = sorted(avail, key=key_fn, reverse=True)[0]
            cid = nm_id_to_str(best.get("cand_nm_id"))
            if cid:
                picked.append(cid)
                roles[role] = cid

        # LEADER: feedbacks, then rating, then rank inverse
        _pick("LEADER", kept_cands, lambda c: (
            safe_int(c.get("feedbacks"), -1) or -1,
            safe_float(c.get("rating"), -1.0) or -1.0,
            -int(c.get("rank") or 10**9),
        ))

        # PRICE_FLOOR: invert sort by using negative price for reverse sort
        _pick("PRICE_FLOOR", kept_cands, lambda c: (
            -(safe_int(c.get("price"), 10**9) or 10**9),
            -int(c.get("rank") or 10**9),
        ))

        # PRICE_CEILING
        _pick("PRICE_CEILING", kept_cands, lambda c: (
            safe_int(c.get("price"), -1) or -1,
            -(int(c.get("rank") or 10**9)),
        ))

        # CONTENT_BEAST (needs competitor_norm ideally)
        _pick("CONTENT_BEAST", kept_cands, lambda c: (
            _media_score(owner_nm, nm_id_to_str(c.get("cand_nm_id"))),
            safe_int(c.get("feedbacks"), -1) or -1,
        ))

        # CLOSEST_MATCH (owner model/features overlap)
        _pick("CLOSEST_MATCH", kept_cands, lambda c: (
            _closest_score(owner_models, owner_feats, c),
            safe_int(c.get("feedbacks"), -1) or -1,
            -int(c.get("rank") or 10**9),
        ))

        # fill remaining by best rank
        rank_map = {nm_id_to_str(c.get("cand_nm_id")): int(c.get("rank") or 10**9) for c in cand_list if isinstance(c, dict)}
        remaining = [cid for cid in keep_ids if cid not in set(picked)]
        remaining = sorted(remaining, key=lambda x: rank_map.get(x, 10**9))

        cap = int(max_competitors)
        if cap < 0:
            cap = 0

        selected_ids = picked + remaining
        if cap > 0:
            selected_ids = selected_ids[:cap]
        else:
            selected_ids = []

        warnings = []
        if not selected_ids:
            warnings.append({"code":"EMPTY_SELECTION", "msg":"No selected competitors", "where":"stage_H"})
        if not roles and keep_ids:
            warnings.append({"code":"ROLE_PICK_FAIL", "msg":"Could not assign any role picks (fallback to rank)", "where":"stage_H"})

        rec = {
            "meta": make_meta(run_id, "H", owner_nm, vendor_code, name=name, source="script"),
            "selection": {
                "selected_ids": selected_ids,
                "max_competitors": int(max_competitors),
                "source": src,
                "roles": roles,
            },
            "errors": [],
            "warnings": warnings,
        }
        validate_competitors_selected_record(rec)
        append_jsonl(out_path, rec)
        if verbose:
            print(f"[H] ({i}/{total}) owner={owner_nm} selected={len(selected_ids)} src={src} roles={len(roles)}")

# --- chunk2 end ---


# =========================
# Stage I: Compute comparison metrics -> comparison_metrics.jsonl
# =========================

def validate_comparison_metrics_record(rec: dict) -> None:
    validate_meta(rec.get("meta", {}))
    m = rec.get("metrics")
    if not isinstance(m, dict):
        raise ValueError("metrics missing")
    if not isinstance(rec.get("competitors"), list):
        raise ValueError("competitors list missing")

def _price_bucket(pr: Optional[int]) -> str:
    if pr is None:
        return "unknown"
    if pr < 700:
        return "<700"
    if pr < 1000:
        return "700-999"
    if pr < 1300:
        return "1000-1299"
    if pr < 1700:
        return "1300-1699"
    if pr < 2200:
        return "1700-2199"
    return "2200+"

def _score_vs_competitors(own: dict, comps: List[dict]) -> dict:
    own_price = safe_int(own.get("pricing", {}).get("sale_price") or own.get("pricing", {}).get("price"), None)
    own_rating = safe_float(own.get("social", {}).get("rating"), None)
    own_fb = safe_int(own.get("social", {}).get("feedbacks"), None)
    own_media = own.get("media", {}) if isinstance(own.get("media"), dict) else {}
    own_ph = safe_int(own_media.get("photos"), 0) or 0
    own_vid = safe_int(own_media.get("video"), 0) or 0

    comp_prices = [safe_int(c.get("pricing", {}).get("sale_price") or c.get("pricing", {}).get("price"), None) for c in comps]
    comp_prices = [p for p in comp_prices if p is not None]
    comp_rating = [safe_float(c.get("social", {}).get("rating"), None) for c in comps]
    comp_rating = [r for r in comp_rating if r is not None]
    comp_fb = [safe_int(c.get("social", {}).get("feedbacks"), None) for c in comps]
    comp_fb = [f for f in comp_fb if f is not None]

    comp_ph = [safe_int((c.get("media") or {}).get("photos"), None) for c in comps]
    comp_ph = [p for p in comp_ph if p is not None]
    comp_vid = [safe_int((c.get("media") or {}).get("video"), None) for c in comps]
    comp_vid = [v for v in comp_vid if v is not None]

    med_price = _percentile_int(comp_prices, 50)
    p10 = _percentile_int(comp_prices, 10)
    p90 = _percentile_int(comp_prices, 90)

    med_rating = None
    if comp_rating:
        xs = sorted(comp_rating)
        med_rating = xs[len(xs)//2] if len(xs)%2==1 else round((xs[len(xs)//2-1]+xs[len(xs)//2])/2, 2)

    med_fb = _median_int(comp_fb) if comp_fb else None
    med_ph = _median_int(comp_ph) if comp_ph else None
    med_vid = _median_int(comp_vid) if comp_vid else None

    def pct_diff(a: Optional[int], b: Optional[int]) -> Optional[float]:
        if a is None or b is None or b == 0:
            return None
        return round((a - b) / b * 100.0, 1)

    return {
        "own_price": own_price,
        "own_price_bucket": _price_bucket(own_price),
        "comp_price_p10": p10,
        "comp_price_median": med_price,
        "comp_price_p90": p90,
        "own_vs_comp_median_price_pct": pct_diff(own_price, med_price),
        "own_rating": own_rating,
        "comp_rating_median": med_rating,
        "own_feedbacks": own_fb,
        "comp_feedbacks_median": med_fb,
        "own_photos": own_ph,
        "comp_photos_median": med_ph,
        "own_video": own_vid,
        "comp_video_median": med_vid,
        "content_gap_photos": (max(0, int(med_ph) - int(own_ph)) if (med_ph is not None and own_ph is not None) else None),
        "content_gap_video": (max(0, int(med_vid) - int(own_vid)) if (med_vid is not None and own_vid is not None) else None),
    }

def stage_I(out_dir: Path, *, resume: bool = False, verbose: bool = False) -> None:
    """
    Compute comparison metrics per SKU.

    V1.4 improvements:
    - keeps SERP-derived market snapshot from Stage D (best valid query)
    - adds keyword_gap using must_terms / ban_terms and competitor/title/query term coverage
    - keeps existing competitor summary + own-vs-comp medians + feature/material gaps
    """
    manifest = read_json(out_dir / "run_manifest.json")
    run_id = manifest["run_id"]
    scope = manifest["scope"]["sku_list"]
    total = len(scope)

    own_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "own_norm.jsonl")}
    sel_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "competitors_selected.jsonl")}
    comp_norm = read_jsonl(out_dir / "competitor_norm.jsonl")
    qv_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "queries_valid.jsonl")}
    qr_map: Dict[str, dict] = {}
    qr_path = out_dir / "queries_raw.jsonl"
    if qr_path.exists():
        qr_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(qr_path)}

    comp_by_owner: Dict[str, Dict[str, dict]] = {}
    for r in comp_norm:
        owner = nm_id_to_str(r.get("meta", {}).get("nm_id"))
        cid = nm_id_to_str(r.get("competitor_nm_id"))
        if not owner or not cid:
            continue
        comp_by_owner.setdefault(owner, {})[cid] = r

    out_path = out_dir / "comparison_metrics.jsonl"
    already = set()
    if resume and out_path.exists():
        for r in read_jsonl(out_path):
            already.add(nm_id_to_str(r.get("meta", {}).get("nm_id")))

    RU_STOP = {
        "и","в","во","на","для","с","со","по","к","ко","от","до","из","у","о","об","а","но","или","ли","же","что","это","этот","эта","эти",
        "как","так","не","нет","да","при","без","под","над","за","про","через","уже","еще","ещё","все","всё","сам","сама","свой","своя","свои",
        "the","and","or","for","with","without","to","of","in","on","a","an"
    }

    def _norm_text(s: str) -> str:
        s = safe_str(s).lower().strip()
        if not s:
            return ""
        s = re.sub(r"[^0-9a-zа-яё+#\-\. ]+", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def _tokenize(s: str) -> List[str]:
        s = _norm_text(s)
        if not s:
            return []
        return [t for t in s.split(" ") if t and len(t) >= 3 and t not in RU_STOP]

    def _contains_phrase(text_norm: str, phrase: str) -> bool:
        p = _norm_text(phrase)
        return bool(p) and p in text_norm

    def _top(counter: Counter, limit: int, min_count: int) -> List[dict]:
        out = []
        for term, cnt in counter.most_common():
            if cnt < min_count:
                continue
            out.append({"term": term, "count": int(cnt)})
            if len(out) >= limit:
                break
        return out

    def _top_missing_terms(own_terms: List[str], comp_terms: List[str], *, limit: int = 10, min_count: int = 2) -> List[dict]:
        own_set = set(_norm_text(x) for x in own_terms if _norm_text(x))
        c = Counter(_norm_text(x) for x in comp_terms if _norm_text(x))
        out = []
        for term, cnt in c.most_common():
            if term in own_set:
                continue
            if cnt < min_count:
                continue
            out.append({"term": term, "count": int(cnt)})
            if len(out) >= limit:
                break
        return out

    def _compute_feature_material_gaps(own_card: dict, comps_cards: List[dict]) -> dict:
        own_attrs = own_card.get("attributes", {}) if isinstance(own_card.get("attributes"), dict) else {}
        own_feats = own_attrs.get("features", []) if isinstance(own_attrs.get("features"), list) else []
        own_mats = own_attrs.get("materials", []) if isinstance(own_attrs.get("materials"), list) else []

        comp_feats_all: List[str] = []
        comp_mats_all: List[str] = []
        for cc in comps_cards:
            if not isinstance(cc, dict):
                continue
            a = cc.get("attributes", {}) if isinstance(cc.get("attributes"), dict) else {}
            cf = a.get("features", []) if isinstance(a.get("features"), list) else []
            cm = a.get("materials", []) if isinstance(a.get("materials"), list) else []
            comp_feats_all.extend([safe_str(x) for x in cf if safe_str(x).strip()])
            comp_mats_all.extend([safe_str(x) for x in cm if safe_str(x).strip()])

        missing_features = _top_missing_terms([safe_str(x) for x in own_feats], comp_feats_all, limit=10, min_count=2)
        missing_materials = _top_missing_terms([safe_str(x) for x in own_mats], comp_mats_all, limit=6, min_count=2)
        return {"missing_features": missing_features, "missing_materials": missing_materials}

    def _compute_keyword_gap(owner: str, own_card: dict, comps_cards: List[dict]) -> dict:
        qp = (qr_map.get(owner) or {}).get("query_pack") if isinstance(qr_map.get(owner), dict) else {}
        must_terms = qp.get("must_terms", []) if isinstance(qp, dict) and isinstance(qp.get("must_terms"), list) else []
        ban_terms = qp.get("ban_terms", []) if isinstance(qp, dict) and isinstance(qp.get("ban_terms"), list) else []
        must_terms = [safe_str(x).strip() for x in must_terms if safe_str(x).strip()]
        ban_terms = [safe_str(x).strip() for x in ban_terms if safe_str(x).strip()]

        title = safe_str((own_card.get("content") or {}).get("title", ""))
        desc = safe_str((own_card.get("content") or {}).get("description", ""))
        attrs = own_card.get("attributes", {}) if isinstance(own_card.get("attributes"), dict) else {}

        extra_parts: List[str] = []
        for k in ["brand", "color", "model", "type", "subject"]:
            v = attrs.get(k)
            if isinstance(v, str) and v.strip():
                extra_parts.append(v)
        for k in ["features", "materials", "phone_models"]:
            v = attrs.get(k)
            if isinstance(v, list):
                extra_parts.extend([safe_str(x) for x in v if safe_str(x).strip()])

        own_norm = _norm_text(" ".join([title, desc] + extra_parts))
        own_tok = set(_tokenize(own_norm))

        must_missing = [t for t in must_terms if not _contains_phrase(own_norm, t)]
        ban_found = [t for t in ban_terms if _contains_phrase(own_norm, t)]

        c_title = Counter()
        c_attr = Counter()
        for cc in comps_cards:
            if not isinstance(cc, dict):
                continue
            ct = safe_str((cc.get("content") or {}).get("title", ""))
            toks = set(_tokenize(ct))
            for t in toks:
                if t not in own_tok:
                    c_title[t] += 1

            a = cc.get("attributes", {}) if isinstance(cc.get("attributes"), dict) else {}
            for k in ["features", "materials", "phone_models"]:
                v = a.get(k)
                if isinstance(v, list):
                    vtoks = set()
                    for x in v:
                        vtoks.update(_tokenize(safe_str(x)))
                    for t in vtoks:
                        if t not in own_tok:
                            c_attr[t] += 1

        missing_title_terms = _top(c_title, limit=12, min_count=2)
        missing_attr_terms = _top(c_attr, limit=10, min_count=2)

        qv = qv_map.get(owner) or {}
        valid_ids = qv.get("valid_query_ids", []) if isinstance(qv.get("valid_query_ids"), list) else []
        rows = (qv.get("validation") or {}).get("queries", []) if isinstance((qv.get("validation") or {}), dict) else []
        q_terms = Counter()
        if isinstance(rows, list) and valid_ids:
            valid_set = set(safe_str(x) for x in valid_ids)
            for r in rows:
                if not isinstance(r, dict):
                    continue
                if safe_str(r.get("query_id")) not in valid_set:
                    continue
                qt = safe_str(r.get("query_text", ""))
                toks = set(_tokenize(qt))
                for t in toks:
                    if t not in own_tok:
                        q_terms[t] += 1
        missing_query_terms = _top(q_terms, limit=10, min_count=2)

        return {
            "must_terms_missing_in_own": must_missing,
            "ban_terms_found_in_own": ban_found,
            "missing_title_terms": missing_title_terms,
            "missing_attr_terms": missing_attr_terms,
            "missing_query_terms": missing_query_terms,
        }

    def _pick_best_serp(owner: str) -> Optional[dict]:
        qv = qv_map.get(owner)
        if not qv:
            return None
        valid_ids = qv.get("valid_query_ids", [])
        rows = (qv.get("validation") or {}).get("queries", [])
        if not isinstance(valid_ids, list) or not isinstance(rows, list):
            return None
        valid_set = set(map(safe_str, valid_ids))
        rows = [r for r in rows if isinstance(r, dict) and safe_str(r.get("query_id")) in valid_set]
        if not rows:
            return None
        rows_sorted = sorted(rows, key=lambda x: (
            safe_int((x.get("metrics") or {}).get("relevant_count_top50"), 0) or 0,
            safe_int((x.get("metrics") or {}).get("relevance_score"), 0) or 0,
        ), reverse=True)
        best = rows_sorted[0]
        m = best.get("metrics", {}) if isinstance(best.get("metrics"), dict) else {}
        return {
            "query_id": safe_str(best.get("query_id")),
            "query_text": safe_str(best.get("query_text")),
            "serp_cache_path": safe_str(best.get("serp_cache_path")),
            "relevant_count_top50": safe_int(m.get("relevant_count_top50"), None),
            "total_estimate": safe_int(m.get("total_estimate"), None),
            "price_p10": safe_int(m.get("price_p10"), None),
            "price_median": safe_int(m.get("price_median"), None),
            "price_p90": safe_int(m.get("price_p90"), None),
            "monster_share_top20": safe_float(m.get("monster_share_top20"), None),
            "relevance_score": safe_int(m.get("relevance_score"), None),
        }

    for i, sku in enumerate(scope, start=1):
        owner = nm_id_to_str(sku["nm_id"])
        if resume and owner in already:
            if verbose:
                print(f"[I] skip nm={owner}")
            continue

        vendor_code = safe_str(sku.get("vendor_code", ""))
        name = safe_str(sku.get("name", ""))

        own_rec = own_map.get(owner)
        sel = sel_map.get(owner)
        if not own_rec or not sel:
            raise ValueError(f"missing deps for stage_I owner={owner}")

        own_card = own_rec.get("own_card", {}) if isinstance(own_rec.get("own_card"), dict) else {}
        selected_ids = sel.get("selection", {}).get("selected_ids", [])
        if not isinstance(selected_ids, list):
            selected_ids = []

        comps: List[dict] = []
        comp_summ = []
        for cid in selected_ids:
            cid_str = nm_id_to_str(cid)
            r = comp_by_owner.get(owner, {}).get(cid_str)
            if not r:
                continue
            cc = r.get("competitor_card", {})
            if isinstance(cc, dict) and cc:
                comps.append(cc)
                title = safe_str(cc.get("content", {}).get("title", ""))[:160]
                pr = safe_int(cc.get("pricing", {}).get("sale_price") or cc.get("pricing", {}).get("price"), None)
                rat = safe_float(cc.get("social", {}).get("rating"), None)
                fb = safe_int(cc.get("social", {}).get("feedbacks"), None)
                comp_summ.append({"nm_id": cid_str, "title": title, "price": pr, "rating": rat, "feedbacks": fb, "url": wb_product_url(cid_str)})

        score = _score_vs_competitors(own_card, comps)
        gaps = _compute_feature_material_gaps(own_card, comps)
        gaps["keyword_gap"] = _compute_keyword_gap(owner, own_card, comps)
        serp_best = _pick_best_serp(owner)

        rec = {
            "meta": make_meta(run_id, "I", owner, vendor_code, name=name, source="script"),
            "metrics": score,
            "gaps": gaps,
            "market": {"best_serp": serp_best} if serp_best else {"best_serp": None},
            "competitors": comp_summ,
            "errors": [],
            "warnings": [] if comps else [{"code": "NO_COMP_DATA", "msg": "No competitor cards fetched", "where": "stage_I"}],
        }
        validate_comparison_metrics_record(rec)
        append_jsonl(out_path, rec)
        if verbose:
            qtxt = serp_best.get("query_text") if serp_best else ""
            print(f"[I] ({i}/{total}) owner={owner} comps={len(comp_summ)} q={qtxt[:40]!r}")

# =========================
# Stage J: Decide verdict -> decisions.jsonl
# =========================

VERDICTS = ["REVIVE_FAST", "REVIVE_REWORK", "CLONE_NEW_CARD", "DROP"]

def validate_decision_record(rec: dict) -> None:
    validate_meta(rec.get("meta", {}))
    d = rec.get("decision")
    if not isinstance(d, dict):
        raise ValueError("decision missing")

    if safe_str(d.get("verdict")) not in VERDICTS:
        raise ValueError("verdict invalid")

    rf = d.get("risk_flags")
    if not isinstance(rf, list):
        raise ValueError("risk_flags must be list")
    for x in rf:
        if safe_str(x).strip() == "":
            raise ValueError("empty risk flag")

    bl = d.get("backlog")
    if not isinstance(bl, list):
        raise ValueError("backlog must be list")
    for it in bl:
        if not isinstance(it, dict):
            raise ValueError("backlog item must be dict")
        if safe_str(it.get("task")).strip() == "":
            raise ValueError("backlog.task missing")
        pr = it.get("prio", 9)
        if not isinstance(pr, int):
            raise ValueError("backlog.prio must be int")

    rat = d.get("rationale")
    if rat is not None:
        if not isinstance(rat, list):
            raise ValueError("rationale must be list")
        for x in rat:
            if safe_str(x).strip() == "":
                raise ValueError("rationale item empty")

def build_verdict_prompt(own_card: dict, met_rec: dict) -> List[dict]:
    """Strict JSON-only prompt for Stage J (Verdict + Backlog).

    Goal: anchor rationale in explicit evidence (numbers) so the final HTML looks like an audit,
    not a vibes-based horoscope.
    """

    own = {
        "title": safe_str((own_card.get("content") or {}).get("title", ""))[:220],
        "desc_len": len(safe_str((own_card.get("content") or {}).get("description", ""))),
        "models": (own_card.get("attributes") or {}).get("phone_models", []) if isinstance((own_card.get("attributes") or {}).get("phone_models"), list) else [],
        "materials": (own_card.get("attributes") or {}).get("materials", []) if isinstance((own_card.get("attributes") or {}).get("materials"), list) else [],
        "features": (own_card.get("attributes") or {}).get("features", []) if isinstance((own_card.get("attributes") or {}).get("features"), list) else [],
        "media": own_card.get("media") if isinstance(own_card.get("media"), dict) else {},
        "social": own_card.get("social") if isinstance(own_card.get("social"), dict) else {},
        "pricing": own_card.get("pricing") if isinstance(own_card.get("pricing"), dict) else {},
    }

    metrics = met_rec.get("metrics") if isinstance(met_rec.get("metrics"), dict) else {}
    market = met_rec.get("market") if isinstance(met_rec.get("market"), dict) else {}
    comps = met_rec.get("competitors") if isinstance(met_rec.get("competitors"), list) else []
    gaps = met_rec.get("gaps") if isinstance(met_rec.get("gaps"), dict) else {}

    best_serp = (market.get("best_serp") if isinstance(market.get("best_serp"), dict) else None)
    evidence = {
        "competitors_compared": int(len(comps)),
        "own_price": safe_int(metrics.get("own_price"), None),
        "comp_price_median": safe_int(metrics.get("comp_price_median"), None),
        "own_vs_comp_median_price_pct": metrics.get("own_vs_comp_median_price_pct") if isinstance(metrics.get("own_vs_comp_median_price_pct"), (int, float)) else None,
        "own_rating": metrics.get("own_rating") if isinstance(metrics.get("own_rating"), (int, float)) else None,
        "own_feedbacks": safe_int(metrics.get("own_feedbacks"), None),
        "own_photos": safe_int(metrics.get("own_photos"), None),
        "own_video": safe_int(metrics.get("own_video"), None),
        "best_query_text": safe_str(best_serp.get("query_text"))[:120] if isinstance(best_serp, dict) else "",
        "best_query_rel50": safe_int(best_serp.get("relevant_count_top50"), None) if isinstance(best_serp, dict) else None,
        "best_query_total_estimate": safe_int(best_serp.get("total_estimate"), None) if isinstance(best_serp, dict) else None,
    }

    sys = (
        "Ты жёсткий, но честный аналитик маркетплейса Wildberries.\n"
        "Реши, что делать с SKU ТОЛЬКО по переданным фактам.\n"
        "Верни ТОЛЬКО JSON-объект. Без markdown и без текста вокруг.\n\n"
        "Допустимые значения verdict (НЕ переводить):\n"
        "- REVIVE_FAST\n"
        "- REVIVE_REWORK\n"
        "- CLONE_NEW_CARD\n"
        "- DROP\n\n"
        "Схема ответа:\n"
        "{\n"
        "  \"verdict\": \"...\",\n"
        "  \"risk_flags\": [\"...\"],\n"
        "  \"backlog\": [{\"task\":\"...\",\"prio\":1}],\n"
        "  \"rationale\": [\"...\"],\n"
        "  \"confidence\": 0.0\n"
        "}\n\n"
        "Правила:\n"
        "- Никаких галлюцинаций. Если данных мало, добавь флаг LOW_EVIDENCE.\n"
        "- Числа и конкретные формулировки бери ТОЛЬКО из блока evidence/metrics/market/gaps (никаких придуманных процентов и 'много/мало' без опоры).\n"
        "- В rationale: 3–6 пунктов и минимум 3 пункта должны ссылаться на конкретные факты/числа из evidence.\n"
        "- risk_flags: пиши UPPER_SNAKE_CASE на английском (как LOW_EVIDENCE), без русских слов.\n"
        "- backlog.task и rationale: пиши ТОЛЬКО по-русски.\n"
        "- Backlog задачи должны быть конкретные и исполнимые.\n"
        "- prio: 1 (срочно) .. 3 (не горит).\n"
        "- Используй gaps.keyword_gap (must_terms_missing_in_own / ban_terms_found_in_own / missing_*_terms) как доказательство для переписывания контента.\n"
    )

    payload = {
        "evidence": evidence,
        "own": {
            "title": own["title"],
            "desc_len": own["desc_len"],
            "models": own["models"][:5],
            "materials": own["materials"][:5],
            "features": own["features"][:12],
            "media": own["media"],
            "social": own["social"],
            "pricing": own["pricing"],
        },
        "metrics": metrics,
        "market": market,
        "gaps": gaps,
        "competitors": comps[:5],
    }

    return [
        {"role": "system", "content": sys},
        {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
    ]

def _validate_llm_decision_payload(obj: dict) -> dict:
    if not isinstance(obj, dict):
        raise ValueError("LLM decision: not a dict")
    verdict = safe_str(obj.get("verdict"))
    if verdict not in VERDICTS:
        raise ValueError("LLM decision: invalid verdict")
    rf = obj.get("risk_flags", [])
    if not isinstance(rf, list):
        raise ValueError("LLM decision: risk_flags not list")
    rf = [safe_str(x).strip() for x in rf if safe_str(x).strip()]
    bl = obj.get("backlog", [])
    if not isinstance(bl, list):
        raise ValueError("LLM decision: backlog not list")
    backlog = []
    for it in bl[:12]:
        if not isinstance(it, dict):
            continue
        task = safe_str(it.get("task")).strip()
        if not task:
            continue
        pr = it.get("prio", 3)
        try:
            pr = int(pr)
        except Exception:
            pr = 3
        pr = 1 if pr < 1 else 3 if pr > 3 else pr
        backlog.append({"task": task, "prio": pr})
    rat = obj.get("rationale", [])
    if not isinstance(rat, list):
        rat = []
    rationale = [safe_str(x).strip() for x in rat if safe_str(x).strip()][:6]
    conf = obj.get("confidence", None)
    try:
        conf = float(conf) if conf is not None else None
    except Exception:
        conf = None
    if conf is not None:
        conf = max(0.0, min(1.0, conf))
    return {"verdict": verdict, "risk_flags": rf, "backlog": backlog, "rationale": rationale, "confidence": conf}


def decide_from_metrics(own_card: dict, metrics: dict):
    """Deterministic fallback for Stage J when LLM is disabled or fails.

    NOTE: This is intentionally conservative. It uses only metrics computed in Stage I
    (own vs competitor medians) plus minimal signals from own_card.
    Returns: (verdict, risk_flags:list[str], backlog:list[dict], rationale:list[str])
    """
    # --- extract signals ---
    own_price = safe_int(metrics.get("own_price"), None)
    comp_med = safe_int(metrics.get("comp_price_median"), None)
    price_pct = metrics.get("own_vs_comp_median_price_pct")
    if not isinstance(price_pct, (int, float)):
        price_pct = None
    else:
        price_pct = float(price_pct)

    own_rating = safe_float(metrics.get("own_rating"), None)
    comp_rating = safe_float(metrics.get("comp_rating_median"), None)
    own_fb = safe_int(metrics.get("own_feedbacks"), None)
    comp_fb = safe_int(metrics.get("comp_feedbacks_median"), None)

    own_ph = safe_int(metrics.get("own_photos"), None)
    comp_ph = safe_int(metrics.get("comp_photos_median"), None)
    own_vid = safe_int(metrics.get("own_video"), None)
    comp_vid = safe_int(metrics.get("comp_video_median"), None)

    gap_ph = safe_int(metrics.get("content_gap_photos"), None)
    gap_vid = safe_int(metrics.get("content_gap_video"), None)

    title = safe_str((own_card.get("content") or {}).get("title", ""))[:200]
    desc_len = len(safe_str((own_card.get("content") or {}).get("description", "")))

    risk_flags = []
    backlog = []
    rationale = []

    # evidence check
    if comp_med is None and comp_fb is None and comp_ph is None:
        risk_flags.append("LOW_EVIDENCE")
        rationale.append("Мало данных по конкурентам: решения будут консервативными.")

    # --- content quality signals ---
    if gap_ph is not None and gap_ph >= 3:
        risk_flags.append("CONTENT_WEAK_PHOTOS")
        backlog.append({"task": f"Добавить фото: +{gap_ph} (до медианы конкурентов)", "prio": 1})
        rationale.append(f"Фото ниже медианы конкурентов на {gap_ph}.")
    elif own_ph is not None and own_ph <= 2:
        risk_flags.append("CONTENT_WEAK_PHOTOS")
        backlog.append({"task": "Усилить фотоконтент: 6–10 фото (детали, пример на телефоне, упаковка)", "prio": 2})
        rationale.append("Мало фото в карточке.")

    if (own_vid is not None and own_vid == 0) and (comp_vid is not None and comp_vid > 0):
        risk_flags.append("CONTENT_NO_VIDEO")
        backlog.append({"task": "Добавить видео 10–25 сек (посадка, карман, кнопки, материал)", "prio": 2})
        rationale.append("У конкурентов чаще есть видео, у товара нет.")

    if desc_len < 350:
        risk_flags.append("CONTENT_THIN_DESCRIPTION")
        backlog.append({"task": "Переписать описание: польза, совместимость, материалы, преимущества, уход", "prio": 2})
        rationale.append("Описание слишком короткое.")

    if len(title) < 25:
        risk_flags.append("CONTENT_WEAK_TITLE")
        backlog.append({"task": "Усилить заголовок: тип чехла + модель + ключевая фича (без спама)", "prio": 2})
        rationale.append("Заголовок слабый/короткий.")

    # --- price positioning ---
    if price_pct is not None:
        if price_pct > 30:
            risk_flags.append("PRICE_TOO_HIGH")
            if comp_med is not None:
                backlog.append({"task": f"Снизить цену/включить промо: цель около медианы рынка {comp_med} ₽", "prio": 1})
            else:
                backlog.append({"task": "Проверить цену относительно конкурентов и скорректировать промо/скидку", "prio": 1})
            rationale.append(f"Цена выше медианы рынка примерно на {price_pct:.1f}%.")
        elif price_pct < -25:
            risk_flags.append("PRICE_TOO_LOW")
            rationale.append(f"Цена ниже медианы рынка примерно на {price_pct:.1f}% (возможна недомонетизация).")

    # --- social proof ---
    if own_fb is not None:
        if own_fb < 10:
            risk_flags.append("LOW_REVIEWS")
            backlog.append({"task": "Нарастить отзывы: вкладыш/чат-поддержка/UGC, цель 20+ отзывов", "prio": 1})
            rationale.append("Мало отзывов, слабый социальный сигнал.")
        elif comp_fb is not None and own_fb < max(5, int(comp_fb * 0.4)):
            risk_flags.append("REVIEWS_BELOW_MARKET")
            backlog.append({"task": "Усилить сбор отзывов, чтобы приблизиться к уровню конкурентов", "prio": 2})
            rationale.append("Отзывы заметно ниже рынка.")

    if own_rating is not None:
        if own_rating < 3.8 and (own_fb or 0) >= 20:
            risk_flags.append("BAD_RATING_HISTORY")
            rationale.append("Низкий рейтинг при заметном числе отзывов: вероятно, системная проблема товара/ожиданий.")
        elif comp_rating is not None and own_rating + 0.2 < comp_rating:
            risk_flags.append("RATING_BELOW_MARKET")
            rationale.append("Рейтинг ниже медианы рынка.")

    # --- verdict decision ---
    verdict = "REVIVE_REWORK"  # default: safe option

    # If established low rating history -> clone new card (or major reposition)
    if "BAD_RATING_HISTORY" in risk_flags:
        verdict = "CLONE_NEW_CARD"
        backlog.insert(0, {"task": "Проверить причины низкого рейтинга и устранить; рассмотреть новую карточку/перезапуск", "prio": 1})

    # Fast revive if only minor fixes needed
    minor = (
        ("PRICE_TOO_HIGH" not in risk_flags) and
        ("CONTENT_WEAK_PHOTOS" not in risk_flags) and
        ("CONTENT_NO_VIDEO" not in risk_flags) and
        (own_rating is None or own_rating >= 4.5) and
        (own_fb is None or own_fb >= 15)
    )
    if minor and "LOW_EVIDENCE" not in risk_flags:
        verdict = "REVIVE_FAST"

    # Drop only if extremely poor signals (conservative)
    if own_rating is not None and own_rating < 3.5 and (own_fb or 0) >= 30:
        verdict = "DROP"
        risk_flags.append("VERY_LOW_RATING")

    # ensure rationale length
    if not rationale:
        rationale = ["Детерминированный фоллбек: не хватило сигналов для точного решения."]

    # cap backlog
    backlog = backlog[:12]
    return verdict, risk_flags, backlog, rationale


def stage_J(
    out_dir: Path,
    *,
    provider: str,
    model: str,
    api_key: str,
    base_url: str = "",
    use_llm: bool,
    llm_timeout: int,
    max_tokens: int,
    temperature: float,
    resume: bool = False,
    verbose: bool = False,
) -> None:
    """
    Decide verdict per SKU.

    V1.2 improvement:
    - optional LLM verdict writer (use_llm)
    - deterministic fallback remains (decide_from_metrics)
    """
    manifest = read_json(out_dir / "run_manifest.json")
    run_id = manifest["run_id"]
    scope = manifest["scope"]["sku_list"]
    total = len(scope)

    own_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "own_norm.jsonl")}
    met_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "comparison_metrics.jsonl")}

    out_path = out_dir / "decisions.jsonl"
    already = set()
    if resume and out_path.exists():
        for r in read_jsonl(out_path):
            already.add(nm_id_to_str(r.get("meta", {}).get("nm_id")))

    for i, sku in enumerate(scope, start=1):
        nm_id = nm_id_to_str(sku["nm_id"])
        if resume and nm_id in already:
            if verbose:
                print(f"[J] skip nm={nm_id}")
            continue

        vendor_code = safe_str(sku.get("vendor_code", ""))
        name = safe_str(sku.get("name", ""))

        own_rec = own_map.get(nm_id)
        met_rec = met_map.get(nm_id)
        if not own_rec or not met_rec:
            raise ValueError(f"missing deps for stage_J nm={nm_id}")

        own_card = own_rec.get("own_card", {}) if isinstance(own_rec.get("own_card"), dict) else {}
        metrics = met_rec.get("metrics", {}) if isinstance(met_rec.get("metrics"), dict) else {}

        warnings: List[dict] = []
        errors: List[dict] = []
        llm_dbg = {"used": False, "provider": provider, "model": model, "http_code": None, "usage": {}}

        decision = None

        if use_llm:
            try:
                messages = build_verdict_prompt(own_card, met_rec)
                parsed, dbg = call_llm_json(
                    provider=provider,
                    model=model,
                    api_key=api_key,
                    base_url=base_url,
                    messages=messages,
                    timeout_sec=llm_timeout,
                    max_tokens=max_tokens,
                    temperature=temperature,
                    force_json=True,
                )
                llm_dbg["used"] = True
                llm_dbg.update(dbg)
                clean = _validate_llm_decision_payload(parsed)

                # add LOW_EVIDENCE if we are missing key signals
                if (metrics.get("comp_price_median") is None) and (metrics.get("comp_feedbacks_median") is None):
                    if "LOW_EVIDENCE" not in clean["risk_flags"]:
                        clean["risk_flags"].append("LOW_EVIDENCE")

                decision = {
                    "verdict": clean["verdict"],
                    "risk_flags": sorted(list(set(clean["risk_flags"]))),
                    "backlog": sorted(clean["backlog"], key=lambda x: int(x.get("prio", 9))),
                    "rationale": clean["rationale"],
                }
                if clean.get("confidence") is not None:
                    decision["confidence"] = clean["confidence"]
            except Exception as e:
                warnings.append({"code": "LLM_FAIL", "msg": safe_str(e)[:240], "where": "stage_J"})
                decision = None

        if decision is None:
            verdict, flags, backlog, rationale = decide_from_metrics(own_card, metrics)
            decision = {
                "verdict": verdict,
                "risk_flags": sorted(list(set(flags))),
                "backlog": sorted(backlog, key=lambda x: int(x.get("prio", 9))),
                "rationale": rationale[:6],
            }

        rec = {
            "meta": make_meta(run_id, "J", nm_id, vendor_code, name=name, source="llm" if llm_dbg["used"] else "script"),
            "decision": decision,
            "llm_debug": llm_dbg,
            "errors": errors,
            "warnings": warnings,
        }
        validate_decision_record(rec)
        append_jsonl(out_path, rec)
        if verbose:
            print(f"[J] ({i}/{total}) nm={nm_id} verdict={decision.get('verdict')} llm={llm_dbg['used']} flags={len(decision.get('risk_flags', []))}")

# --- chunk3 end ---


# =========================
# Stage K: Reports (XLSX + HTML) + exec_summary.json
# =========================

def _xlsx_autofit(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = safe_str(v)
            max_len = max(max_len, len(s))
        ws.column_dimensions[get_column_letter(col)].width = min(60, max(10, max_len + 2))
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

def _style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F2937")  # dark gray
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.freeze_panes = ws["A2"]

def _fill_by_verdict(verdict: str) -> PatternFill:
    # ok-ish colors but without being too loud
    if verdict == "REVIVE_FAST":
        return PatternFill("solid", fgColor="D1FAE5")  # green tint
    if verdict == "REVIVE_REWORK":
        return PatternFill("solid", fgColor="FEF3C7")  # amber tint
    if verdict == "CLONE_NEW_CARD":
        return PatternFill("solid", fgColor="DBEAFE")  # blue tint
    return PatternFill("solid", fgColor="FEE2E2")      # red tint



def summarize_run_facts(out_dir: Path) -> dict:
    """Aggregate evidence numbers from canonical artifacts.

    Used to make final LLM outputs more accountable (no invented numbers).
    All values are best-effort and tolerate missing files.
    """

    def _try_jsonl(path: Path):
        try:
            return read_jsonl(path) if path.exists() else []
        except Exception:
            return []

    def _try_json(path: Path):
        try:
            return read_json(path) if path.exists() else {}
        except Exception:
            return {}

    mf = _try_json(out_dir / 'run_manifest.json')
    sku_count = None
    try:
        sku_count = len(((mf.get('scope', {}) or {}).get('sku_list', [])) or [])
    except Exception:
        sku_count = None

    # ---- Stage D: queries_valid ----
    qv = _try_jsonl(out_dir / 'queries_valid.jsonl')
    valid_queries_total = 0
    rel50_gt0 = 0
    rel50_vals = []
    for r in qv:
        if not isinstance(r, dict):
            continue
        valid_ids = r.get('valid_query_ids', []) if isinstance(r.get('valid_query_ids'), list) else []
        valid_set = set(safe_str(x) for x in valid_ids)
        rows = (r.get('validation') or {}).get('queries', []) if isinstance((r.get('validation') or {}), dict) else []
        if not isinstance(rows, list):
            rows = []
        valid_queries_total += len(valid_ids)
        for it in rows:
            if not isinstance(it, dict):
                continue
            if safe_str(it.get('query_id')) not in valid_set:
                continue
            m = it.get('metrics', {}) if isinstance(it.get('metrics'), dict) else {}
            rel = safe_int(m.get('relevant_count_top50'), None)
            if rel is None:
                continue
            rel50_vals.append(int(rel))
            if rel > 0:
                rel50_gt0 += 1

    rel50_median = None
    if rel50_vals:
        xs = sorted(rel50_vals)
        rel50_median = xs[len(xs)//2]

    rel50_gt0_share = None
    if valid_queries_total:
        rel50_gt0_share = round(rel50_gt0 / valid_queries_total * 100.0, 1)

    # ---- Stage E: competitor_pool ----
    pool = _try_jsonl(out_dir / 'competitor_pool.jsonl')
    pool_candidates_total = 0
    pool_selected_total = 0
    pool_selected_per_sku = []
    for r in pool:
        if not isinstance(r, dict):
            continue
        p = r.get('pool', {}) if isinstance(r.get('pool'), dict) else {}
        cand = p.get('candidates', []) if isinstance(p.get('candidates'), list) else []
        sel = p.get('selected_ids', []) if isinstance(p.get('selected_ids'), list) else []
        pool_candidates_total += len(cand)
        pool_selected_total += len(sel)
        pool_selected_per_sku.append(len(sel))

    pool_selected_median = None
    if pool_selected_per_sku:
        xs = sorted(pool_selected_per_sku)
        pool_selected_median = xs[len(xs)//2]

    # ---- Stage F: competitor_norm ----
    comp_norm = _try_jsonl(out_dir / 'competitor_norm.jsonl')
    comp_total = 0
    comp_ok = 0
    comp_fail = 0
    for r in comp_norm:
        if not isinstance(r, dict):
            continue
        comp_total += 1
        errs = r.get('errors', []) if isinstance(r.get('errors'), list) else []
        card = r.get('competitor_card', {}) if isinstance(r.get('competitor_card'), dict) else {}
        if errs or not card:
            comp_fail += 1
        else:
            comp_ok += 1

    # ---- Stage G: llm_relevance ----
    rel = _try_jsonl(out_dir / 'llm_relevance.jsonl')
    rel_items_total = 0
    rel_keep_total = 0
    for r in rel:
        if not isinstance(r, dict):
            continue
        items = (r.get('relevance') or {}).get('items', []) if isinstance((r.get('relevance') or {}), dict) else []
        if not isinstance(items, list):
            continue
        for it in items:
            if not isinstance(it, dict):
                continue
            rel_items_total += 1
            if safe_str(it.get('label')) == 'KEEP':
                rel_keep_total += 1

    rel_keep_share = None
    if rel_items_total:
        rel_keep_share = round(rel_keep_total / rel_items_total * 100.0, 1)

    # ---- Stage H: competitors_selected ----
    sel = _try_jsonl(out_dir / 'competitors_selected.jsonl')
    final_selected_total = 0
    final_selected_per_sku = []
    for r in sel:
        if not isinstance(r, dict):
            continue
        s = r.get('selection', {}) if isinstance(r.get('selection'), dict) else {}
        ids = s.get('selected_ids', []) if isinstance(s.get('selected_ids'), list) else []
        final_selected_total += len(ids)
        final_selected_per_sku.append(len(ids))

    final_selected_median = None
    if final_selected_per_sku:
        xs = sorted(final_selected_per_sku)
        final_selected_median = xs[len(xs)//2]

    # ---- Stage I: comparisons (competitors compared) ----
    met = _try_jsonl(out_dir / 'comparison_metrics.jsonl')
    comparisons_total = 0
    best_serp_present = 0
    for r in met:
        if not isinstance(r, dict):
            continue
        comps = r.get('competitors', []) if isinstance(r.get('competitors'), list) else []
        comparisons_total += len(comps)
        best = (r.get('market') or {}).get('best_serp') if isinstance((r.get('market') or {}), dict) else None
        if isinstance(best, dict) and safe_str(best.get('query_text')).strip():
            best_serp_present += 1

    return {
        'sku_count': sku_count,
        'valid_queries_total': valid_queries_total,
        'valid_queries_rel50_gt0': rel50_gt0,
        'valid_queries_rel50_gt0_share_pct': rel50_gt0_share,
        'valid_queries_rel50_median': rel50_median,
        'competitor_pool_candidates_total': pool_candidates_total,
        'competitor_pool_selected_total': pool_selected_total,
        'competitor_pool_selected_median_per_sku': pool_selected_median,
        'competitor_cards_total': comp_total,
        'competitor_cards_ok': comp_ok,
        'competitor_cards_fail': comp_fail,
        'llm_relevance_items_total': rel_items_total,
        'llm_relevance_keep_total': rel_keep_total,
        'llm_relevance_keep_share_pct': rel_keep_share,
        'final_competitors_selected_total': final_selected_total,
        'final_competitors_selected_median_per_sku': final_selected_median,
        'comparisons_total': comparisons_total,
        'best_serp_present_sku': best_serp_present,
    }


def build_exec_summary(
    out_dir: Path,
    *,
    llm_provider: str = "openrouter",
    model: str = "gpt-4o-mini",
    api_key: str = "",
    base_url: str = "",
    use_llm: bool = False,
    llm_timeout: int = 60,
    max_tokens: int = 800,
    temperature: float = 0.2,
) -> dict:
    decisions = read_jsonl(out_dir / "decisions.jsonl")
    met = read_jsonl(out_dir / "comparison_metrics.jsonl")

    sku_count = None
    try:
        mf = read_json(out_dir / "run_manifest.json")
        sku_count = len(mf.get("scope", {}).get("sku_list", []) or [])
    except Exception:
        sku_count = None

    run_facts = summarize_run_facts(out_dir)

    by_nm = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in decisions}
    met_by_nm = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in met}

    buckets: Dict[str, List[str]] = {v: [] for v in VERDICTS}
    risk_counts: Dict[str, int] = {}
    task_counts: Dict[str, int] = {}

    for nm, r in by_nm.items():
        v = r.get("decision", {}).get("verdict")
        if v in buckets:
            buckets[v].append(nm)

        for f in r.get("decision", {}).get("risk_flags", []):
            f = safe_str(f).strip()
            if not f:
                continue
            risk_counts[f] = risk_counts.get(f, 0) + 1

        for t in r.get("decision", {}).get("backlog", []):
            if not isinstance(t, dict):
                continue
            task = safe_str(t.get("task", "")).strip()
            if not task:
                continue
            task_counts[task] = task_counts.get(task, 0) + 1

    top_risks = sorted(risk_counts.items(), key=lambda x: x[1], reverse=True)[:12]
    top_tasks = sorted(task_counts.items(), key=lambda x: x[1], reverse=True)[:15]

    own_prices = []
    price_vs_med = []
    for nm, mr in met_by_nm.items():
        m = mr.get("metrics", {}) if isinstance(mr.get("metrics"), dict) else {}
        p = safe_int(m.get("own_price"), None)
        if p is not None:
            own_prices.append(p)
        pct = m.get("own_vs_comp_median_price_pct")
        if isinstance(pct, (int, float)):
            price_vs_med.append(float(pct))

    summary = {
        "schema_version": SCHEMA_VERSION,
        "created_at": utc_now_iso(),
        "sku_count": sku_count,
        "counts": {k: len(v) for k, v in buckets.items()},
        "buckets": buckets,
        "top_risk_flags": [{"flag": k, "count": v} for k, v in top_risks],
        "top_backlog_tasks": [{"task": k, "count": v} for k, v in top_tasks],
        "own_price_stats": {
            "min": min(own_prices) if own_prices else None,
            "median": _percentile_int(own_prices, 50) if own_prices else None,
            "max": max(own_prices) if own_prices else None,
        },
        "own_vs_comp_median_price_pct_stats": {
            "median": round(sorted(price_vs_med)[len(price_vs_med)//2], 1) if price_vs_med else None,
        },
        "run_facts": run_facts,
        "llm_summary": None,
        "llm_debug": {"used": bool(use_llm), "error": None},
    }

    if use_llm:
        try:
            payload = {
                "facts": run_facts,
                "sku_count": summary.get("sku_count"),
                "counts": summary["counts"],
                "top_risk_flags": summary["top_risk_flags"][:10],
                "top_backlog_tasks": summary["top_backlog_tasks"][:10],
                "own_price_stats": summary["own_price_stats"],
                "own_vs_comp_median_price_pct_stats": summary["own_vs_comp_median_price_pct_stats"],
                "bucket_examples": {k: v[:5] for k, v in buckets.items()},
            }
            messages = build_exec_summary_prompt(payload)
            parsed, dbg = call_llm_json(
                provider=llm_provider,
                model=model,
                api_key=api_key,
                base_url=base_url,
                messages=messages,
                timeout_sec=llm_timeout,
                max_tokens=max_tokens,
                temperature=temperature,
            )
            if isinstance(parsed, dict):
                summary["llm_summary"] = parsed
            summary["llm_debug"] = dbg
        except Exception as e:
            summary["llm_debug"] = {"used": True, "error": f"{type(e).__name__}: {e}"}

    return summary
def build_exec_summary_prompt(payload: dict) -> List[dict]:
    facts = payload.get("facts") if isinstance(payload.get("facts"), dict) else {}
    system = (
        "Ты пишешь Executive Summary по результатам проекта WB Revival (Wildberries).\n"
        "У тебя есть агрегированные числа и примеры nm_id по корзинам, плюс блок FACTS.\n"
        "Критично: ЗАПРЕЩЕНО выдумывать цифры или факты вне FACTS/входного JSON. Если цифры нет, так и пиши: 'нет данных'.\n\n"
        "Формат: верни ТОЛЬКО валидный JSON-объект без markdown и без текста вокруг.\n"
        "Схема ответа:\n"
        "{\n"
        "  \"headline\": string,\n"
        "  \"overall\": string,\n"
        "  \"key_findings\": [string, ...],\n"
        "  \"recommended_actions\": [string, ...],\n"
        "  \"watchouts\": [string, ...]\n"
        "}\n\n"
        "Ограничения: headline <= 140 символов; overall <= 700 символов; в списках по 3–7 пунктов.\n"
        "В overall ОБЯЗАТЕЛЬНО добавь строку, начинающуюся с 'Основание:' и упомяни там 3–5 чисел из FACTS "
        "(например: валидные запросы, доля rel50>0, карточек конкурентов OK, финальных сравнений).\n"
        "Не используй слова типа 'лучший', 'номер 1', 'топ'."
    )

    user = (
        "FACTS (используй числа только отсюда):\n" + json.dumps(facts, ensure_ascii=False, indent=2) +
        "\n\nDATA (остальные агрегаты):\n" + json.dumps(payload, ensure_ascii=False)
    )

    return [
        {"role": "system", "content": system},
        {"role": "user", "content": user},
    ]
def stage_K(out_dir: Path, *, xlsx_name: str = "", html_name: str = "", verbose: bool = False, llm_provider: str = "openrouter", llm_model: str = "openai/gpt-4o-mini", api_key: str = "", llm_base_url: str = "", use_llm_exec_summary: bool = False, llm_timeout: int = 60, llm_max_tokens: int = 800, llm_temperature: float = 0.2) -> None:
    manifest = read_json(out_dir / "run_manifest.json")
    run_id = manifest["run_id"]
    scope = manifest["scope"]["sku_list"]
    n_scope = len(scope)
    if not safe_str(xlsx_name).strip():
        xlsx_name = f"WB_REVIVE_{n_scope}.xlsx"
    if not safe_str(html_name).strip():
        html_name = f"WB_REVIVE_{n_scope}.html"
    total = len(scope)

    own_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "own_norm.jsonl")}
    met_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "comparison_metrics.jsonl")}
    dec_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "decisions.jsonl")}
    sel_map = {nm_id_to_str(r.get("meta", {}).get("nm_id")): r for r in read_jsonl(out_dir / "competitors_selected.jsonl")}

    wb = Workbook()
    ws = wb.active
    ws.title = "REPORT"

    headers = [
        "nm_id","vendor_code","name","wb_url",
        "own_price","comp_price_median","own_vs_comp_median_price_pct",
        "own_rating","own_feedbacks","own_photos","own_video",
        "verdict","risk_flags","top_tasks",
        "competitors_count","competitors_urls",
    ]
    ws.append(headers)

    for sku in scope:
        nm = nm_id_to_str(sku["nm_id"])
        vc = safe_str(sku.get("vendor_code",""))
        nm_name = safe_str(sku.get("name",""))
        own_card = own_map.get(nm, {}).get("own_card", {})
        m = met_map.get(nm, {}).get("metrics", {})
        d = dec_map.get(nm, {}).get("decision", {})
        sel = sel_map.get(nm, {}).get("selection", {})
        comp_ids = sel.get("selected_ids", []) if isinstance(sel.get("selected_ids"), list) else []
        comp_urls = [wb_product_url(x) for x in comp_ids]

        own_price = safe_int(m.get("own_price"), None)
        own_rating = m.get("own_rating")
        own_fb = safe_int(m.get("own_feedbacks"), None)
        own_ph = safe_int(m.get("own_photos"), 0) or 0
        own_vid = safe_int(m.get("own_video"), 0) or 0

        verdict = safe_str(d.get("verdict",""))
        flags = ", ".join(map(str, d.get("risk_flags", [])[:10]))
        tasks = "; ".join([safe_str(t.get("task","")) for t in d.get("backlog", [])[:3]])

        ws.append([
            nm, vc, nm_name, wb_product_url(nm),
            own_price,
            safe_int(m.get("comp_price_median"), None),
            m.get("own_vs_comp_median_price_pct"),
            own_rating,
            own_fb,
            own_ph,
            own_vid,
            verdict,
            flags,
            tasks,
            len(comp_ids),
            "\n".join(comp_urls),
        ])

        # style verdict row
        fill = _fill_by_verdict(verdict)
        r = ws.max_row
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

    _style_header(ws, 1)
    _xlsx_autofit(ws)

    # add DETAILS sheet with competitor summaries
    ws2 = wb.create_sheet("COMPETITORS")
    ws2.append(["owner_nm_id","competitor_nm_id","title","price","rating","feedbacks","url"])
    comp_norm = read_jsonl(out_dir / "competitor_norm.jsonl")
    for r in comp_norm:
        owner = nm_id_to_str(r.get("meta", {}).get("nm_id"))
        cid = nm_id_to_str(r.get("competitor_nm_id"))
        cc = r.get("competitor_card", {})
        title = safe_str(cc.get("content", {}).get("title",""))
        pr = safe_int(cc.get("pricing", {}).get("sale_price") or cc.get("pricing", {}).get("price"), None)
        rat = safe_float(cc.get("social", {}).get("rating"), None)
        fb = safe_int(cc.get("social", {}).get("feedbacks"), None)
        ws2.append([owner, cid, title, pr, rat, fb, wb_product_url(cid)])
    _style_header(ws2, 1)
    _xlsx_autofit(ws2)

    xlsx_path = out_dir / xlsx_name
    wb.save(xlsx_path)

    # exec summary
    exec_summary = build_exec_summary(out_dir, llm_provider=llm_provider, model=llm_model, api_key=api_key, base_url=llm_base_url, use_llm=use_llm_exec_summary, llm_timeout=llm_timeout, max_tokens=llm_max_tokens, temperature=llm_temperature)
    exec_summary["run_id"] = run_id
    write_json(out_dir / "exec_summary.json", exec_summary)

    # HTML report
    html = build_html_report(out_dir, scope, met_map, dec_map)
    write_text_atomic(out_dir / html_name, html, encoding="utf-8")

    if verbose:
        print(f"[K] wrote {xlsx_path.name}, {html_name}, exec_summary.json")

def build_html_report(out_dir: Path, scope: List[dict], met_map: dict, dec_map: dict) -> str:
    """
    Pretty (human-first) HTML report.

    Goals:
      - Show key info immediately (counts + summary + top issues)
      - Keep technical junk collapsible
      - Russian UI labels
      - Offline single-file (no external deps)
    """
    import math
    import re
    from datetime import datetime, timezone

    def utc_now_iso_local() -> str:
        return datetime.now(timezone.utc).isoformat(timespec="seconds")

    # -------- load exec summary (if exists) --------
    exec_summary: Dict[str, Any] = {}
    p_exec = out_dir / "exec_summary.json"
    if p_exec.exists():
        try:
            exec_summary = read_json(p_exec)
        except Exception:
            exec_summary = {}

    llm_sum = exec_summary.get("llm_summary", {}) if isinstance(exec_summary, dict) else {}
    counts = exec_summary.get("counts", {}) if isinstance(exec_summary, dict) else {}
    run_facts = exec_summary.get("run_facts", {}) if isinstance(exec_summary, dict) else {}

    # -------- load REPORT sheet from the newest XLSX (preferred) --------
    report_rows: List[Dict[str, Any]] = []
    try:
        xlsx_candidates = sorted(out_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        xlsx_path = xlsx_candidates[0] if xlsx_candidates else None
        if xlsx_path is not None and xlsx_path.exists():
            wb = load_workbook(xlsx_path, data_only=True)
            if "REPORT" in wb.sheetnames:
                ws = wb["REPORT"]
                headers = []
                for cell in ws[1]:
                    headers.append(safe_str(cell.value).strip())
                # map columns
                idx_map = {h: i for i, h in enumerate(headers)}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or all(v is None for v in row):
                        continue
                    d: Dict[str, Any] = {}
                    for h, i in idx_map.items():
                        d[h] = row[i] if i < len(row) else None
                    report_rows.append(d)
    except Exception:
        report_rows = []

    # Fallback: reconstruct from maps (older runs). We keep it minimal.
    if not report_rows:
        for sku in scope:
            nm = nm_id_to_str(sku["nm_id"])
            name = safe_str(sku.get("name", ""))
            m = met_map.get(nm, {}).get("metrics", {})
            d = dec_map.get(nm, {})
            report_rows.append({
                "nm_id": nm,
                "vendor_code": safe_str(sku.get("vendor_code","")),
                "name": name,
                "wb_url": f"https://www.wildberries.ru/catalog/{nm}/detail.aspx",
                "own_price": m.get("own_price"),
                "comp_price_median": m.get("comp_price_median"),
                "own_vs_comp_median_price_pct": m.get("own_vs_comp_median_price_pct"),
                "own_rating": m.get("own_rating"),
                "own_feedbacks": m.get("own_feedbacks"),
                "own_photos": m.get("own_photos"),
                "own_video": m.get("own_video"),
                "verdict": d.get("verdict"),
                "risk_flags": ", ".join(d.get("risk_flags", []) or []) if isinstance(d.get("risk_flags"), list) else safe_str(d.get("risk_flags","")),
                "top_tasks": "; ".join(d.get("top_tasks", []) or []) if isinstance(d.get("top_tasks"), list) else safe_str(d.get("top_tasks","")),
                "competitors_count": m.get("competitors_count", ""),
                "competitors_urls": ", ".join(m.get("competitors_urls", []) or []) if isinstance(m.get("competitors_urls"), list) else safe_str(m.get("competitors_urls","")),
            })

    # -------- helpers --------
    FLAG_RU = {
        "NO_PRICE_SET":"Цена не установлена",
        "PRICE_NOT_SET":"Цена не установлена",
        "MISSING_PRICE":"Нет цены в карточке",
        "NO_PRICE":"Нет цены",
        "NO_OWN_PRICE":"Наша цена не указана",
        "PRICE_UNKNOWN":"Цена неизвестна",
        "NO_PRICING_DATA":"Нет данных по цене",
        "LOW_PRICE_DATA":"Мало данных по цене",
        "OVERPRICED_VS_COMPETITORS":"Цена выше конкурентов",
        "PRICE_ABOVE_COMPETITORS":"Цена выше конкурентов",
        "UNDERPRICED":"Сильно ниже рынка",
        "PRICE_UNDERCUTTING":"Демпинг",
        "MISSING_KEYWORDS":"Нет ключевых слов",
        "MISSING_KEYWORD_COVERAGE":"Слабое покрытие ключей",
        "KEYWORD_GAP":"Разрыв по ключам",
        "KEYWORD_GAPS":"Разрыв по ключам",
        "KEYWORD_GAP_ATTR_TERMS":"Разрыв по атрибутам",
        "MISSING_ATTR_TERMS":"Нет атрибутов",
        "MISSING_KEY_TERMS":"Нет ключевых терминов",
        "MISSING_TITLE_TERMS":"Слабый заголовок",
        "MISSING_MUST_TERMS":"Нет must-терминов",
        "MISSING_MUST_KEYWORDS":"Нет must-ключей",
        "MISSING_MUST_HAVE_KEYWORDS":"Нет must-ключей",
        "MISSING_MUST_HAVE_TERMS":"Нет must-терминов",
        "MUST_TERMS_MISSING_IN_OWN":"Must-термины отсутствуют",
        "PHOTO_DEFICIT":"Мало фото",
        "NO_PHOTOS":"Нет фото",
        "LOW_MEDIA_COUNT":"Мало медиа",
        "NO_DESCRIPTION":"Нет описания",
        "CONTENT_GAP":"Пробелы в контенте",
        "CONTENT_GAPS":"Пробелы в контенте",
        "CONTENT_GAP_FEATURES":"Нет функций в тексте",
        "CONTENT_GAP_MATERIALS":"Нет материалов в тексте",
        "MISSING_FEATURES":"Нет функций",
        "MISSING_FEATURE_TERMS":"Нет терминов функций",
        "MISSING_MATERIALS":"Нет материалов",
        "MISSING_MATERIAL_TERMS":"Нет терминов материалов",
        "FEATURES_GAP":"Разрыв по функциям",
        "MATERIALS_GAP":"Разрыв по материалам",
        "LOW_EVIDENCE":"Мало данных (низкая уверенность)",
        "LOW_DEMAND":"Низкий спрос",
        "NO_COMPETITOR_DATA":"Нет данных по конкурентам",
        "LOW_FEEDBACKS":"Мало отзывов",
        "NO_FEEDBACKS":"Нет отзывов",
        "LOW_FEEDBACKS_VS_COMPETITORS":"Отзывов меньше конкурентов",
        "LOW_RATING":"Низкий рейтинг",
        "NO_RATING":"Нет рейтинга",
        "NO_TITLE":"Нет заголовка",
    }

    def flag_kind(raw: str) -> str:
        if raw in {"NO_PRICE_SET","PRICE_NOT_SET","MISSING_PRICE","NO_PRICE","NO_OWN_PRICE","PRICE_UNKNOWN","NO_PRICING_DATA","LOW_PRICE_DATA","OVERPRICED_VS_COMPETITORS","PRICE_ABOVE_COMPETITORS","UNDERPRICED","PRICE_UNDERCUTTING"}:
            return "price"
        if ("KEYWORD" in raw) or ("MUST" in raw) or ("ATTR" in raw) or ("TITLE" in raw):
            return "seo"
        if ("PHOTO" in raw) or ("CONTENT" in raw) or ("DESCRIPTION" in raw) or ("MATERIAL" in raw) or ("FEATURE" in raw) or ("MEDIA" in raw):
            return "content"
        if ("FEEDBACK" in raw) or ("RATING" in raw):
            return "social"
        if ("EVIDENCE" in raw) or ("DEMAND" in raw) or ("COMPETITOR" in raw):
            return "meta"
        return "other"

    def parse_flags(s: Any) -> List[Dict[str,str]]:
        out: List[Dict[str,str]] = []
        for raw in [x.strip() for x in safe_str(s).split(",")]:
            if not raw:
                continue
            out.append({"raw": raw, "label": FLAG_RU.get(raw, raw), "kind": flag_kind(raw)})
        return out

    def parse_tasks(s: Any) -> List[str]:
        s2 = safe_str(s).strip()
        if not s2:
            return []
        parts = re.split(r"\s*;\s*|\s*\n\s*|\s*<br\s*/?>\s*", s2)
        return [p.strip() for p in parts if p.strip()]

    def fmt_num(x: Any, digits: int = 1) -> str:
        if x is None:
            return "—"
        try:
            if isinstance(x, str):
                x = x.strip()
                if not x:
                    return "—"
                # try float
                xf = float(x)
                x = xf
            if isinstance(x, (int, float)):
                if isinstance(x, float) and math.isnan(x):
                    return "—"
                if float(x).is_integer():
                    return str(int(x))
                return f"{x:.{digits}f}"
        except Exception:
            pass
        return safe_str(x) or "—"

    def verdict_badge(v: str) -> str:
        cls = {"REVIVE_FAST":"b-fast","REVIVE_REWORK":"b-rework","CLONE_NEW_CARD":"b-clone","DROP":"b-drop"}.get(v,"b-unk")
        ru = {"REVIVE_FAST":"Быстрый ревайв","REVIVE_REWORK":"Нужна доработка","CLONE_NEW_CARD":"Новая карточка","DROP":"Снять/архив"}.get(v, v)
        return f"<span class='badge {cls}' title='{html_escape(v)}'>{html_escape(ru)}</span>"

    def has_any(s: str, keys: List[str]) -> bool:
        return any(k in s for k in keys)

    def calc_cat(flags_str: str) -> Dict[str,int]:
        s = flags_str
        return {
            "price": int(has_any(s, ["NO_PRICE_SET","PRICE_NOT_SET","MISSING_PRICE","NO_PRICE","PRICE_UNKNOWN","NO_OWN_PRICE","NO_PRICING_DATA","LOW_PRICE_DATA","OVERPRICED_VS_COMPETITORS","PRICE_ABOVE_COMPETITORS","UNDERPRICED","PRICE_UNDERCUTTING"])),
            "seo": int(has_any(s, ["MISSING_KEYWORDS","MISSING_KEYWORD_COVERAGE","KEYWORD_GAP","KEYWORD_GAPS","MISSING_KEY_TERMS","MISSING_TITLE_TERMS","MISSING_ATTR_TERMS","MISSING_MUST_TERMS","MISSING_MUST_KEYWORDS","MISSING_MUST_HAVE_KEYWORDS","MISSING_MUST_HAVE_TERMS","MUST_TERMS_MISSING_IN_OWN"])),
            "content": int(has_any(s, ["PHOTO_DEFICIT","NO_PHOTOS","LOW_MEDIA_COUNT","NO_DESCRIPTION","CONTENT_GAP","CONTENT_GAPS","CONTENT_GAP_FEATURES","CONTENT_GAP_MATERIALS","MISSING_FEATURES","MISSING_FEATURE_TERMS","MISSING_MATERIALS","MISSING_MATERIAL_TERMS","FEATURES_GAP","MATERIALS_GAP"])),
            "social": int(has_any(s, ["LOW_FEEDBACKS","NO_FEEDBACKS","LOW_FEEDBACKS_VS_COMPETITORS","LOW_RATING","NO_RATING"])),
        }

    def list_to_ul(lst: List[str]) -> str:
        """Render list of strings as UL. Returns empty string for empty list."""
        if not lst:
            return ""
        return "<ul>" + "".join(f"<li>{html_escape(safe_str(x))}</li>" for x in lst) + "</ul>"

    # -------- aggregates for header cards --------
    cat_counts = {"price":0,"seo":0,"content":0,"social":0}
    for r in report_rows:
        cats = calc_cat(safe_str(r.get("risk_flags","")))
        for k in cat_counts:
            cat_counts[k] += cats[k]

    # -------- top flags / tasks (from exec_summary if present) --------
    top_flags = exec_summary.get("top_risk_flags", []) if isinstance(exec_summary, dict) else []
    top_tasks = exec_summary.get("top_backlog_tasks", []) if isinstance(exec_summary, dict) else []

    def render_top_flags(limit: int = 12) -> str:
        if not isinstance(top_flags, list):
            return "—"
        items = []
        for it in top_flags[:limit]:
            raw = safe_str(it.get("flag",""))
            if not raw:
                continue
            label = FLAG_RU.get(raw, raw)
            kind = flag_kind(raw)
            items.append(f"<span class='tag t-{kind}' title='{html_escape(raw)}'>{html_escape(label)} <span class='cnt'>{it.get('count',0)}</span></span>")
        return " ".join(items) if items else "—"

    def render_top_tasks(limit: int = 10) -> str:
        if not isinstance(top_tasks, list):
            return "—"
        li = []
        for it in top_tasks[:limit]:
            task = safe_str(it.get("task",""))
            if not task:
                continue
            li.append(f"<li><span class='cnt'>{it.get('count',0)}</span> {html_escape(task)}</li>")
        return "<ol class='toplist'>" + "".join(li) + "</ol>" if li else "—"

    def render_bullets(arr: Any) -> str:
        if not isinstance(arr, list) or not arr:
            return "—"
        return "<ul>" + "".join([f"<li>{html_escape(safe_str(x))}</li>" for x in arr]) + "</ul>"

    # -------- build table rows --------
    row_html: List[str] = []
    for r in report_rows:
        nm = nm_id_to_str(r.get("nm_id"))
        url = safe_str(r.get("wb_url")) or f"https://www.wildberries.ru/catalog/{nm}/detail.aspx"
        name = safe_str(r.get("name",""))
        vendor_code = safe_str(r.get("vendor_code",""))

        verdict = safe_str(r.get("verdict",""))
        flags = parse_flags(r.get("risk_flags",""))
        tasks = parse_tasks(r.get("top_tasks",""))

        cats = calc_cat(safe_str(r.get("risk_flags","")))
        tags_short = " ".join([f"<span class='tag t-{f['kind']}' title='{html_escape(f['raw'])}'>{html_escape(f['label'])}</span>" for f in flags[:4]])
        more = max(0, len(flags) - 4)
        if more > 0:
            tags_short += f" <span class='tag t-more'>+{more}</span>"

        own = fmt_num(r.get("own_price"), 0)
        comp = fmt_num(r.get("comp_price_median"), 0)
        dp = r.get("own_vs_comp_median_price_pct")
        dp_s = "—"
        try:
            if dp is not None and not (isinstance(dp, float) and math.isnan(dp)):
                dp_s = f"{float(dp):+.1f}%"
        except Exception:
            pass

        rating = fmt_num(r.get("own_rating"), 1)
        fb = fmt_num(r.get("own_feedbacks"), 0)
        photos = fmt_num(r.get("own_photos"), 0)
        video = "да" if safe_str(r.get("own_video")).strip() not in {"","0","—","None"} else "нет"

        comp_cnt = fmt_num(r.get("competitors_count"), 0)
        comp_urls = safe_str(r.get("competitors_urls",""))
        comp_links: List[str] = []
        if comp_urls.strip():
            for u in comp_urls.split(","):
                u = u.strip()
                if not u:
                    continue
                m = re.search(r"/catalog/(\d+)/", u)
                label = m.group(1) if m else "ссылка"
                comp_links.append(f"<a href='{html_escape(u)}' target='_blank'>{html_escape(label)}</a>")
        comp_links_html = " ".join(comp_links) if comp_links else "—"

        flags_full = " ".join([f"<span class='tag t-{f['kind']}' title='{html_escape(f['raw'])}'>{html_escape(f['label'])}</span>" for f in flags]) if flags else "—"
        tasks_html = "<ol>" + "".join([f"<li>{html_escape(t)}</li>" for t in tasks]) + "</ol>" if tasks else "—"

        # --- why (rationale) + evidence snapshot ---
        dec_rec = dec_map.get(nm, {}) if isinstance(dec_map, dict) else {}
        dec_obj = dec_rec.get('decision', {}) if isinstance(dec_rec, dict) and isinstance(dec_rec.get('decision'), dict) else {}
        rat_list = dec_obj.get('rationale', []) if isinstance(dec_obj.get('rationale'), list) else []
        rationale_html = list_to_ul(rat_list)

        met_rec = met_map.get(nm, {}) if isinstance(met_map, dict) else {}
        best = (met_rec.get('market') or {}).get('best_serp') if isinstance((met_rec.get('market') or {}), dict) else None
        ev_lines = []
        if isinstance(best, dict) and safe_str(best.get('query_text')).strip():
            rel50 = safe_int(best.get('relevant_count_top50'), None)
            tot = safe_int(best.get('total_estimate'), None)
            q = safe_str(best.get('query_text'))[:120]
            if rel50 is not None:
                ev_lines.append(f"Лучший запрос: ‘{q}’ • rel50={rel50} • total≈{tot if tot is not None else '—'}")
            else:
                ev_lines.append(f"Лучший запрос: ‘{q}’")
        comp_list = met_rec.get('competitors', []) if isinstance(met_rec, dict) and isinstance(met_rec.get('competitors'), list) else []
        ev_lines.append(f"Конкурентов сравнено: {len(comp_list)}")
        if dp_s != '—':
            ev_lines.append(f"Цена: наша {own}₽ vs рынок p50 {comp}₽ (Δ {dp_s})")
        evidence_html = list_to_ul(ev_lines)

        first_task = tasks[0] if tasks else "—"
        more_tasks = (f"<div class='muted'>+{len(tasks)-1} ещё</div>" if len(tasks) > 1 else "")

        row_html.append(f"""
<tr class="row" data-verdict="{html_escape(verdict)}" data-name="{html_escape(name.lower())}" data-nm="{html_escape(nm)}"
    data-price="{cats['price']}" data-seo="{cats['seo']}" data-content="{cats['content']}" data-social="{cats['social']}">
  <td class="mono"><a href="{html_escape(url)}" target="_blank">{html_escape(nm)}</a><div class="muted">{html_escape(vendor_code)}</div></td>
  <td><div class="name">{html_escape(name)}</div></td>
  <td>{verdict_badge(verdict)}</td>
  <td class="num">
    <div><span class="muted">Наша:</span> <b>{html_escape(own)}</b></div>
    <div><span class="muted">Рынок p50:</span> {html_escape(comp)}</div>
    <div class="muted">Δ: {html_escape(dp_s)}</div>
  </td>
  <td class="num">
    <div><span class="muted">Рейт:</span> {html_escape(rating)}</div>
    <div><span class="muted">Отзывы:</span> {html_escape(fb)}</div>
    <div class="muted">Фото: {html_escape(photos)} • Видео: {html_escape(video)}</div>
  </td>
  <td>{tags_short}</td>
  <td class="tasks">{html_escape(first_task)}{more_tasks}</td>
  <td class="details">
    <details>
      <summary>Подробнее</summary>
      <div class="detail-grid">
        <div><div class="h">Основание</div><div class="box">{evidence_html}</div></div>
        <div><div class="h">Почему</div><div class="box">{rationale_html}</div></div>
        <div><div class="h">Риски</div><div class="box">{flags_full}</div></div>
        <div><div class="h">Задачи</div><div class="box">{tasks_html}</div></div>
        <div><div class="h">Конкуренты</div><div class="box"><div class="muted">Количество: {html_escape(comp_cnt)}</div><div class="links">{comp_links_html}</div></div></div>
      </div>
    </details>
  </td>
</tr>
""")

    # -------- top panels --------
    headline = safe_str(llm_sum.get("headline","")) or "WB Revival: сводка"
    overall = safe_str(llm_sum.get("overall",""))
    key_findings = llm_sum.get("key_findings", [])
    recommended = llm_sum.get("recommended_actions", [])
    watchouts = llm_sum.get("watchouts", [])

    title = "WB Revival: отчёт по SKU"
    gen_at = safe_str(exec_summary.get("created_at","")) or utc_now_iso_local()

    cards_html = f"""
 <div class="cards">
   <div class="card"><div class="k">SKU в отчёте</div><div class="v">{len(report_rows)}</div></div>
   <div class="card"><div class="k">Быстрый ревайв</div><div class="v ok">{counts.get("REVIVE_FAST",0)}</div></div>
   <div class="card"><div class="k">Нужна доработка</div><div class="v warn">{counts.get("REVIVE_REWORK",0)}</div></div>
   <div class="card"><div class="k">Новая карточка</div><div class="v info">{counts.get("CLONE_NEW_CARD",0)}</div></div>
   <div class="card"><div class="k">Снять/архив</div><div class="v bad">{counts.get("DROP",0)}</div></div>

   <div class="card"><div class="k">Проблемы с ценой</div><div class="v bad">{cat_counts["price"]}</div><div class="hint">по флагам</div></div>
   <div class="card"><div class="k">Проблемы SEO</div><div class="v warn">{cat_counts["seo"]}</div><div class="hint">по флагам</div></div>
   <div class="card"><div class="k">Проблемы контента</div><div class="v warn">{cat_counts["content"]}</div><div class="hint">по флагам</div></div>
   <div class="card"><div class="k">Отзывы и рейтинг</div><div class="v info">{cat_counts["social"]}</div><div class="hint">по флагам</div></div>

   <div class="card"><div class="k">Валидных запросов</div><div class="v">{fmt_num(run_facts.get("valid_queries_total"), 0)}</div><div class="hint">Stage D</div></div>
   <div class="card"><div class="k">Доля запросов с rel50&gt;0</div><div class="v">{fmt_num(run_facts.get("valid_queries_rel50_gt0_share_pct"), 1)}%</div><div class="hint">Stage D</div></div>
   <div class="card"><div class="k">Карточек конкурентов OK</div><div class="v info">{fmt_num(run_facts.get("competitor_cards_ok"), 0)}</div><div class="hint">Stage F</div></div>
 </div>
 """

    css = """
:root{
  --bg:#f7f8fb; --card:#ffffff; --text:#0f172a; --muted:#64748b; --border:#e2e8f0;
  --accent:#2563eb; --ok:#16a34a; --warn:#d97706; --bad:#dc2626; --info:#0284c7;
}
*{box-sizing:border-box}
body{margin:0;font-family:ui-sans-serif,system-ui,-apple-system,Segoe UI,Roboto,Arial; background:var(--bg); color:var(--text);}
header{position:sticky;top:0;z-index:50;background:rgba(247,248,251,.9);backdrop-filter: blur(8px); border-bottom:1px solid var(--border);}
.header-inner{max-width:1400px;margin:0 auto;padding:14px 18px;display:flex;gap:14px;align-items:center;justify-content:space-between;flex-wrap:wrap}
h1{font-size:18px;margin:0}
.meta{font-size:12px;color:var(--muted)}
.container{max-width:1400px;margin:0 auto;padding:16px 18px}
.cards{display:grid;grid-template-columns:repeat(6,minmax(140px,1fr));gap:10px;margin:10px 0 16px}
.card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:10px 12px;box-shadow:0 1px 0 rgba(15,23,42,.03)}
.card .k{font-size:12px;color:var(--muted)}
.card .v{font-size:22px;margin-top:6px;font-weight:700}
.card .v.ok{color:var(--ok)} .card .v.warn{color:var(--warn)} .card .v.bad{color:var(--bad)} .card .v.info{color:var(--info)}
.card .hint{margin-top:4px;font-size:11px;color:var(--muted)}
.grid{display:grid;grid-template-columns:1.2fr .8fr;gap:12px}
@media(max-width:1000px){.grid{grid-template-columns:1fr}.cards{grid-template-columns:repeat(2,minmax(140px,1fr));}}
.panel{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:12px 14px}
.panel h2{font-size:14px;margin:0 0 8px}
.panel p{margin:6px 0;color:var(--text);line-height:1.35}
.panel .muted{color:var(--muted);font-size:12px}
.badge{display:inline-flex;align-items:center;gap:6px;border-radius:999px;padding:4px 10px;font-size:12px;font-weight:700;border:1px solid transparent;white-space:nowrap}
.b-fast{background:rgba(22,163,74,.08);color:var(--ok);border-color:rgba(22,163,74,.22)}
.b-rework{background:rgba(217,119,6,.10);color:var(--warn);border-color:rgba(217,119,6,.25)}
.b-clone{background:rgba(2,132,199,.10);color:var(--info);border-color:rgba(2,132,199,.22)}
.b-drop{background:rgba(220,38,38,.10);color:var(--bad);border-color:rgba(220,38,38,.22)}
.tag{display:inline-flex;align-items:center;gap:8px;border-radius:999px;padding:3px 9px;font-size:11px;border:1px solid var(--border);background:#fff;margin:2px 4px 2px 0}
.tag .cnt{color:var(--muted);font-weight:700}
.t-price{background:rgba(220,38,38,.06);border-color:rgba(220,38,38,.18)}
.t-seo{background:rgba(37,99,235,.06);border-color:rgba(37,99,235,.18)}
.t-content{background:rgba(2,132,199,.06);border-color:rgba(2,132,199,.18)}
.t-social{background:rgba(217,119,6,.08);border-color:rgba(217,119,6,.22)}
.t-meta{background:rgba(100,116,139,.08)}
.t-other{background:rgba(100,116,139,.08)}
.t-more{background:#f1f5f9}
.controls{display:flex;flex-wrap:wrap;gap:10px;align-items:center;margin:14px 0}
input[type="text"], select{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:9px 10px;font-size:13px;min-width:260px}
.chk{display:flex;align-items:center;gap:8px;font-size:13px;color:var(--text);background:var(--card);border:1px solid var(--border);border-radius:10px;padding:8px 10px}
button{border:1px solid var(--border);background:var(--card);border-radius:10px;padding:9px 10px;font-size:13px;cursor:pointer}
button:hover{border-color:#cbd5e1}
table{width:100%;border-collapse:separate;border-spacing:0;background:var(--card);border:1px solid var(--border);border-radius:12px;overflow:hidden}
th,td{padding:10px 10px;border-bottom:1px solid var(--border);vertical-align:top}
th{font-size:12px;color:var(--muted);text-align:left;background:#f8fafc;position:sticky;top:60px;z-index:10}
tr:last-child td{border-bottom:none}
.mono{font-family:ui-monospace,SFMono-Regular,Menlo,Monaco,Consolas,monospace;font-size:12px}
.num{white-space:nowrap}
.name{font-weight:700}
.muted{color:var(--muted);font-size:12px}
.tasks{max-width:420px}
.details summary{cursor:pointer;color:var(--accent);font-weight:700}
.details summary::-webkit-details-marker{display:none}
.details summary:before{content:"▸";display:inline-block;margin-right:8px;color:var(--muted);transition:transform .15s}
details[open] summary:before{transform:rotate(90deg)}
.detail-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-top:10px}
@media(max-width:1100px){.detail-grid{grid-template-columns:1fr}}
.detail-grid .h{font-size:12px;color:var(--muted);margin-bottom:6px}
.box{background:#f8fafc;border:1px solid var(--border);border-radius:10px;padding:10px}
.box ol{margin:0;padding-left:18px}
.links a{display:inline-block;margin:2px 8px 2px 0}
.footer{margin:16px 0;color:var(--muted);font-size:12px}
.toplist{margin:0;padding-left:18px}
.toplist .cnt{display:inline-block;min-width:24px;color:var(--muted);font-weight:700}
"""

    js = """
const rows = Array.from(document.querySelectorAll('tr.row'));
const shownEl = document.getElementById('shown');

function applyFilters(){
  const term = (document.getElementById('q').value || '').trim().toLowerCase();
  const v = document.getElementById('verdict').value;
  const fPrice = document.getElementById('fPrice').checked;
  const fSeo = document.getElementById('fSeo').checked;
  const fContent = document.getElementById('fContent').checked;
  const fSocial = document.getElementById('fSocial').checked;

  let shown = 0;
  for (const r of rows){
    let ok = true;

    if (term){
      const hay = (r.dataset.nm + ' ' + r.dataset.name + ' ' + r.cells[0].innerText).toLowerCase();
      ok = hay.includes(term);
    }
    if (ok && v !== 'ALL'){
      ok = r.dataset.verdict === v;
    }
    if (ok && fPrice) ok = r.dataset.price === '1';
    if (ok && fSeo) ok = r.dataset.seo === '1';
    if (ok && fContent) ok = r.dataset.content === '1';
    if (ok && fSocial) ok = r.dataset.social === '1';

    r.style.display = ok ? '' : 'none';
    if (ok) shown++;
  }
  shownEl.textContent = shown;
}

function toggleDetails(open){
  document.querySelectorAll('td.details details').forEach(d => d.open = open);
}

document.getElementById('q').addEventListener('input', applyFilters);
document.getElementById('verdict').addEventListener('change', applyFilters);
['fPrice','fSeo','fContent','fSocial'].forEach(id => {
  document.getElementById(id).addEventListener('change', applyFilters);
});
document.getElementById('openAll').addEventListener('click', () => toggleDetails(true));
document.getElementById('closeAll').addEventListener('click', () => toggleDetails(false));

applyFilters();
"""

    # summary panels
    # run facts panel (evidence numbers)
    rf_lines = []
    if isinstance(run_facts, dict):
        def _rf(label, key):
            v = run_facts.get(key)
            if v is None or safe_str(v).strip() == '':
                return
            rf_lines.append(f"{label}: {v}")
        _rf('Валидных запросов (D)', 'valid_queries_total')
        _rf('Запросов с rel50>0 (D)', 'valid_queries_rel50_gt0')
        _rf('Доля rel50>0 (D, %)', 'valid_queries_rel50_gt0_share_pct')
        _rf('Карточек конкурентов OK (F)', 'competitor_cards_ok')
        _rf('Карточек конкурентов FAIL (F)', 'competitor_cards_fail')
        _rf('Финальных сравнений (I)', 'comparisons_total')
    rf_panel = (
        '<section class="panel">'
        '<h2>Доказательная база</h2>'
        + (list_to_ul(rf_lines) if rf_lines else '<div class="muted">Нет данных</div>')
        + '<div class="muted" style="margin-top:8px">Цифры из артефактов пайплайна, не из головы модели.</div>'
        + '</section>'
    )

    summary_panel = f"""
<div class="grid">
  <section class="panel">
    <h2>{html_escape(headline)}</h2>
    <p>{html_escape(overall)}</p>
    <div class="muted">Сгенерировано: {html_escape(gen_at)} • Папка: {html_escape(str(out_dir))}</div>
  </section>

  <section class="panel">
    <h2>Главные проблемы (по флагам)</h2>
    <div>{render_top_flags()}</div>
    <div style="margin-top:10px" class="muted">
      Категории: цена {cat_counts["price"]}, SEO {cat_counts["seo"]}, контент {cat_counts["content"]}, отзывы {cat_counts["social"]}.
    </div>
  </section>
  {rf_panel}
</div>

<div class="grid" style="margin-top:12px">
  <section class="panel">
    <h2>Что нашли</h2>
    {render_bullets(key_findings)}
  </section>
  <section class="panel">
    <h2>Что делать</h2>
    {render_bullets(recommended)}
  </section>
</div>

<div class="grid" style="margin-top:12px">
  <section class="panel">
    <h2>Самые частые задачи</h2>
    {render_top_tasks()}
  </section>
  <section class="panel">
    <h2>На что смотреть (риски)</h2>
    {render_bullets(watchouts)}
  </section>
</div>
"""

    tech = f"""
<details class="panel" style="margin-top:14px">
  <summary style="cursor:pointer;font-weight:800;color:var(--accent)">Технические детали (скрыто, чтобы не пугать людей)</summary>
  <div class="muted" style="margin-top:8px">
    <div><b>run_id:</b> {html_escape(safe_str(exec_summary.get("run_id","")))}</div>
    <div><b>created_at:</b> {html_escape(safe_str(exec_summary.get("created_at","")))}</div>
    <div><b>schema_version:</b> {html_escape(safe_str(exec_summary.get("schema_version","")))}</div>
    <div><b>LLM debug:</b>
      <pre style="white-space:pre-wrap;background:#f8fafc;border:1px solid var(--border);padding:10px;border-radius:10px">{html_escape(json.dumps(exec_summary.get("llm_debug",{}), ensure_ascii=False, indent=2))}</pre>
    </div>
  </div>
</details>
"""

    return f"""<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{html_escape(title)}</title>
<style>{css}</style>
</head>
<body>
<header>
  <div class="header-inner">
    <div>
      <h1>{html_escape(title)}</h1>
      <div class="meta">Показано: <span id="shown">0</span> / {len(report_rows)} • Источники: XLSX REPORT + exec_summary.json</div>
    </div>
    <div class="meta">Ссылки WB открываются в новой вкладке. Потому что так удобнее, чем плакать.</div>
  </div>
</header>

<div class="container">
  {cards_html}
  {summary_panel}

  <div class="controls">
    <input id="q" type="text" placeholder="Поиск: nm_id, vendor_code, название..." />
    <select id="verdict">
      <option value="ALL">Все решения</option>
      <option value="REVIVE_FAST">Быстрый ревайв</option>
      <option value="REVIVE_REWORK">Нужна доработка</option>
      <option value="CLONE_NEW_CARD">Новая карточка</option>
      <option value="DROP">Снять/архив</option>
    </select>

    <label class="chk"><input type="checkbox" id="fPrice"> Только проблемы с ценой</label>
    <label class="chk"><input type="checkbox" id="fSeo"> Только SEO и ключи</label>
    <label class="chk"><input type="checkbox" id="fContent"> Только контент</label>
    <label class="chk"><input type="checkbox" id="fSocial"> Только отзывы и рейтинг</label>

    <button id="openAll" title="Открыть детали для всех строк">Развернуть все</button>
    <button id="closeAll" title="Свернуть детали для всех строк">Свернуть все</button>
  </div>

  <table>
    <thead>
      <tr>
        <th style="width:150px">SKU</th>
        <th>Товар</th>
        <th style="width:160px">Решение</th>
        <th style="width:160px">Цена</th>
        <th style="width:170px">Соцдоказ</th>
        <th style="width:320px">Теги проблем</th>
        <th style="width:420px">Первое действие</th>
        <th style="width:140px">Детали</th>
      </tr>
    </thead>
    <tbody>
      {''.join(row_html)}
    </tbody>
  </table>

  {tech}

  <div class="footer">
    Подсказка: теги проблем показывают человеческие формулировки, но исходные флаги (EN) остаются в подсказках.
  </div>
</div>

<script>{js}</script>
</body>
</html>"""


def html_escape(s: str) -> str:
    s = safe_str(s)
    return (s.replace("&","&amp;")
             .replace("<","&lt;")
             .replace(">","&gt;")
             .replace('"',"&quot;")
             .replace("'","&#39;"))

# =========================
# Integrity checks
# =========================

def check_integrity(out_dir: Path, end_stage: str = "K", *, report_xlsx: Optional[str] = None, report_html: Optional[str] = None) -> None:
    """Basic artifact sanity checks.

    - Stage-aware: checks only files that should exist up to end_stage.
    - Per-SKU JSONL files must contain all nm_id records from the manifest scope (unless the file is optional for that stage).
    """
    manifest_path = out_dir / "run_manifest.json"
    if not manifest_path.exists():
        raise RuntimeError("Missing run_manifest.json (stage A)")

    manifest = read_json(manifest_path)
    scope = manifest["scope"]["sku_list"]
    total = len(scope)
    expected = set(nm_id_to_str(s["nm_id"]) for s in scope)

    n_scope = len(scope)
    if end_stage == "K":
        if not safe_str(report_xlsx or "").strip():
            report_xlsx = f"WB_REVIVE_{n_scope}.xlsx"
        if not safe_str(report_html or "").strip():
            report_html = f"WB_REVIVE_{n_scope}.html"

    stage_to_files = {
        "A": ["run_manifest.json"],
        "B": ["own_norm.jsonl"],
        "C": ["queries_raw.jsonl"],
        "D": ["queries_valid.jsonl"],
        "E": ["competitor_pool.jsonl"],
        "F": ["competitor_norm.jsonl"],  # may be empty but should exist
        "G": ["llm_relevance.jsonl"],
        "H": ["competitors_selected.jsonl"],
        "I": ["comparison_metrics.jsonl"],
        "J": ["decisions.jsonl"],
        "K": ["exec_summary.json"],
    }

    if end_stage not in STAGE_ORDER:
        raise ValueError(f"Invalid end_stage: {end_stage}")

    idx = STAGE_ORDER.index(end_stage)
    required: List[str] = []
    for st in STAGE_ORDER[:idx+1]:
        required.extend(stage_to_files.get(st, []))

    # stage K outputs are parameterized
    if end_stage == "K":
        if report_xlsx:
            required.append(report_xlsx)
        if report_html:
            required.append(report_html)

    missing_files = [f for f in required if not (out_dir / f).exists()]
    if missing_files:
        raise RuntimeError(f"Missing required files up to stage {end_stage}: {missing_files}")

    per_sku_files = [
        "own_norm.jsonl",
        "queries_raw.jsonl",
        "queries_valid.jsonl",
        "competitor_pool.jsonl",
        "llm_relevance.jsonl",
        "competitors_selected.jsonl",
        "comparison_metrics.jsonl",
        "decisions.jsonl",
    ]
    for fn in per_sku_files:
        if fn not in required:
            continue
        path = out_dir / fn
        ids = set(nm_id_to_str(r.get("meta", {}).get("nm_id")) for r in read_jsonl(path))
        miss = sorted(expected - ids)
        if miss:
            raise RuntimeError(f"{fn}: missing nm_id records: {miss[:5]} ... total {len(miss)}")

# =========================
# Stage runner + CLI
# =========================

def pause_between_stages(pause: bool, *, current_stage: str, next_stage: Optional[str], args: argparse.Namespace, llm_caps: Dict[str, int]) -> None:
    """Interactive pause between stages with stage labels + VPN/LLM hints."""
    if not pause:
        return
    if not next_stage:
        return

    title = _stage_title(next_stage)
    net = _stage_network(next_stage)
    llm_on = _stage_llm_enabled(next_stage, args)
    vpn_hint = _stage_vpn_hint(next_stage, args)

    parts = [f"Следующая стадия: [{next_stage}] {title}", f"сеть: {net}", f"LLM: {'да' if llm_on else 'нет'}", f"VPN: {vpn_hint}"]
    if llm_on and next_stage in llm_caps:
        parts.append(f"max_tokens: {llm_caps[next_stage]}")
    msg = " | ".join(parts)

    try:
        ans = input(msg + " | Продолжить? (y/n) > ").strip().lower()
        if ans not in {"y", "yes", "д", "да"}:
            print("Остановлено пользователем.")
            sys.exit(1)
    except KeyboardInterrupt:
        print("\nОстановлено.")
        sys.exit(1)

def run_pipeline(args: argparse.Namespace) -> None:
    out_dir = Path(args.out).resolve()
    ensure_dir(out_dir)

    load_env_file(Path(__file__).with_name(".env"), override=False)
    api_key = ""
    if args.llm_provider == "openai":
        api_key = os.environ.get("OPENAI_API_KEY", "")
    else:
        api_key = os.environ.get("OPENROUTER_API_KEY", "")


    # Check-only mode (no network, no LLM). Useful when you resume or you just want to verify outputs.
    if getattr(args, "check", False):
        upto = safe_str(getattr(args, "check_upto", "") or args.end_stage or "K").strip().upper()
        check_integrity(out_dir, end_stage=upto, report_xlsx=args.report_xlsx, report_html=args.report_html)
        print(f"[CHECK] OK up to stage {upto} in {out_dir}")
        return

    # Stage selection
    start = args.start_stage
    end = args.end_stage
    if start not in STAGE_ORDER or end not in STAGE_ORDER:
        raise ValueError("Invalid stage bounds")
    sidx = STAGE_ORDER.index(start)
    eidx = STAGE_ORDER.index(end)
    stages = STAGE_ORDER[sidx:eidx+1]
    llm_caps = resolve_llm_caps(args)


    # Stage A is special because it creates the manifest
    if "A" in stages:
        stage_A(
            out_dir,
            Path(args.input),
            args.sheet,
            llm_provider=args.llm_provider,
            model_small=args.model_small,
            model_main=args.model_main,
            dests=args.dests,
            search_limit=args.search_limit,
            deep_card_enabled=(not args.no_deep_card),
            expect_count=args.expect_count,
            dedupe=args.dedupe,
            case_like_terms=args.case_like_terms,
        )
        if args.verbose:
            print("[A] manifest created")
        pause_between_stages(args.pause_between_stages, current_stage="A", next_stage=_next_stage_in_run(stages, "A"), args=args, llm_caps=llm_caps)

    if "B" in stages:
        stage_B(out_dir, timeout=args.wb_timeout, sleep_sec=args.sleep, resume=args.resume, verbose=args.verbose, deep_card_enabled=(not args.no_deep_card))
        pause_between_stages(args.pause_between_stages, current_stage="B", next_stage=_next_stage_in_run(stages, "B"), args=args, llm_caps=llm_caps)

    if "C" in stages:
        stage_C(
            out_dir,
            provider=args.llm_provider,
            model=args.model_small,
            api_key=api_key,
            base_url=args.llm_base_url,
            use_llm=args.use_llm,
            llm_timeout=args.llm_timeout,
            max_tokens=llm_caps["C"],
            temperature=args.llm_temp,
            resume=args.resume,
            verbose=args.verbose,
        )
        pause_between_stages(args.pause_between_stages, current_stage="C", next_stage=_next_stage_in_run(stages, "C"), args=args, llm_caps=llm_caps)

    if "D" in stages:
        stage_D(out_dir, timeout=args.wb_timeout, sleep_sec=args.sleep, search_limit=args.search_limit, resume=args.resume, verbose=args.verbose)
        pause_between_stages(args.pause_between_stages, current_stage="D", next_stage=_next_stage_in_run(stages, "D"), args=args, llm_caps=llm_caps)

    if "E" in stages:
        stage_E(out_dir, pool_limit=args.pool_limit, resume=args.resume, verbose=args.verbose)
        pause_between_stages(args.pause_between_stages, current_stage="E", next_stage=_next_stage_in_run(stages, "E"), args=args, llm_caps=llm_caps)

    if "F" in stages:
        stage_F(out_dir, timeout=args.wb_timeout, sleep_sec=args.sleep, resume=args.resume, verbose=args.verbose, deep_card_enabled=(not args.no_deep_card))
        pause_between_stages(args.pause_between_stages, current_stage="F", next_stage=_next_stage_in_run(stages, "F"), args=args, llm_caps=llm_caps)

    if "G" in stages:
        stage_G(
            out_dir,
            provider=args.llm_provider,
            model=args.model_small,
            api_key=api_key,
            base_url=args.llm_base_url,
            use_llm=args.use_llm_relevance,
            llm_timeout=args.llm_timeout,
            max_tokens=llm_caps["G"],
            temperature=args.llm_temp,
            resume=args.resume,
            verbose=args.verbose,
        )
        pause_between_stages(args.pause_between_stages, current_stage="G", next_stage=_next_stage_in_run(stages, "G"), args=args, llm_caps=llm_caps)

    if "H" in stages:
        stage_H(out_dir, max_competitors=args.max_competitors, resume=args.resume, verbose=args.verbose)
        pause_between_stages(args.pause_between_stages, current_stage="H", next_stage=_next_stage_in_run(stages, "H"), args=args, llm_caps=llm_caps)

    if "I" in stages:
        stage_I(out_dir, resume=args.resume, verbose=args.verbose)
        pause_between_stages(args.pause_between_stages, current_stage="I", next_stage=_next_stage_in_run(stages, "I"), args=args, llm_caps=llm_caps)

    if "J" in stages:
        stage_J(
            out_dir,
            provider=args.llm_provider,
            model=args.model_main,
            api_key=api_key,
            base_url=args.llm_base_url,
            use_llm=args.use_llm_verdict,
            llm_timeout=args.llm_timeout,
            max_tokens=llm_caps["J"],
            temperature=args.llm_temp,
            resume=args.resume,
            verbose=args.verbose,
        )
        pause_between_stages(args.pause_between_stages, current_stage="J", next_stage=_next_stage_in_run(stages, "J"), args=args, llm_caps=llm_caps)

    if "K" in stages:
        stage_K(out_dir, xlsx_name=args.report_xlsx, html_name=args.report_html, verbose=args.verbose, llm_provider=args.llm_provider, llm_model=args.model_main, api_key=api_key, llm_base_url=args.llm_base_url, use_llm_exec_summary=args.use_llm_exec_summary, llm_timeout=args.llm_timeout, llm_max_tokens=llm_caps["K"], llm_temperature=args.llm_temp)

    if not args.no_auto_check:
        check_integrity(out_dir, end_stage=end, report_xlsx=args.report_xlsx, report_html=args.report_html)

# =========================
# CLI + Interactive Menu
# =========================

LLM_PRESETS_FALLBACK = {
    "openai": {
        "base_url": "https://api.openai.com/v1",
        "small": [
            ("gpt-4o-mini", "cheap small (queries/relevance)"),
            ("gpt-4o-mini-2024-07-18", "pinned id (if available)"),
            ("gpt-4.1-mini", "alt small (if available)"),
            ("o3-mini", "reasoning small (if available)"),
            ("gpt-4o", "not small, but people insist"),
        ],
        "main": [
            ("gpt-4o", "default main"),
            ("gpt-4.1", "alt main (if available)"),
            ("o3", "reasoning main (if available)"),
            ("gpt-5", "future-ish (if available)"),
            ("gpt-5.2", "even more future-ish (if available)"),
        ],
    },
    "openrouter": {
        "base_url": "https://openrouter.ai/api/v1",
        "small": [
            ("openai/gpt-4o-mini", "cheap small"),
            ("deepseek/deepseek-chat", "DeepSeek chat"),
            ("mistralai/mistral-small", "Mistral small"),
            ("google/gemini-flash-1.5", "Gemini flash"),
            ("qwen/qwen-2.5-72b-instruct", "Qwen instruct"),
        ],
        "main": [
            ("openai/gpt-4o", "default main"),
            ("anthropic/claude-3.5-sonnet", "Claude Sonnet"),
            ("deepseek/deepseek-r1", "DeepSeek reasoning"),
            ("google/gemini-1.5-pro", "Gemini Pro"),
            ("x-ai/grok-2", "Grok"),
        ],
    },
}

ENV_COMPAT_KEYS = {
    "WB_LLM_SMALL": "LLM_MODEL_SMALL",
    "WB_LLM_MAIN": "LLM_MODEL_MAIN",
}

def read_dotenv(path: Path) -> Dict[str, str]:
    if not path.exists():
        return {}
    out: Dict[str, str] = {}
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        s = line.strip()
        if not s or s.startswith("#") or "=" not in s:
            continue
        k, v = s.split("=", 1)
        out[k.strip()] = v.strip()
    for old, new in ENV_COMPAT_KEYS.items():
        if new not in out and old in out:
            out[new] = out[old]
    return out

def write_dotenv(path: Path, kv: Dict[str, str]) -> None:
    existing = read_dotenv(path) if path.exists() else {}
    merged = dict(existing)
    merged.update({k: v for k, v in kv.items() if v is not None})
    keys_order = [
        "LLM_PROVIDER",
        "LLM_BASE_URL",
        "LLM_MODEL_SMALL",
        "LLM_MODEL_MAIN",
        "OPENAI_API_KEY",
        "OPENROUTER_API_KEY",
        "OPENROUTER_HTTP_REFERER",
        "OPENROUTER_X_TITLE",
    ]
    lines = ["# WB Revival .env (generated/updated by menu)"]
    for k in keys_order:
        if k in merged:
            lines.append(f"{k}={merged[k]}")
    for k in sorted(set(merged.keys()) - set(keys_order)):
        lines.append(f"{k}={merged[k]}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")

def _env_get(env: Dict[str, str], k: str, default: str = "") -> str:
    v = env.get(k, "")
    return v if v else default

def _yn(prompt: str, default: bool = False) -> bool:
    d = "y" if default else "n"
    while True:
        s = input(f"{prompt} [y/n] (default {d}): ").strip().lower()
        if not s:
            return default
        if s in ("y", "yes", "д", "да"):
            return True
        if s in ("n", "no", "н", "нет"):
            return False
        print("Нужно y/n.")

def _prompt_str(prompt: str, default: str = "") -> str:
    s = input(f"{prompt} [{default}]: ").strip()
    return s if s else default

def _prompt_int(prompt: str, default: int) -> int:
    while True:
        s = input(f"{prompt} [{default}]: ").strip()
        if not s:
            return default
        try:
            return int(s)
        except Exception:
            print("Нужно целое число.")

def _prompt_float(prompt: str, default: float) -> float:
    while True:
        s = input(f"{prompt} [{default}]: ").strip()
        if not s:
            return default
        try:
            return float(s.replace(",", "."))
        except Exception:
            print("Нужно число (например 0.8).")

def _pick_from_list(title: str, items: List[Tuple[str, str]], current: str) -> str:
    print("\n" + title)
    for i, (mid, desc) in enumerate(items, start=1):
        marker = " (current)" if mid == current else ""
        print(f"  {i}) {mid} — {desc}{marker}")
    print("  0) оставить как есть")
    while True:
        s = input("Выбор: ").strip()
        if s == "0" or s == "":
            return current
        try:
            n = int(s)
            if 1 <= n <= len(items):
                return items[n-1][0]
        except Exception:
            pass
        print("Нужно число из списка.")

def interactive_menu() -> None:
    env_path = Path(__file__).with_name(".env")
    env = read_dotenv(env_path)

    provider = _env_get(env, "LLM_PROVIDER", "openrouter").lower()
    if provider not in ("openai", "openrouter"):
        provider = "openrouter"
    base_url = _env_get(env, "LLM_BASE_URL", LLM_PRESETS_FALLBACK[provider]["base_url"])
    model_small = _env_get(env, "LLM_MODEL_SMALL", LLM_PRESETS_FALLBACK[provider]["small"][0][0])
    model_main = _env_get(env, "LLM_MODEL_MAIN", LLM_PRESETS_FALLBACK[provider]["main"][0][0])

    default_in = DEFAULT_INPUT_XLSX
    default_sheet = DEFAULT_INPUT_SHEET
    default_out = "WB_REVIVAL_RUN"

    stage_desc = {
        "A": "Input + Manifest (LOCAL)",
        "B": "Collect OWN cards (WB)",
        "C": "Generate Queries (rules + optional LLM)",
        "D": "SERP Validate Queries (WB)",
        "E": "Build Competitor Pool (LOCAL)",
        "F": "Collect Competitor Cards (WB)",
        "G": "LLM Relevance Filter (optional LLM)",
        "H": "Select Final 1–5 Competitors (LOCAL)",
        "I": "Compare + Score (LOCAL)",
        "J": "Verdict + Backlog (optional LLM)",
        "K": "Render Reports + Exec Summary (LOCAL + optional LLM)",
    }

    while True:
        print("\nРежимы:")
        print("  1) Полный прогон A..K (WB + LLM + отчёты)")
        print("  2) Только WB-часть A..F (без LLM)")
        print("  3) Только LLM/отчёты G..K (после A..F)")
        print("  4) Настройки LLM (провайдер/модели/.env)")
        print("  5) Проверка целостности (--check)")
        print("  6) Выход")
        print("  7) Пользовательский прогон (выбор start/end стадии)")

        choice = input("Номер: ").strip()
        if choice in ("6", "q", "quit", "exit"):
            return

        if choice not in ("1", "2", "3", "4", "5", "7"):
            print("Нужно число 1..7 (6=выход).")
            continue

        if choice == "4":
            print("\nТекущие LLM настройки:")
            print(f"  provider: {provider}")
            print(f"  base_url: {base_url}")
            print(f"  small:    {model_small}")
            print(f"  main:     {model_main}")

            provider = _pick_from_list(
                "Выбери провайдера:",
                [("openrouter", "OpenRouter (агрегатор)"), ("openai", "OpenAI напрямую")],
                provider,
            )
            base_url = LLM_PRESETS_FALLBACK[provider]["base_url"]

            model_small = _pick_from_list("Модель SMALL (Stage C/G):", LLM_PRESETS_FALLBACK[provider]["small"], model_small)
            model_main = _pick_from_list("Модель MAIN (Stage J/K):", LLM_PRESETS_FALLBACK[provider]["main"], model_main)

            env_updates = {
                "LLM_PROVIDER": provider,
                "LLM_BASE_URL": base_url,
                "LLM_MODEL_SMALL": model_small,
                "LLM_MODEL_MAIN": model_main,
            }

            if provider == "openai":
                if _yn("Вбить/обновить OPENAI_API_KEY?", default=False):
                    env_updates["OPENAI_API_KEY"] = _prompt_str("OPENAI_API_KEY", _env_get(env, "OPENAI_API_KEY", ""))
            else:
                if _yn("Вбить/обновить OPENROUTER_API_KEY?", default=False):
                    env_updates["OPENROUTER_API_KEY"] = _prompt_str("OPENROUTER_API_KEY", _env_get(env, "OPENROUTER_API_KEY", ""))
                if _yn("Указать OPENROUTER_HTTP_REFERER?", default=False):
                    env_updates["OPENROUTER_HTTP_REFERER"] = _prompt_str("OPENROUTER_HTTP_REFERER", _env_get(env, "OPENROUTER_HTTP_REFERER", ""))
                if _yn("Указать OPENROUTER_X_TITLE?", default=False):
                    env_updates["OPENROUTER_X_TITLE"] = _prompt_str("OPENROUTER_X_TITLE", _env_get(env, "OPENROUTER_X_TITLE", ""))

            write_dotenv(env_path, env_updates)
            env = read_dotenv(env_path)
            print("Сохранено в .env.")
            continue

        # Common run settings
        out_dir = _prompt_str("Папка вывода (--out)", default_out)
        verbose = _yn("Verbose лог (--verbose)?", default=True)
        resume = _yn("Resume (--resume)?", default=True)

        proxy = _prompt_str("Proxy для WB (--proxy), пусто = без прокси", "")
        no_env_proxy = False
        if proxy:
            no_env_proxy = _yn("Игнорировать HTTP_PROXY/HTTPS_PROXY из env (--no-env-proxy)?", default=False)

        pause = _yn("Пауза между стадиями? (покажет название следующей стадии + LLM/VPN) (--pause-between-stages)?", default=False)
        deep_card = _yn("Собирать расширенную карточку (wbbasket card.json) (Stage B/F)?", default=True)

        argv = ["--out", out_dir]
        if verbose:
            argv.append("--verbose")
        if resume:
            argv.append("--resume")
        if pause:
            argv.append("--pause-between-stages")
        if not deep_card:
            argv.append("--no-deep-card")
        if proxy:
            argv += ["--proxy", proxy]
        if no_env_proxy:
            argv.append("--no-env-proxy")

        if choice in ("1", "2"):
            in_path = _prompt_str("Input XLSX (--input)", default_in)
            sheet = _prompt_str("Sheet (--sheet)", default_sheet)
            argv += ["--input", in_path, "--sheet", sheet]

        if choice == "1":
            use_llm = _yn("Использовать LLM для запросов? [C] Generate Queries (LLM API, VPN может понадобиться) (--use-llm)?", default=True)
            use_rel = _yn("Использовать LLM релевантность? [G] LLM Relevance Filter (LLM API, VPN может понадобиться) (--use-llm-relevance)?", default=True)
            use_ver = _yn("Использовать LLM для вердикта? [J] Verdict + Backlog (LLM API, VPN может понадобиться) (--use-llm-verdict)?", default=True)
            use_sum = _yn("Использовать LLM для exec summary? [K] Executive Summary (LLM API, VPN может понадобиться) (--use-llm-exec-summary)?", default=False)

            if use_llm:
                argv.append("--use-llm")
            if use_rel:
                argv.append("--use-llm-relevance")
            if use_ver:
                argv.append("--use-llm-verdict")
            if use_sum:
                argv.append("--use-llm-exec-summary")

            if use_llm or use_rel or use_ver or use_sum:
                print("\nЛимиты LLM max_tokens по стадиям (оставь дефолты, если не хочешь думать):")
                tok_c = _prompt_int("Stage C (queries)", 800)
                tok_g = _prompt_int("Stage G (relevance)", 3000)
                tok_j = _prompt_int("Stage J (verdict)", 6000)
                tok_k = _prompt_int("Stage K (exec summary)", 15000)
                argv += [
                    "--llm-max-tokens-c", str(tok_c),
                    "--llm-max-tokens-g", str(tok_g),
                    "--llm-max-tokens-j", str(tok_j),
                    "--llm-max-tokens-k", str(tok_k),
                ]

            argv += ["--start-stage", "A", "--end-stage", "K"]

        elif choice == "2":
            sleep = _prompt_float("Sleep между WB-запросами (--sleep)", 0.8)
            argv += ["--sleep", str(sleep), "--start-stage", "A", "--end-stage", "F"]

        elif choice == "3":
            use_rel = _yn("Использовать LLM релевантность? [G] LLM Relevance Filter (LLM API, VPN может понадобиться) (--use-llm-relevance)?", default=True)
            use_ver = _yn("Использовать LLM для вердикта? [J] Verdict + Backlog (LLM API, VPN может понадобиться) (--use-llm-verdict)?", default=True)
            use_sum = _yn("Использовать LLM для exec summary? [K] Executive Summary (LLM API, VPN может понадобиться) (--use-llm-exec-summary)?", default=False)

            if use_rel:
                argv.append("--use-llm-relevance")
            if use_ver:
                argv.append("--use-llm-verdict")
            if use_sum:
                argv.append("--use-llm-exec-summary")

            if use_rel or use_ver or use_sum:
                print("\nЛимиты LLM max_tokens по стадиям (оставь дефолты, если не хочешь думать):")
                tok_g = _prompt_int("Stage G (relevance)", 3000)
                tok_j = _prompt_int("Stage J (verdict)", 6000)
                tok_k = _prompt_int("Stage K (exec summary)", 15000)
                argv += ["--llm-max-tokens-g", str(tok_g), "--llm-max-tokens-j", str(tok_j), "--llm-max-tokens-k", str(tok_k)]

            argv += ["--start-stage", "G", "--end-stage", "K"]

        elif choice == "5":
            upto = _prompt_str("Проверять до стадии (--check-upto)", "K").upper().strip()
            if upto not in STAGE_ORDER:
                upto = "K"
            argv += ["--check", "--check-upto", upto]

        elif choice == "7":
            print("\nДоступные стадии (A..K):")
            for s in STAGE_ORDER:
                print(f"  {s} — {stage_desc.get(s, '')}")

            while True:
                start_stage = _prompt_str("Start stage (--start-stage)", "A").upper().strip()
                if start_stage not in STAGE_ORDER:
                    print("Нужна буква A..K.")
                    continue
                end_stage = _prompt_str("End stage (--end-stage)", "K").upper().strip()
                if end_stage not in STAGE_ORDER:
                    print("Нужна буква A..K.")
                    continue
                if STAGE_ORDER.index(end_stage) < STAGE_ORDER.index(start_stage):
                    print("End stage не может быть раньше start stage.")
                    continue
                break

            stages = STAGE_ORDER[STAGE_ORDER.index(start_stage):STAGE_ORDER.index(end_stage) + 1]

            if "A" in stages:
                in_path = _prompt_str("Input XLSX (--input)", default_in)
                sheet = _prompt_str("Sheet (--sheet)", default_sheet)
                argv += ["--input", in_path, "--sheet", sheet]

            if set(stages) & {"B", "D", "F"}:
                sleep = _prompt_float("Sleep между WB-запросами (--sleep)", 0.8)
                argv += ["--sleep", str(sleep)]

            if "D" in stages:
                sl = _prompt_int("SERP search limit (--search-limit)", 100)
                argv += ["--search-limit", str(sl)]

            if "E" in stages:
                pl = _prompt_int("Pool limit после hard filters (--pool-limit)", 30)
                argv += ["--pool-limit", str(pl)]
            if "H" in stages:
                mc = _prompt_int("Max competitors per SKU (--max-competitors)", 12)
                argv += ["--max-competitors", str(mc)]

            use_llm = False
            use_rel = False
            use_ver = False
            use_sum = False

            if "C" in stages:
                use_llm = _yn("Использовать LLM для запросов? [C] (--use-llm)", default=False)
                if use_llm:
                    argv.append("--use-llm")
            if "G" in stages:
                use_rel = _yn("Использовать LLM релевантность? [G] (--use-llm-relevance)", default=True)
                if use_rel:
                    argv.append("--use-llm-relevance")
            if "J" in stages:
                use_ver = _yn("Использовать LLM для вердикта? [J] (--use-llm-verdict)", default=True)
                if use_ver:
                    argv.append("--use-llm-verdict")
            if "K" in stages:
                use_sum = _yn("Использовать LLM для exec summary? [K] (--use-llm-exec-summary)", default=False)
                if use_sum:
                    argv.append("--use-llm-exec-summary")

            cap_args = []
            if use_llm and "C" in stages:
                cap_args += ["--llm-max-tokens-c", str(_prompt_int("Stage C max_tokens", 800))]
            if use_rel and "G" in stages:
                cap_args += ["--llm-max-tokens-g", str(_prompt_int("Stage G max_tokens", 3000))]
            if use_ver and "J" in stages:
                cap_args += ["--llm-max-tokens-j", str(_prompt_int("Stage J max_tokens", 6000))]
            if use_sum and "K" in stages:
                cap_args += ["--llm-max-tokens-k", str(_prompt_int("Stage K max_tokens", 15000))]
            if cap_args:
                argv += cap_args

            argv += ["--start-stage", start_stage, "--end-stage", end_stage]

        argv += ["--llm-provider", provider, "--llm-base-url", base_url, "--model-small", model_small, "--model-main", model_main]

        print("\nКоманда (для истории/копипасты):")
        print("python", SCRIPT_NAME, " ".join([shlex.quote(a) for a in argv]))
        print("-" * 78)

        try:
            main(argv)
        except SystemExit as e:
            if getattr(e, "code", 0) not in (0, None):
                print(f"[MENU] Завершилось с кодом {e.code}")
        except Exception as e:
            print(f"[MENU] Ошибка: {e}")

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog=SCRIPT_NAME,
        description="WB Revival V1.5 staged pipeline (dynamic scope + deep card.json) with canonical JSONL artifacts + model presets menu",
    )
    p.add_argument("--input", default=DEFAULT_INPUT_XLSX, help="Input XLSX (must contain a sheet with nm_id column; any row count)")
    p.add_argument("--sheet", default=DEFAULT_INPUT_SHEET, help="Sheet name to read SKU scope from")
    p.add_argument("--expect-count", type=int, default=None, help="Optional: enforce exact SKU row count in the sheet")
    p.add_argument("--dedupe", action="store_true", help="If duplicate nm_id rows exist, keep first occurrence instead of failing")
    p.add_argument("--out", default="WB_REVIVAL_RUN", help="Output directory")

    p.add_argument("--menu", action="store_true", help="Interactive menu (also auto-starts if no args)")

    p.add_argument("--start-stage", default="A", choices=STAGE_ORDER, help="Start stage (A..K)")
    p.add_argument("--end-stage", default="K", choices=STAGE_ORDER, help="End stage (A..K)")

    p.add_argument("--dests", nargs="+", type=int, default=DEFAULT_DESTS, help="WB dests list")
    p.add_argument("--search-limit", type=int, default=100, help="SERP fetch limit per query")

    p.add_argument("--no-deep-card", action="store_true", help="Disable fetching deep wbbasket card.json (more description/attributes) in stages B/F")

    p.add_argument("--wb-timeout", type=int, default=60, help="WB request timeout seconds")
    p.add_argument("--llm-timeout", type=int, default=60, help="LLM request timeout seconds")
    p.add_argument("--sleep", type=float, default=0.8, help="Sleep between WB requests")

    p.add_argument("--resume", action="store_true", help="Resume from existing JSONL outputs")
    p.add_argument("--pause-between-stages", action="store_true", help="Ask y/n between stages (for VPN switching)")

    p.add_argument("--proxy", default="", help="HTTP proxy for WB requests (optional)")
    p.add_argument("--no-env-proxy", action="store_true", help="Ignore env proxy variables for WB requests")

    p.add_argument("--use-llm", action="store_true", help="Use LLM for query enrichment (stage C)")
    p.add_argument("--use-llm-relevance", action="store_true", help="Use LLM for competitor relevance (stage G)")
    p.add_argument("--use-llm-verdict", action="store_true", help="Use LLM for verdict/backlog (stage J)")
    p.add_argument("--use-llm-exec-summary", action="store_true", help="Use LLM to write executive summary (stage K -> exec_summary.json)")

    p.add_argument("--check", action="store_true", help="Only run integrity checks and exit")
    p.add_argument("--check-upto", "--check_upto", dest="check_upto", default="", choices=STAGE_ORDER,
                   help="For --check: verify files up to this stage (default: --end-stage)")
    p.add_argument("--no-auto-check", action="store_true", help="Disable automatic integrity check after pipeline run")

    p.add_argument("--llm-provider", choices=["openai", "openrouter"], default="")
    p.add_argument("--llm-base-url", default="", help="Root or full endpoint for Chat Completions (overrides env)")
    p.add_argument("--model-small", default="", help="Small/cheap LLM model (stage C/G). Empty -> env/preset default.")
    p.add_argument("--model-main", default="", help="Main LLM model (stage J/K). Empty -> env/preset default.")
    # LLM output caps (max_tokens). Per-stage defaults are tuned for this pipeline:
    # - C (queries): small JSON, cheap
    # - G (relevance): can be larger JSON list
    # - J (verdict): structured JSON with backlog/risk flags
    # - K (exec summary): can be long if enabled
    p.add_argument("--llm-max-tokens", type=int, default=0,
               help="Legacy/global max_tokens for ALL LLM stages (overrides per-stage caps if set > 0)")
    p.add_argument("--llm-max-tokens-c", type=int, default=800, help="Stage C max_tokens (query enrichment)")
    p.add_argument("--llm-max-tokens-g", type=int, default=3000, help="Stage G max_tokens (relevance filter JSON)")
    p.add_argument("--llm-max-tokens-j", type=int, default=6000, help="Stage J max_tokens (verdict/backlog JSON)")
    p.add_argument("--llm-max-tokens-k", type=int, default=15000, help="Stage K max_tokens (exec summary JSON, if enabled)")
    p.add_argument("--llm-temp", type=float, default=0.2)

    p.add_argument("--case-like-terms", nargs="+", default=[], help="Override case-like keywords for hard filter (stage E). If empty uses defaults.")
    p.add_argument("--pool-limit", type=int, default=30, help="Initial competitor pool after hard filters")
    p.add_argument("--max-competitors", type=int, default=12, help="Final competitors per SKU")

    # Canon file names, as promised in the spec
    p.add_argument("--report-xlsx", default="", help="Report XLSX name. Empty -> auto WB_REVIVE_<N>.xlsx")
    p.add_argument("--report-html", default="", help="Report HTML name. Empty -> auto WB_REVIVE_<N>.html")

    p.add_argument("--verbose", action="store_true", help="Verbose logs")
    return p

def _resolve_llm_config(args: argparse.Namespace) -> Tuple[str, str, str, str]:
    env = read_dotenv(Path(__file__).with_name(".env"))
    provider = safe_str(getattr(args, "llm_provider", "")).strip().lower() or _env_get(env, "LLM_PROVIDER", "openrouter").lower()
    if provider not in ("openai", "openrouter"):
        provider = "openrouter"

    base_url = safe_str(getattr(args, "llm_base_url", "")).strip() or _env_get(env, "LLM_BASE_URL", LLM_PRESETS_FALLBACK[provider]["base_url"])
    model_small = safe_str(getattr(args, "model_small", "")).strip() or _env_get(env, "LLM_MODEL_SMALL", LLM_PRESETS_FALLBACK[provider]["small"][0][0])
    model_main = safe_str(getattr(args, "model_main", "")).strip() or _env_get(env, "LLM_MODEL_MAIN", LLM_PRESETS_FALLBACK[provider]["main"][0][0])

    if provider == "openrouter" and "/" not in model_small and model_small.startswith("gpt-"):
        model_small = "openai/" + model_small
    if provider == "openrouter" and "/" not in model_main and model_main.startswith("gpt-"):
        model_main = "openai/" + model_main

    return provider, base_url, model_small, model_main

def main(argv: Optional[List[str]] = None) -> None:
    parser = build_parser()
    if argv is None:
        argv = sys.argv[1:]

    if not argv:
        interactive_menu()
        return

    args = parser.parse_args(argv)

    if getattr(args, "menu", False):
        interactive_menu()
        return

    configure_wb_network(getattr(args, "proxy", ""), no_env_proxy=getattr(args, "no_env_proxy", False))

    provider, base_url, model_small, model_main = _resolve_llm_config(args)
    args.llm_provider = provider
    args.llm_base_url = base_url
    args.model_small = model_small
    args.model_main = model_main

    if getattr(args, "check", False):
        out_dir = Path(args.out).resolve()
        upto = getattr(args, "check_upto", "") or getattr(args, "end_stage", "K")
        check_integrity(out_dir, end_stage=upto, report_xlsx=args.report_xlsx, report_html=args.report_html)
        print(f"[CHECK] OK up to stage {upto} in {out_dir}")
        return

    run_pipeline(args)

if __name__ == "__main__":
    main()
