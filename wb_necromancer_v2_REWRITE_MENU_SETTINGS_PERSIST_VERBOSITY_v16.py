# -*- coding: utf-8 -*-
"""
WB Revival v2 "Некромант" — переписанная версия (one-file, но без колхоза).

Важно:
- Комменты тут на "быдло-русском" как просили. Да, я тоже страдаю.
- Стадии A–M: вход -> own fetch -> intent -> queries -> SERP -> pool -> lite -> relevance -> pulse -> supply -> verdicts -> decisions -> reports
- Каноничные артефакты: JSONL + сырой кэш в .wb_cache + финал XLSX + HTML

Если ты это читаешь потому что скрипт упал — поздравляю, ты нашёл очередной угол вселенной, где всё ломается.
"""

from __future__ import annotations

import argparse
import dataclasses
from dataclasses import dataclass, asdict, fields
import getpass
import hashlib
import html as _html
import json
import math
import os
import random
import re
import sys
import traceback
import time
from datetime import datetime, timezone, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# deps
try:
    import requests
except Exception:
    print("Нужен requests. Поставь: pip install requests", file=sys.stderr)
    raise

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl import load_workbook
except Exception:
    print("Нужен openpyxl. Поставь: pip install openpyxl", file=sys.stderr)
    raise


# =========================
# Версии, константы, дефолты
# =========================

SCRIPT_NAME = Path(__file__).name
SCRIPT_VERSION = "2.0.11-rewrite-2026-02-23-reviews-v1get-logfix"
SCHEMA_VERSION = "2.0.11-rewrite-2026-02-23-reviews-v1get-logfix"

FOCUS_TYPE = "tpu_pocket"   # да, мы некроманты, но чёткие
CLUSTERS = ("phone", "type")

ALLOWED_VERDICTS = ("REVIVE_FAST", "REVIVE_REWORK", "CLONE_NEW_CARD", "DROP")

DEFAULT_DESTS = [-1257786, -1216601, -115136, -421732, 123585595]
DEFAULT_SEARCH_HOSTS = ["u-search.wb.ru", "search.wb.ru"]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "ru,en-US;q=0.9,en;q=0.8",
}

# Дефолтные термины. Списки грузим ОДИН раз при старте и дальше не трогаем. Иначе будет ад.
DEFAULT_LEXICON = {
    "case_intent_terms": ["чехол", "чехол для", "накладка", "бампер", "кейс", "case"],
    "tpu_terms": ["tpu", "силикон", "силиконовый", "термополиуретан", "soft", "soft touch"],
    "pocket_terms": ["карман", "картхолдер", "карт холдер", "держатель карт", "для карт",
                     "cardholder", "card holder", "wallet", "кошелек", "кошелёк", "отделение для карт"],
    "ban_terms_default": [
        "стекло", "пленка", "плёнка", "защитное стекло", "гидрогель",
        "зарядка", "кабель", "наушники", "ремешок", "часы",
        "бампер для мебели", "держатель в авто", "крепление",
    ],
}

DEFAULT_PULSE_RULES = {
    # из пульса рынка делаем ALIVE/SLOW/DEAD. Если будешь тюнить — тюнить здесь/в манифесте.
    "alive_r30_med_gte": 1.0,
    "alive_days_since_last_min_lte": 21,
    "alive_r90_med_gte": 3.0,
    "alive_days_since_last_med_lte": 45,
    "slow_r90_med_gte": 1.0,
    "slow_days_since_last_med_lte": 90,
    "slow_days_since_last_min_lte": 120,
}

DEFAULT_SUPPLY_THRESHOLDS = {
    "min_n": 8,
    "dumping": {
        "min_unique_sellers": 10,
        "p10_ratio_lt": 0.75,
        "outlier_rate_gte": 0.10,
        "min_price_n_or": 12,
    },
    "monopoly": {
        "top1_share_gte": 0.35,
        "hhi_gte": 0.25,
    },
}

DEFAULT_REVIEWS_CFG = {
    # Основной (и самый стабильный) публичный источник: feedbacks1/2.
    # Он отдаёт свежие отзывы и даты без seller API.
    "mode": "v1_get",  # v1_get | post (legacy)
    "v1_urls": [
        "https://feedbacks1.wb.ru/feedbacks/v1/{imt_id}",
        "https://feedbacks2.wb.ru/feedbacks/v1/{imt_id}",
    ],
    # Для v1_get пробуем пагинацию через take/skip (если сервер её игнорирует — это поймаем по дублям).
    "v1_take": 100,
    "v1_max_pages": 20,
    "v1_max_skip": 5000,

    # Legacy POST endpoint (часто 403/anti-bot). Оставлен как fallback по желанию.
    "allow_post_fallback": True,
    "base_urls": [
        "https://public-feedbacks.wildberries.ru/api/v1/feedbacks/site",
        "https://public-feedbacks.wildberries.ru/api/v1/feedbacks",
    ],
    "post_take": 30,
    "post_max_pages": 40,
    "post_max_skip": 1000,

    "early_stop_days": 90,
    "cache_ttl_hours": 24,
}

DEFAULT_DECISION_CFG = {
    # токсичная карма: при достаточном числе отзывов и низком рейтинге лучше клонировать
    "karma_min_feedbacks": 30,
    "karma_min_rating": 3.7,
    # если общий рынок модели жив, а тип (TPU+карман) мёртв — можно подсказать альтернативу
    "alt_strategy_on_type_dead": True,
}


# Модельные пресеты: дешёвка, чтоб не разориться.
MODEL_PRESETS = {
    "openai": {
        "cheap": ["gpt-5-mini", "gpt-4.1-mini", "gpt-5-nano", "gpt-4.1-nano"],
        "default": "gpt-5-mini",
    },
    "openrouter": {
        "cheap": [
            "openai/gpt-5-mini", "openai/gpt-4.1-mini", "openai/gpt-5-nano", "openai/gpt-4.1-nano",
            "deepseek/deepseek-chat", "google/gemini-2.0-flash-001", "anthropic/claude-3.5-haiku",
            "meta-llama/llama-3.1-8b-instruct",
        ],
        "default": "openai/gpt-5-mini",
    }
}

STAGE_ORDER = list("ABCDEFGHIJKLM")

# Описание стадий. Тут без поэзии: что делаем, нужна ли сеть WB, нужен ли LLM, и когда трогать VPN.
# VPN hint: WB лучше гонять БЕЗ VPN (РФ/IP), LLM обычно наоборот требует VPN (или хотя бы не РФ).
STAGE_META = {
    "A": {"title": "Вход + манифест (scope/run_id/лексикон)", "network": "LOCAL", "llm_flag": None, "vpn": "ANY"},
    "B": {"title": "Сбор своей карточки (v4/v1 + deep card.json)", "network": "WB", "llm_flag": None, "vpn": "OFF"},
    "C": {"title": "Извлечение интента (модель телефона, must/ban, карма)", "network": "LOCAL", "llm_flag": None, "vpn": "ANY"},
    "D": {"title": "Запросы для 2 рынков: rules + опц LLM-обогащение", "network": "LOCAL", "llm_flag": "use_llm_d", "vpn": "ON"},
    "E": {"title": "SERP: снять выдачу и провалидировать запросы", "network": "WB", "llm_flag": None, "vpn": "OFF"},
    "F": {"title": "Пул конкурентов (лидеры + ближайшие + диверсификация)", "network": "LOCAL", "llm_flag": None, "vpn": "ANY"},
    "G": {"title": "Конкуренты: lite-фетч карточек (цены/остатки/продавцы)", "network": "WB", "llm_flag": None, "vpn": "OFF"},
    "H": {"title": "Релевантность конкурентов: rules + опц LLM для пограничных", "network": "LOCAL", "llm_flag": "use_llm_h", "vpn": "ON"},
    "I": {"title": "Market pulse: отзывы (скорость/давность) + кэш", "network": "WB", "llm_flag": None, "vpn": "OFF"},
    "J": {"title": "Supply/Structure: цены, стоки, концентрация продавцов", "network": "LOCAL", "llm_flag": None, "vpn": "ANY"},
    "K": {"title": "Вердикт кластеров: ALIVE/SLOW/DEAD + confidence", "network": "LOCAL", "llm_flag": None, "vpn": "ANY"},
    "L": {"title": "Финал по SKU: матрица решений + флаги + бэклог", "network": "LOCAL", "llm_flag": None, "vpn": "ANY"},
    "M": {"title": "Отчёты XLSX/HTML + опц LLM-выжимка (exec summary)", "network": "LOCAL", "llm_flag": "use_llm_m", "vpn": "ON"},
}


STAGE_IO = {
    "A": {"in": "INPUT XLSX", "out": "run_manifest.json"},
    "B": {"in": "run_manifest.json", "out": "own_norm.jsonl (+ .wb_cache/deep_*.json)"},
    "C": {"in": "own_norm.jsonl", "out": "intent.jsonl"},
    "D": {"in": "intent.jsonl", "out": "queries_raw.jsonl"},
    "E": {"in": "queries_raw.jsonl", "out": "queries_valid.jsonl (+ .wb_cache/serp/*.json)"},
    "F": {"in": "queries_valid.jsonl", "out": "competitor_pool.jsonl"},
    "G": {"in": "competitor_pool.jsonl", "out": "competitor_lite.jsonl (+ .wb_cache/cards/*.json)"},
    "H": {"in": "competitor_lite.jsonl + intent.jsonl", "out": "relevance.jsonl"},
    "I": {"in": "relevance.jsonl + competitor_lite.jsonl", "out": "market_pulse.jsonl (+ .wb_cache/reviews/*.json)"},
    "J": {"in": "market_pulse.jsonl + relevance.jsonl", "out": "supply_structure.jsonl"},
    "K": {"in": "supply_structure.jsonl", "out": "cluster_verdicts.jsonl"},
    "L": {"in": "cluster_verdicts.jsonl + intent.jsonl", "out": "decisions.jsonl"},
    "M": {"in": "decisions.jsonl", "out": "WB_NECROMANCER_REPORT.xlsx + WB_NECROMANCER_REPORT.html"},
}

def _stage_title(code: str) -> str:
    return STAGE_META.get(code, {}).get("title") or code

def _stage_base_network(code: str) -> str:
    return STAGE_META.get(code, {}).get("network") or "LOCAL"

def _stage_llm_flag(code: str) -> Optional[str]:
    return STAGE_META.get(code, {}).get("llm_flag")

def _stage_llm_enabled(code: str, args: argparse.Namespace) -> bool:
    flag = _stage_llm_flag(code)
    if not flag:
        return False
    return bool(getattr(args, flag, False))

def _stage_effective_network(code: str, args: argparse.Namespace) -> str:
    net = _stage_base_network(code)
    # D/H/M считаются "LOCAL", но если флаг включён — это LLM стадия по факту.
    if net == "LOCAL" and _stage_llm_enabled(code, args):
        return "LLM"
    return net

def _stage_vpn_hint(code: str, args: argparse.Namespace) -> str:
    net = _stage_effective_network(code, args)
    if net == "WB":
        return "VPN: OFF (WB, иначе может послать)"
    if net == "LLM":
        return "VPN: ON (LLM API, иначе может не работать)"
    return "VPN: ANY (локалка)"

def print_stage_table(args: Optional[argparse.Namespace] = None) -> None:
    # Если args нет — считаем, что LLM флаги выключены (чтобы не пугать людей).
    fake = args or argparse.Namespace(use_llm_d=False, use_llm_h=False, use_llm_m=False)
    print("")
    print("Стадии (A..M):")
    for c in STAGE_ORDER:
        base = _stage_base_network(c)
        flag = _stage_llm_flag(c)
        if flag:
            llm_tag = f"LLM: опц ({flag})" if not bool(getattr(fake, flag, False)) else f"LLM: ON ({flag})"
            net = _stage_effective_network(c, fake)
        else:
            llm_tag = "LLM: нет"
            net = base
        vpn = _stage_vpn_hint(c, fake)
        print(f"  {c}) {_stage_title(c)}")
        print(f"     net={net} | {llm_tag} | {vpn}")



# =========================
# Утилиты, чтоб не повторять одно и то же как идиот
# =========================

def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()

def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def safe_str(x: Any, default: str = "") -> str:
    if x is None:
        return default
    try:
        return str(x)
    except Exception:
        return default

def safe_int(x: Any, default: Optional[int] = None) -> Optional[int]:
    try:
        s = safe_str(x, "").strip()
        if not s:
            return default
        if re.fullmatch(r"-?\d+(\.0+)?", s):
            return int(float(s))
        return int(s)
    except Exception:
        return default

def safe_float(x: Any, default: Optional[float] = None) -> Optional[float]:
    try:
        s = safe_str(x, "").strip()
        if not s:
            return default
        v = float(s)
        if not math.isfinite(v):
            return default
        return v
    except Exception:
        return default

def nm_id_to_str(x: Any) -> str:
    # Excel любит превращать цифры в научную нотацию. Мы не любим Excel.
    if x is None or isinstance(x, bool):
        return ""
    if isinstance(x, int):
        return str(x)
    if isinstance(x, float):
        if math.isfinite(x):
            return str(int(x))
        return safe_str(x, "").strip()
    s = safe_str(x, "").strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+(\.0+)?", s):
        try:
            return str(int(Decimal(s)))
        except Exception:
            return str(int(float(s)))
    if re.fullmatch(r"\d+(\.\d+)?[eE][\+\-]?\d+", s):
        try:
            return str(int(Decimal(s)))
        except (InvalidOperation, ValueError, OverflowError):
            try:
                return str(int(float(s)))
            except Exception:
                return s
    return s

def sha1_hex(s: str) -> str:
    # никаких usedforsecurity, потому что это не крипта, это просто отпечаток
    return hashlib.sha1(s.encode("utf-8")).hexdigest()

def sha1_short(s: str, n: int = 12) -> str:
    return sha1_hex(s)[: max(4, int(n))]

def write_json(path: Path, obj: Any) -> None:
    ensure_dir(path.parent)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)

def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))

def append_jsonl(path: Path, obj: dict) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")

def iter_jsonl(path: Path) -> List[dict]:
    if not path.exists():
        return []
    out: List[dict] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                if isinstance(j, dict):
                    out.append(j)
            except Exception:
                continue
    return out

def read_jsonl_done_ids(path: Path, meta_key: str = "nm_id") -> set:
    done = set()
    if not path.exists():
        return done
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                j = json.loads(line)
                meta = j.get("meta") or {}
                k = nm_id_to_str(meta.get(meta_key))
                if k:
                    done.add(k)
            except Exception:
                continue
    return done

def mask_secret(s: str, keep: int = 4) -> str:
    s = safe_str(s)
    if len(s) <= keep:
        return "*" * len(s)
    return "*" * (len(s) - keep) + s[-keep:]


# =========================
# .env (минималка, чтобы не ныть)
# =========================

def load_env_file(env_path: Path, override: bool = False) -> None:
    # мини-dotenv: без фанатизма
    try:
        if not env_path.exists():
            return
        for raw in env_path.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            k = k.strip()
            v = v.strip().strip('"').strip("'")
            if not k:
                continue
            if (not override) and os.getenv(k) not in (None, ""):
                continue
            os.environ[k] = v
    except Exception:
        return

def write_dotenv(env_path: Path, kv: Dict[str, str]) -> None:
    # обновляем или добавляем ключи, комменты оставляем (если повезёт)
    try:
        lines = env_path.read_text(encoding="utf-8").splitlines() if env_path.exists() else []
    except Exception:
        lines = []
    out_lines: List[str] = []
    seen = set()
    for raw in lines:
        s = raw.strip()
        if not s or s.startswith("#") or "=" not in s:
            out_lines.append(raw)
            continue
        k, _ = s.split("=", 1)
        k = k.strip()
        if k in kv:
            out_lines.append(f"{k}={kv[k]}")
            seen.add(k)
        else:
            out_lines.append(raw)
    for k, v in kv.items():
        if k not in seen:
            out_lines.append(f"{k}={v}")
    env_path.write_text("\n".join(out_lines).rstrip() + "\n", encoding="utf-8")

# Грузим .env сразу, чтобы потом не плакать
try:
    load_env_file(Path(__file__).with_name(".env"), override=False)
    load_env_file(Path.cwd() / ".env", override=False)
except Exception:
    pass


# =========================
# Прогресс-бар без tqdm (чтоб не ставить лишнее)
# =========================

class SimpleProgress:
    def __init__(self, total: int, desc: str = ""):
        self.total = max(int(total), 0)
        self.desc = desc
        self.n = 0
        self.postfix: Dict[str, Any] = {}
        self._last_len = 0

    def set_postfix(self, d: Dict[str, Any]):
        self.postfix = d or {}

    def update(self, inc: int = 1):
        self.n += int(inc)
        self._render()

    def _render(self):
        if not v(1):
            return
        if self.total <= 0:
            return
        width = 24
        frac = min(max(self.n / self.total, 0.0), 1.0)
        fill = int(round(frac * width))
        bar = "#" * fill + "-" * (width - fill)
        pct = int(frac * 100)
        pf = " ".join([f"{k}={v}" for k, v in self.postfix.items()])
        msg = f"{self.desc} [{bar}] {self.n}/{self.total} {pct}% {pf}".strip()
        pad = " " * max(self._last_len - len(msg), 0)
        sys.stderr.write("\r" + msg + pad)
        sys.stderr.flush()
        self._last_len = len(msg)

    def close(self):
        if not v(1):
            return
        if self.total > 0:
            sys.stderr.write("\n")
            sys.stderr.flush()


# =========================
# Verbosity
# =========================

# Уровни:
# 0 = тихо (без прогресс-баров, минимум вспомогательного мусора)
# 1 = нормально (дефолт)
# 2 = дебаг (чуть больше инфы о настройках/фоллбеках)
GLOBAL_VERBOSITY: int = 1

def set_global_verbosity(level: int) -> None:
    global GLOBAL_VERBOSITY
    try:
        GLOBAL_VERBOSITY = int(level)
    except Exception:
        GLOBAL_VERBOSITY = 1
    if GLOBAL_VERBOSITY < 0:
        GLOBAL_VERBOSITY = 0
    if GLOBAL_VERBOSITY > 2:
        GLOBAL_VERBOSITY = 2

def v(level: int = 1) -> bool:
    return GLOBAL_VERBOSITY >= int(level)

def vprint(level: int, *args, **kwargs) -> None:
    if v(level):
        print(*args, **kwargs)

def veprint(level: int, *args, **kwargs) -> None:
    if v(level):
        print(*args, file=sys.stderr, **kwargs)


# =========================
# Run logger (в файл, чтобы не проёбывать день из-за отсутствия вывода)
# =========================

RUN_LOG_PATH: Optional[Path] = None
RUN_EVENTS_PATH: Optional[Path] = None

def init_run_logger(out_dir: Path) -> None:
    """Инициализирует лог-файлы в out/logs. Пишем ВСЕ события, печать в консоль зависит от verbosity."""
    global RUN_LOG_PATH, RUN_EVENTS_PATH
    try:
        logs_dir = out_dir / "logs"
        ensure_dir(logs_dir)
        RUN_LOG_PATH = logs_dir / "run.log"
        RUN_EVENTS_PATH = logs_dir / "events.jsonl"
        if not RUN_LOG_PATH.exists():
            RUN_LOG_PATH.write_text("", encoding="utf-8")
        if not RUN_EVENTS_PATH.exists():
            RUN_EVENTS_PATH.write_text("", encoding="utf-8")
    except Exception:
        RUN_LOG_PATH = None
        RUN_EVENTS_PATH = None

def _log_to_files(code: str, msg: str, level: int, err: bool = False, nm_id: Optional[str] = None) -> None:
    ts = utc_now_iso()
    try:
        if RUN_LOG_PATH:
            line = f"{ts}\t{level}\t{code or '-'}\t{nm_id or '-'}\t{msg}\n"
            with RUN_LOG_PATH.open("a", encoding="utf-8") as f:
                f.write(line)
    except Exception:
        pass
    try:
        if RUN_EVENTS_PATH:
            ev = {"ts": ts, "level": int(level), "stage": code or "", "nm_id": nm_id or "", "msg": msg, "err": bool(err)}
            with RUN_EVENTS_PATH.open("a", encoding="utf-8") as f:
                f.write(json.dumps(ev, ensure_ascii=False) + "\n")
    except Exception:
        pass

def stage_sku(code: str, nm_id: str, msg: str, level: int = 2, err: bool = False) -> None:
    stage_line(code, f"nm={nm_id} {msg}", level=level, err=err, nm_id=nm_id)


# =========================
# Артефакты (для понятных OK/FAIL логов)
# =========================

# Каноничные выходные файлы по стадиям (для коротких сводок выполнения).
STAGE_ARTIFACTS = {
    "A": ["run_manifest.json"],
    "B": ["own_norm.jsonl"],
    "C": ["intent.jsonl"],
    "D": ["queries_raw.jsonl"],
    "E": ["queries_valid.jsonl"],
    "F": ["competitor_pool.jsonl"],
    "G": ["competitor_lite.jsonl"],
    "H": ["relevance.jsonl"],
    "I": ["market_pulse.jsonl"],
    "J": ["supply_structure.jsonl"],
    "K": ["cluster_verdicts.jsonl"],
    "L": ["decisions.jsonl"],
    "M": ["WB_NECROMANCER_REPORT.xlsx", "WB_NECROMANCER_REPORT.html"],
}

# Опциональные кэши (для дебага, не критично)
STAGE_CACHES = {
    "B": [".wb_cache"],
    "E": [".wb_cache/serp"],
    "G": [".wb_cache/cards"],
    "I": [".wb_cache/reviews"],
}


def _fmt_bytes(n: int) -> str:
    try:
        n = int(n)
    except Exception:
        return "?B"
    units = ["B", "KB", "MB", "GB"]
    x = float(n)
    for u in units:
        if x < 1024.0 or u == units[-1]:
            if u == "B":
                return f"{int(x)}{u}"
            return f"{x:.1f}{u}"
        x /= 1024.0
    return f"{n}B"


def _count_lines(path: Path, limit: int = 2_000_000) -> int:
    # JSONL обычно небольшой, но чтобы не словить боль, есть лимит.
    n = 0
    try:
        with path.open("rb") as f:
            for _ in f:
                n += 1
                if n >= limit:
                    break
    except Exception:
        return -1
    return n


def _artifact_stat(path: Path) -> dict:
    stat = {"exists": False, "path": str(path), "bytes": 0, "lines": None}
    try:
        if not path.exists():
            return stat
        stat["exists"] = True
        stat["bytes"] = path.stat().st_size
        if path.suffix.lower() == ".jsonl":
            stat["lines"] = _count_lines(path)
        return stat
    except Exception:
        return stat


def _dir_file_count(path: Path, cap: int = 200_000) -> int:
    # Для дебага: сколько файлов в кэше. cap чтобы не умереть на гигантских папках.
    try:
        if not path.exists() or not path.is_dir():
            return 0
        n = 0
        for _ in path.rglob("*"):
            n += 1
            if n >= cap:
                break
        return n
    except Exception:
        return 0


def stage_line(code: str, msg: str, level: int = 1, err: bool = False, nm_id: Optional[str] = None) -> None:
    _log_to_files(code, msg, level, err=err, nm_id=nm_id)
    if not v(level):
        return
    prefix = f"[{code}] " if code else ""
    if err:
        print(prefix + msg, file=sys.stderr)
    else:
        print(prefix + msg)


def stage_ok(code: str, msg: str, level: int = 1) -> None:
    stage_line(code, f"OK: {msg}", level=level)


def stage_dbg(code: str, msg: str) -> None:
    stage_line(code, f"DBG: {msg}", level=2)


def stage_warn(code: str, msg: str) -> None:
    stage_line(code, f"WARN: {msg}", level=1, err=True)


def stage_fail(code: str, msg: str) -> None:
    stage_line(code, f"FAIL: {msg}", level=0, err=True)


def stage_artifacts_summary(code: str, out_dir: Path) -> str:
    files = STAGE_ARTIFACTS.get(code, [])
    parts = []
    for fn in files:
        st = _artifact_stat(out_dir / fn)
        if not st["exists"]:
            parts.append(f"{fn}: missing")
        else:
            if st.get("lines") is not None:
                ln = st["lines"]
                ln_s = "?" if ln < 0 else str(ln)
                parts.append(f"{fn}: {ln_s} lines, {_fmt_bytes(st['bytes'])}")
            else:
                parts.append(f"{fn}: {_fmt_bytes(st['bytes'])}")
    return "; ".join(parts) if parts else "(no artifacts)"


def stage_caches_summary(code: str, out_dir: Path) -> str:
    caches = STAGE_CACHES.get(code, [])
    parts = []
    for rel in caches:
        n = _dir_file_count(out_dir / rel)
        parts.append(f"{rel}: {n} files")
    return "; ".join(parts)


# =========================
# Лексикон (один раз при старте)
# =========================

@dataclass(frozen=True)
class Lexicon:
    case_intent_terms: Tuple[str, ...]
    tpu_terms: Tuple[str, ...]
    pocket_terms: Tuple[str, ...]
    ban_terms_default: Tuple[str, ...]


def load_lexicon(path: Optional[str]) -> Lexicon:
    base = DEFAULT_LEXICON
    data = {}
    if path:
        p = Path(path)
        if p.exists():
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                data = {}
    def _get_list(k: str) -> List[str]:
        v = data.get(k)
        if isinstance(v, list):
            return [safe_str(x).strip() for x in v if safe_str(x).strip()]
        return [safe_str(x).strip() for x in base.get(k, []) if safe_str(x).strip()]
    return Lexicon(
        case_intent_terms=tuple(_get_list("case_intent_terms")),
        tpu_terms=tuple(_get_list("tpu_terms")),
        pocket_terms=tuple(_get_list("pocket_terms")),
        ban_terms_default=tuple(_get_list("ban_terms_default")),
    )


# =========================
# WB HTTP клиент и эндпоинты
# =========================

def req_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    s.trust_env = True  # если у тебя прокси в окружении, это оно
    return s

def backoff_sleep(attempt: int, base: float = 0.4, cap: float = 6.0) -> None:
    time.sleep(min(cap, base * (2 ** attempt) + random.random() * 0.25))

def wb_product_url(nm_id: str) -> str:
    return f"https://www.wildberries.ru/catalog/{nm_id_to_str(nm_id)}/detail.aspx"

def wb_api_v4_url(nm_id: str, dest: int) -> str:
    return f"https://card.wb.ru/cards/v4/detail?appType=1&curr=rub&dest={dest}&nm={nm_id_to_str(nm_id)}"

def wb_api_v1_url(nm_id: str, dest: int) -> str:
    return f"https://card.wb.ru/cards/v1/detail?appType=1&curr=rub&dest={dest}&nm={nm_id_to_str(nm_id)}"

def wb_search_v18_url(query: str, dest: int, *, page: int = 1, limit: int = 100, sort: str = "popular", host: str = "u-search.wb.ru") -> str:
    return (f"https://{host}/exactmatch/ru/common/v18/search"
            f"?appType=1&curr=rub&dest={dest}"
            f"&lang=ru&inheritFilters=false&suppressSpellcheck=false"
            f"&query={requests.utils.quote(query)}"
            f"&page={page}&resultset=catalog&sort={sort}&spp=30&limit={limit}")

def wb_get_json(sess: requests.Session, url: str, *, timeout: int, params: Optional[Dict[str, Any]] = None,
                retries: int = 3, backoff: float = 0.8) -> Tuple[Optional[dict], Dict[str, Any]]:
    last_err = ""
    for a in range(retries + 1):
        try:
            r = sess.get(url, params=params, timeout=timeout)
            info = {"url": url, "params": params, "status": r.status_code}
            if r.status_code == 200:
                try:
                    return r.json(), info
                except Exception as e:
                    return None, {**info, "error": f"bad_json:{e!r}", "text_snip": r.text[:400]}
            if r.status_code in (429, 500, 502, 503, 504):
                backoff_sleep(a, base=float(backoff))
                continue
            return None, {**info, "error": f"http_{r.status_code}", "text_snip": r.text[:400]}
        except Exception as e:
            last_err = repr(e)
            backoff_sleep(a, base=float(backoff))
    return None, {"url": url, "params": params, "status": None, "error": last_err or "unknown"}

def post_json(sess: requests.Session, url: str, payload: Dict[str, Any], *, timeout: int, retries: int = 3) -> Tuple[Optional[dict], Dict[str, Any]]:
    last_err = ""
    for a in range(retries + 1):
        try:
            r = sess.post(url, json=payload, timeout=timeout)
            info = {"url": url, "status": r.status_code, "payload_keys": list(payload.keys())[:12]}
            if r.status_code == 200:
                try:
                    return r.json(), info
                except Exception as e:
                    return None, {**info, "error": f"bad_json:{e!r}", "text_snip": r.text[:400]}
            if r.status_code in (429, 500, 502, 503, 504):
                backoff_sleep(a)
                continue
            return None, {**info, "error": f"http_{r.status_code}", "text_snip": r.text[:400]}
        except Exception as e:
            last_err = repr(e)
            backoff_sleep(a)
    return None, {"url": url, "status": None, "error": last_err or "unknown"}


# =========================
# Deep card.json (wbbasket)
# =========================

def wb_basket_host_by_vol(vol: int) -> str:
    # да, это эмпирика. да, WB любит менять. да, поэтому есть fallback probe.
    v = int(vol)
    if 0 <= v <= 143:   return "01"
    if 144 <= v <= 287: return "02"
    if 288 <= v <= 431: return "03"
    if 432 <= v <= 719: return "04"
    if 720 <= v <= 1007:return "05"
    if 1008 <= v <= 1061:return "06"
    if 1062 <= v <= 1115:return "07"
    if 1116 <= v <= 1169:return "08"
    if 1170 <= v <= 1313:return "09"
    if 1314 <= v <= 1601:return "10"
    if 1602 <= v <= 1655:return "11"
    if 1656 <= v <= 1919:return "12"
    if 1920 <= v <= 2045:return "13"
    if 2046 <= v <= 2189:return "14"
    if 2190 <= v <= 2405:return "15"
    if 2406 <= v <= 2621:return "16"
    if 2622 <= v <= 2837:return "17"
    return "18"

def wb_basket_card_json_url(nm_id: str, *, host_num: Optional[str] = None, lang: str = "ru") -> str:
    nm = safe_int(nm_id, None)
    if nm is None:
        nm = int(nm_id_to_str(nm_id) or "0")
    vol = nm // 100000
    part = nm // 1000
    host = (host_num or wb_basket_host_by_vol(vol)).zfill(2)
    return f"https://basket-{host}.wbbasket.ru/vol{vol}/part{part}/{nm}/info/{lang}/card.json"

def is_valid_deep_card_json(nm_id: str, js: dict) -> bool:
    if not isinstance(js, dict) or not js:
        return False
    if "imt_id" in js or "nm_id" in js or "data" in js or "options" in js:
        return True
    if isinstance(js.get("data"), dict):
        return True
    return False

def fetch_deep_card(sess: requests.Session, nm_id: str, *, timeout: int, cache_dir: Path) -> Tuple[Optional[dict], Dict[str, Any]]:
    ensure_dir(cache_dir)
    nm = nm_id_to_str(nm_id)
    cache_path = cache_dir / f"deep_{nm}.json"
    if cache_path.exists():
        try:
            js = read_json(cache_path)
            if is_valid_deep_card_json(nm, js):
                return js, {"cache": True, "path": str(cache_path)}
        except Exception:
            pass

    nm_int = safe_int(nm, 0) or 0
    vol = nm_int // 100000
    base_host = int(wb_basket_host_by_vol(vol))
    host_try = [base_host, base_host - 1, base_host + 1]
    info_all = {"cache": False, "tries": []}
    for h in host_try:
        if h < 1 or h > 18:
            continue
        url = wb_basket_card_json_url(nm, host_num=str(h).zfill(2))
        js, info = wb_get_json(sess, url, timeout=timeout, retries=2, backoff=0.6)
        info_all["tries"].append(info)
        if js and is_valid_deep_card_json(nm, js):
            write_json(cache_path, js)
            return js, {"cache": False, "path": str(cache_path), "used_url": url, "tries": info_all["tries"]}
        time.sleep(0.12)
    return None, {"cache": False, "error": "deep_card_not_found", "tries": info_all["tries"]}


# =========================
# Парсеры WB: карточка, поиск
# =========================

def norm_lc(s: str) -> str:
    s = (s or "").lower().replace("ё", "е")
    s = re.sub(r"[\t\r\n]+", " ", s)
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s

def parse_card_v4(js: dict) -> dict:
    out: dict = {"content": {}, "pricing": {}, "metrics": {}, "ids": {}, "raw": None}
    if not isinstance(js, dict):
        return out
    out["raw"] = js
    products = None
    # WB умеет отдавать продукты в разных обёртках, потому что единый формат это роскошь
    if isinstance(js.get("data"), dict) and isinstance(js["data"].get("products"), list):
        products = js["data"]["products"]
    elif isinstance(js.get("payload"), dict) and isinstance(js["payload"].get("products"), list):
        products = js["payload"]["products"]
    elif isinstance(js.get("payload"), dict) and isinstance(js["payload"].get("data"), dict) and isinstance(js["payload"]["data"].get("products"), list):
        products = js["payload"]["data"]["products"]
    elif isinstance(js.get("products"), list):
        products = js["products"]
    if not products:
        return out
    p = products[0] if isinstance(products[0], dict) else {}
    out["ids"] = {
        "nm_id": nm_id_to_str(p.get("id") or p.get("nmId") or ""),
        "imt_id": safe_int(p.get("root") or p.get("imtId") or p.get("rootId") or p.get("imt_id")),
        "subject_id": safe_int(p.get("subjectId") or p.get("subject")),
    }
    out["content"] = {
        "title": safe_str(p.get("name") or ""),
        "brand": safe_str(p.get("brand") or ""),
        "seller": safe_str(p.get("supplier") or p.get("supplierName") or p.get("seller") or ""),
        "subject_name": safe_str(p.get("subjectName") or p.get("subjName") or ""),
    }
    out["pricing"] = {
        "priceU": p.get("priceU") or p.get("price"),
        "salePriceU": p.get("salePriceU") or p.get("salePrice"),
    }
    out["metrics"] = {
        "rating": p.get("rating") if p.get("rating") is not None else p.get("reviewRating") or p.get("nmReviewRating"),
        "feedbacks": p.get("feedbacks") if p.get("feedbacks") is not None else p.get("nmFeedbacks"),
        "total_quantity": p.get("totalQuantity") or p.get("quantity") or None,
    }
    return out

def parse_card_v1(js: dict) -> dict:
    return parse_card_v4(js)

def parse_search_items(js: dict) -> List[dict]:
    if not isinstance(js, dict):
        return []
    products = None
    if isinstance(js.get("products"), list):
        products = js.get("products")
    else:
        data = js.get("data")
        if isinstance(data, dict) and isinstance(data.get("products"), list):
            products = data.get("products")
    if not isinstance(products, list):
        return []
    out: List[dict] = []
    for idx, p in enumerate(products, start=1):
        if not isinstance(p, dict):
            continue
        nm = nm_id_to_str(p.get("id") or p.get("nmId") or "")
        if not nm.isdigit():
            continue
        imt = safe_int(p.get("root") or p.get("imtId") or p.get("rootId") or p.get("imt_id"))
        seller_id = safe_int(p.get("supplierId") or p.get("sellerId") or p.get("supplier_id"))
        seller_name = safe_str(p.get("supplier") or p.get("supplierName") or p.get("seller") or "")
        subject_id = p.get("subjectId") or p.get("subject")
        subject_name = safe_str(p.get("subjectName") or p.get("subjName") or "")
        rating = p.get("rating")
        if rating is None:
            rating = p.get("reviewRating") or p.get("nmReviewRating")
        feedbacks = p.get("feedbacks")
        if feedbacks is None:
            feedbacks = p.get("nmFeedbacks")
        out.append({
            "pos": idx,
            "nm_id": nm,
            "imt_id": imt,
            "seller_id": seller_id,
            "seller_name": seller_name,
            "name": safe_str(p.get("name") or ""),
            "brand": safe_str(p.get("brand") or ""),
            "subject_id": subject_id,
            "subject_name": subject_name,
            "rating": rating,
            "feedbacks": feedbacks,
            "priceU": p.get("priceU") or p.get("price"),
            "salePriceU": p.get("salePriceU") or p.get("salePrice"),
            "raw": p,
        })
    return out

def search_price_rub(item: dict) -> Optional[int]:
    raw = item.get("raw") if isinstance(item.get("raw"), dict) else None
    def to_int(x: Any) -> Optional[int]:
        try:
            if x is None:
                return None
            return int(float(x))
        except Exception:
            return None
    price_u = to_int(item.get("priceU"))
    sale_u = to_int(item.get("salePriceU"))
    use = sale_u if sale_u is not None else price_u
    if use is None:
        return None
    raw_has_u = isinstance(raw, dict) and (("priceU" in raw) or ("salePriceU" in raw))
    if raw_has_u:
        return int(max(0, use) // 100)
    if use >= 10000:
        return int(use // 100)
    return int(use)

def query_tokens(q: str) -> List[str]:
    q = norm_lc(q)
    q = re.sub(r"[^a-zа-я0-9 ]+", " ", q)
    toks = [t for t in q.split() if len(t) >= 4]
    stop = {"для", "в", "на", "и", "с", "по", "как", "или", "без", "под", "все"}
    toks = [t for t in toks if t not in stop]
    return toks[:10]

def relevance_score_token_overlap(query: str, items: List[dict]) -> int:
    if not items:
        return 0
    q = safe_str(query).strip()
    if q.isdigit():
        qnm = q
        for it in items[:30]:
            if nm_id_to_str(it.get("nm_id")) == qnm:
                return 100
        return 0
    toks = query_tokens(q)
    if not toks:
        return 0
    topnames = " ".join(safe_str(x.get("name") or "") for x in items[:12]).lower().replace("ё", "е")
    return sum(1 for t in toks if t in topnames)

def fetch_search_best(sess: requests.Session, query: str, *, dests: List[int], hosts: List[str], timeout: int, limit: int) -> Tuple[str, int, Optional[dict], str, int, int, str]:
    best = None  # (score, count, code, js, url, dest, host)
    for host in hosts:
        for dest in dests:
            url = wb_search_v18_url(query, dest, page=1, limit=limit, host=host)
            js, info = wb_get_json(sess, url, timeout=timeout, retries=2, backoff=0.6)
            code = info.get("status") or 0
            if code != 200 or not js:
                continue
            items = parse_search_items(js)
            score = relevance_score_token_overlap(query, items)
            cand = (score, len(items), code, js, url, dest, host)
            if best is None or cand[:2] > best[:2]:
                best = cand
            if (score >= 1 or (safe_str(query).isdigit() and score >= 100)) and len(items) >= min(10, limit):
                return "ok", code, js, url, dest, score, host
    if best:
        score, _, code, js, url, dest, host = best
        return "ok", code, js, url, dest, score, host
    return "not_found", 404, None, "", 0, 0, ""


# =========================
# Извлечение модели телефона (rules-first)
# =========================

PHONE_MODEL_PATTERNS = [
    r"(iphone\s?\d{1,2}\s?(pro\s?max|pro|max|mini)?)",
    r"(iphone\s?se\s?\d?)",
    r"(samsung\s?(galaxy\s?)?(s|a|m)\d{1,2}\s?(ultra|plus|\+|fe)?)",
    r"(xiaomi\s?(redmi\s?)?(note\s?)?\d{1,2}\s?(pro\s?\+?|pro|plus|\+)?)",
    r"(poco\s?[a-z]?\d{1,2}\s?(pro|max|\+)?)",
    r"(realme\s?\d{1,2}\s?(pro\s?\+?|pro|plus|\+)?)",
    r"(honor\s?\d{1,2}\s?(lite|pro|plus)?)",
    r"(huawei\s?(p|mate)\d{1,2}\s?(pro|plus)?)",
    r"(oneplus\s?\d{1,2}\s?(pro|r)?)",
    r"(tecno\s?(camon|pova|spark)\s?\d{1,2}\s?(pro\s?max|pro|plus|neo|premier|5g|4g)?)",
    r"(infinix\s?(hot|note|zero)\s?\d{1,2}\s?(pro|plus|5g|4g)?)",
    r"(oppo\s?(reno|a|find)\s?\d{1,2}\s?(pro|plus|5g|4g)?)",
    r"(vivo\s?(v|y|x)\s?\d{1,2}\s?(pro|plus|5g|4g)?)",
    r"(айфон\s?\d{1,2}\s?(про\s?макс|про|max|мини)?)",
    r"(самсунг\s?(галакси\s?)?(s|a|m)\d{1,2}\s?(ультра|плюс|фе)?)",
    r"(редми\s?(ноут\s?)?\d{1,2}\s?(про\s?плюс|про|плюс)?)",
]

def extract_phone_models(text: str, limit: int = 5) -> List[str]:
    t = norm_lc(text)
    t = re.sub(r"(\d)([a-zа-я])", r"\1 \2", t, flags=re.IGNORECASE)
    t = re.sub(r"([a-zа-я])(\d)", r"\1 \2", t, flags=re.IGNORECASE)
    t = re.sub(r"\s+", " ", t).strip()
    t = re.sub(r"\b(\d{1,3})\s+\1\b", r"\1", t)
    found: List[str] = []
    for pat in PHONE_MODEL_PATTERNS:
        for m in re.finditer(pat, t, flags=re.IGNORECASE):
            s = m.group(0)
            s = re.sub(r"\s+", " ", s).strip()
            if s and s not in found:
                found.append(s)
            if len(found) >= limit:
                return found
    return found

def pass_must_any_groups(text: str, must_any_groups: List[dict]) -> bool:
    t = norm_lc(text)
    for g in must_any_groups or []:
        any_terms = g.get("any") or []
        if any_terms and not any(norm_lc(str(x)) in t for x in any_terms if str(x).strip()):
            return False
    return True

def hit_ban_terms(text: str, ban_terms: List[str]) -> bool:
    t = norm_lc(text)
    for b in ban_terms or []:
        b = norm_lc(str(b))
        if b and b in t:
            return True
    return False


# =========================
# LLM: OpenAI / OpenRouter (chat completions), строгий JSON
# =========================

LLM_SESSION = requests.Session()
LLM_SESSION.headers.update({"User-Agent": HEADERS["User-Agent"]})

def extract_json_object_from_text(raw: str) -> Optional[dict]:
    if not isinstance(raw, str):
        return None
    s = raw.strip()
    if not s:
        return None
    s2 = re.sub(r"^```(?:json)?\s*", "", s, flags=re.I)
    s2 = re.sub(r"\s*```$", "", s2)
    i = s2.find("{")
    j = s2.rfind("}")
    if i != -1 and j != -1 and j > i:
        try:
            obj = json.loads(s2[i:j+1])
            return obj if isinstance(obj, dict) else None
        except Exception:
            pass
    i = s2.find("[")
    j = s2.rfind("]")
    if i != -1 and j != -1 and j > i:
        try:
            arr = json.loads(s2[i:j+1])
            if isinstance(arr, list):
                return {"items": arr}
        except Exception:
            pass
    return None

def llm_endpoint(provider: str, base_url: str) -> str:
    provider = safe_str(provider).strip().lower()
    b = safe_str(base_url).strip()
    if not b:
        b = os.environ.get("LLM_BASE_URL", "").strip()
    if not b:
        b = os.environ.get("OPENAI_BASE_URL" if provider == "openai" else "OPENROUTER_BASE_URL", "").strip()
    if not b:
        b = "https://api.openai.com/v1" if provider == "openai" else "https://openrouter.ai/api/v1"
    b = b.rstrip("/")
    if b.endswith("/chat/completions") or b.endswith("/v1/chat/completions"):
        return b
    return b + "/chat/completions"

def llm_api_key(provider: str) -> str:
    provider = safe_str(provider).strip().lower()
    if provider == "openrouter":
        return os.environ.get("OPENROUTER_API_KEY", "").strip() or os.environ.get("LLM_API_KEY", "").strip()
    return os.environ.get("OPENAI_API_KEY", "").strip() or os.environ.get("LLM_API_KEY", "").strip()

def call_llm_json(*, provider: str, model: str, api_key: str, messages: List[dict],
                  base_url: str = "", timeout_sec: int = 60, max_tokens: int = 900,
                  temperature: float = 0.2, force_json: bool = True) -> Tuple[dict, dict]:
    provider = safe_str(provider).strip().lower()
    if provider not in ("openai", "openrouter"):
        raise ValueError(f"Unsupported provider: {provider}")
    if not api_key:
        raise ValueError("Missing LLM API key")

    url = llm_endpoint(provider, base_url)

    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    if provider == "openrouter":
        ref = os.environ.get("OPENROUTER_HTTP_REFERER", "").strip()
        title = os.environ.get("OPENROUTER_X_TITLE", "").strip()
        if ref:
            headers["HTTP-Referer"] = ref
        if title:
            headers["X-Title"] = title

    payload_base: Dict[str, Any] = {"model": model, "messages": messages, "temperature": float(temperature)}
    if force_json:
        payload_base["response_format"] = {"type": "json_object"}

    token_variants: List[Tuple[str, int]] = [("max_tokens", int(max_tokens))]
    if provider == "openai":
        token_variants = [("max_completion_tokens", int(max_tokens)), ("max_tokens", int(max_tokens))]

    def _post(payload: Dict[str, Any]) -> Tuple[dict, dict]:
        t0 = time.time()
        r = LLM_SESSION.post(url, headers=headers, json=payload, timeout=timeout_sec)
        dt = time.time() - t0
        dbg = {
            "url": url,
            "http_code": r.status_code,
            "elapsed_sec": round(dt, 3),
            "usage": {},
            "retried_without_response_format": False,
            "retried_token_param": False,
            "raw_http_preview": safe_str(r.text)[:1600],
        }
        if r.status_code != 200:
            raise RuntimeError(f"LLM HTTP {r.status_code}: {dbg['raw_http_preview'][:700]}")
        js = r.json()
        usage = js.get("usage") if isinstance(js, dict) else None
        if isinstance(usage, dict):
            dbg["usage"] = usage
        return js, dbg

    last_err: Optional[Exception] = None
    dbg_final: dict = {}

    for tkey, tval in token_variants:
        payload = dict(payload_base)
        payload[tkey] = tval
        try:
            js, dbg = _post(payload)
            dbg_final = dbg
        except Exception as e:
            last_err = e
            continue

        content = ""
        try:
            choices = js.get("choices", [])
            if choices and isinstance(choices[0], dict):
                msg = choices[0].get("message", {})
                if isinstance(msg, dict):
                    content = safe_str(msg.get("content", ""))
        except Exception:
            content = ""

        obj = extract_json_object_from_text(content)
        if obj is not None:
            return obj, dbg_final

        if force_json:
            try:
                payload2 = dict(payload)
                payload2.pop("response_format", None)
                js2, dbg2 = _post(payload2)
                dbg2["retried_without_response_format"] = True
                dbg_final = dbg2

                content2 = ""
                choices = js2.get("choices", [])
                if choices and isinstance(choices[0], dict):
                    msg = choices[0].get("message", {})
                    if isinstance(msg, dict):
                        content2 = safe_str(msg.get("content", ""))

                obj2 = extract_json_object_from_text(content2)
                if obj2 is not None:
                    return obj2, dbg_final
            except Exception as e2:
                last_err = e2

        last_err = RuntimeError("LLM returned no JSON")

    raise RuntimeError(f"LLM request failed: {safe_str(last_err)}")


# =========================
# Meta + schema helpers
# =========================

def make_meta(run_id: str, stage: str, nm_id: str, vendor_code: str, name: str = "") -> dict:
    return {
        "schema_version": SCHEMA_VERSION,
        "run_id": safe_str(run_id),
        "stage": safe_str(stage),
        "ts": utc_now_iso(),
        "nm_id": nm_id_to_str(nm_id),
        "vendor_code": safe_str(vendor_code),
        "name": safe_str(name),
        "source": "necromancer_rewrite",
    }

def load_manifest(out_dir: Path) -> dict:
    p = out_dir / "run_manifest.json"
    if not p.exists():
        raise FileNotFoundError(f"Нет run_manifest.json в {out_dir} (сделай Stage A)")
    return read_json(p)

def scope_from_manifest(manifest: dict) -> List[dict]:
    scope = ((manifest.get("scope") or {}).get("sku_list")) or []
    if not isinstance(scope, list):
        return []
    return [x for x in scope if isinstance(x, dict) and nm_id_to_str(x.get("nm_id"))]

def map_by_nm(path: Path) -> Dict[str, dict]:
    m: Dict[str, dict] = {}
    for r in iter_jsonl(path):
        meta = r.get("meta") or {}
        nm = nm_id_to_str(meta.get("nm_id"))
        if nm:
            m[nm] = r
    return m


# =========================
# Stage A: input -> manifest
# =========================

INPUT_COLS_CANDIDATES = {
    "nm_id": ["nm_id", "NMID", "Артикул", "Артикул WB", "АртикулWB"],
    "vendor_code": ["vendor_code", "Артикул продавца", "АртикулПродавца"],
    "name": ["name", "Наименование", "Название", "Наименование товара"],
    "potential_qty": ["potential_qty", "Потенциал", "potential", "Потенциал продаж"],
}

def _find_col(header_row: List[Any], candidates: List[str]) -> Optional[int]:
    hdr = [norm_lc(safe_str(x)) for x in header_row]
    for c in candidates:
        c = norm_lc(c)
        for i, h in enumerate(hdr):
            if h == c:
                return i
    return None

def read_input_xlsx(input_path: Path, sheet: str, expect_count: Optional[int], dedupe: bool) -> List[dict]:
    wb = load_workbook(filename=str(input_path), read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Нет листа {sheet!r}. Есть: {wb.sheetnames}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Пустой лист, ну и что ты ожидал?")
    header = list(rows[0])
    idx_nm = _find_col(header, INPUT_COLS_CANDIDATES["nm_id"])
    if idx_nm is None:
        raise ValueError("Не нашёл колонку nm_id (артикул WB). Добавь столбец nm_id.")
    idx_vc = _find_col(header, INPUT_COLS_CANDIDATES["vendor_code"])
    idx_name = _find_col(header, INPUT_COLS_CANDIDATES["name"])
    idx_qty = _find_col(header, INPUT_COLS_CANDIDATES["potential_qty"])

    out = []
    seen = set()
    for r in rows[1:]:
        if not r:
            continue
        nm = nm_id_to_str(r[idx_nm] if idx_nm < len(r) else "")
        if not nm:
            continue
        if nm in seen:
            if dedupe:
                continue
            raise ValueError(f"Дубликат nm_id={nm} (включи --dedupe если хочешь молча игнорить)")
        seen.add(nm)
        vc = safe_str(r[idx_vc]) if idx_vc is not None and idx_vc < len(r) else ""
        name = safe_str(r[idx_name]) if idx_name is not None and idx_name < len(r) else ""
        qty = safe_int(r[idx_qty], None) if idx_qty is not None and idx_qty < len(r) else None
        out.append({"nm_id": nm, "vendor_code": vc, "name": name, "potential_qty": qty})
    if expect_count is not None and expect_count > 0 and len(out) != expect_count:
        raise ValueError(f"Ожидали {expect_count} строк, а получили {len(out)}.")
    return out

def stage_A_manifest(out_dir: Path, *, input_xlsx: Path, sheet: str, expect_count: Optional[int], dedupe: bool,
                     dests: List[int], search_hosts: List[str], search_limit: int,
                     lexicon_path: Optional[str]) -> Path:
    sku_list = read_input_xlsx(input_xlsx, sheet, expect_count, dedupe)
    run_id = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S") + "_" + sha1_short(str(input_xlsx) + sheet + str(len(sku_list)), 8)

    lex = load_lexicon(lexicon_path)

    manifest = {
        "meta": {
            "script": SCRIPT_NAME,
            "script_version": SCRIPT_VERSION,
            "schema_version": SCHEMA_VERSION,
            "created_at": utc_now_iso(),
            "run_id": run_id,
        },
        "scope": {"input_xlsx": str(input_xlsx), "sheet": sheet, "sku_n": len(sku_list), "sku_list": sku_list},
        "config": {
            "wb": {"dests": dests, "search_hosts": search_hosts, "search_limit": int(search_limit)},
            "lexicon": dataclasses.asdict(lex),
            "pulse_rules": DEFAULT_PULSE_RULES,
            "supply_thresholds": DEFAULT_SUPPLY_THRESHOLDS,
            "reviews": DEFAULT_REVIEWS_CFG,
        },
    }
    ensure_dir(out_dir)
    write_json(out_dir / "run_manifest.json", manifest)
    stage_dbg("A", f"run_id={run_id} sku_n={len(sku_list)} -> {out_dir/'run_manifest.json'}")
    return out_dir / "run_manifest.json"


# =========================
# Stage B: own fetch
# =========================

def fetch_card_any(sess: requests.Session, nm_id: str, *, dests: List[int], timeout: int) -> Tuple[str, int, Optional[dict], str, int]:
    # возвращает (status, http_code, json, url, dest_used)
    for dest in dests:
        url = wb_api_v4_url(nm_id, dest)
        js, info = wb_get_json(sess, url, timeout=timeout, retries=2, backoff=0.6)
        code = info.get("status") or 0
        if code == 200 and js:
            return "ok_v4", code, js, url, dest
    for dest in dests:
        url = wb_api_v1_url(nm_id, dest)
        js, info = wb_get_json(sess, url, timeout=timeout, retries=2, backoff=0.6)
        code = info.get("status") or 0
        if code == 200 and js:
            return "ok_v1", code, js, url, dest
    return "not_found", 404, None, "", 0

def stage_B_own_fetch(out_dir: Path, *, timeout: int, sleep_s: float, resume: bool, deep_card: bool) -> Path:
    manifest = load_manifest(out_dir)
    run_id = manifest["meta"]["run_id"]
    scope = scope_from_manifest(manifest)

    wb_cfg = (manifest.get("config") or {}).get("wb") or {}
    dests = list(wb_cfg.get("dests") or DEFAULT_DESTS)

    out_path = out_dir / "own_norm.jsonl"
    err_path = out_dir / "own_errors.jsonl"
    done = read_jsonl_done_ids(out_path) if resume else set()

    sess = req_session()
    cache_dir = out_dir / ".wb_cache" / "own"
    ensure_dir(cache_dir)

    pbar = SimpleProgress(total=len(scope), desc="[B] own_fetch")
    for sku in scope:
        nm = nm_id_to_str(sku.get("nm_id"))
        if resume and nm in done:
            pbar.update(1)
            continue
        vendor_code = safe_str(sku.get("vendor_code", ""))
        name = safe_str(sku.get("name", ""))

        status, http_code, js, api_url, dest_used = fetch_card_any(sess, nm, dests=dests, timeout=timeout)
        card = parse_card_v4(js) if js else {}
        deep_js = None
        deep_info = {}
        if deep_card:
            try:
                deep_js, deep_info = fetch_deep_card(sess, nm, timeout=timeout, cache_dir=cache_dir)
            except Exception as e:
                deep_js, deep_info = None, {"error": repr(e)}

        
        # если v4 вернул пустоту, но deep-card есть — выдернем хотя бы название, чтобы дальше не умирать как идиоты
        if (not (card.get("content", {}).get("title") or "")) and isinstance(deep_js, dict):
            deep_title = safe_str(deep_js.get("imt_name") or deep_js.get("name") or deep_js.get("goods_name") or deep_js.get("title") or "")
            deep_brand = safe_str(deep_js.get("brand") or deep_js.get("brand_name") or "")
            if deep_title:
                card.setdefault("content", {})["title"] = deep_title
            if deep_brand and not safe_str(card.get("content", {}).get("brand") or ""):
                card.setdefault("content", {})["brand"] = deep_brand

        rec = {
            "meta": make_meta(run_id, "B", nm, vendor_code, name),
            "fetch": {"status": status, "http": http_code, "api_url": api_url, "dest_used": dest_used, "deep": deep_info},
            "own_card": {"product_url": wb_product_url(nm), "v4": card, "deep": deep_js},
        }
        append_jsonl(out_path, rec)
        # построчно, чтобы было видно что реально собрали
        title_ok = bool((card.get("content", {}).get("title") or "").strip())
        stage_sku("B", nm, f"fetch={status} http={http_code} title={'ok' if title_ok else 'missing'} deep={'yes' if isinstance(deep_js, dict) else 'no'}", level=1, err=(status!="ok" or not title_ok))

        if status == "not_found" or not (card.get("content", {}).get("title") or ""):
            append_jsonl(err_path, {"meta": rec["meta"], "error": "own_card_missing_or_empty", "fetch": rec["fetch"]})

        time.sleep(max(0.0, float(sleep_s)))
        pbar.update(1)
    pbar.close()
    stage_dbg("B", f"wrote {out_path}")
    return out_path


# =========================
# Stage C: intent extract
# =========================

def stage_C_intent(out_dir: Path, *, resume: bool = False) -> Path:
    manifest = load_manifest(out_dir)
    run_id = manifest["meta"]["run_id"]
    scope = scope_from_manifest(manifest)

    lex = Lexicon(**((manifest.get("config") or {}).get("lexicon") or dataclasses.asdict(load_lexicon(None))))

    own_map = map_by_nm(out_dir / "own_norm.jsonl")
    out_path = out_dir / "intent.jsonl"
    done = read_jsonl_done_ids(out_path) if resume else set()

    for sku in scope:
        nm = nm_id_to_str(sku.get("nm_id"))
        if resume and nm in done:
            continue
        vendor_code = safe_str(sku.get("vendor_code", ""))
        name = safe_str(sku.get("name", ""))

        own = own_map.get(nm, {})
        v4 = ((own.get("own_card") or {}).get("v4") or {})
        content = v4.get("content") or {}
        title = safe_str(content.get("title") or name)
        brand = safe_str(content.get("brand") or "")
        seller = safe_str(content.get("seller") or "")
        subject_name = safe_str(content.get("subject_name") or "")

        deep = (own.get("own_card") or {}).get("deep")
        deep_text = ""
        if isinstance(deep, dict):
            deep_text = json.dumps(deep, ensure_ascii=False)[:8000]

        blob = " ".join([title, brand, seller, subject_name, deep_text])
        models = extract_phone_models(blob, limit=8)
        phone_model = models[0] if models else ""
        model_tokens = query_tokens(phone_model) if phone_model else []

        must_phone = [{"name": "case", "any": list(lex.case_intent_terms)}]
        if model_tokens:
            must_phone.append({"name": "model", "any": model_tokens[:6]})

        must_type = [
            {"name": "case", "any": list(lex.case_intent_terms)},
            {"name": "tpu", "any": list(lex.tpu_terms)},
            {"name": "pocket", "any": list(lex.pocket_terms)},
        ]
        if model_tokens:
            must_type.append({"name": "model", "any": model_tokens[:6]})

        ban_terms = list(lex.ban_terms_default)

        metrics = (v4.get("metrics") or {})
        karma = {"rating": metrics.get("rating"), "feedbacks": metrics.get("feedbacks")}

        intent = {
            "phone_model": phone_model,
            "phone_model_candidates": models[:8],
            "model_tokens": model_tokens[:10],
            "focus_type": FOCUS_TYPE,
            "karma": karma,
            "clusters": [
                {"cluster": "phone", "must_any_groups": must_phone, "ban_terms": ban_terms},
                {"cluster": "type", "must_any_groups": must_type, "ban_terms": ban_terms},
            ],
        }

        rec = {"meta": make_meta(run_id, "C", nm, vendor_code, name), "intent": intent}
        append_jsonl(out_path, rec)

    stage_dbg("C", f"wrote {out_path}")
    return out_path


# =========================
# Stage D: query build (rules + optional LLM)
# =========================

def validate_query(q: str, *, min_len: int, max_len: int, ban_terms: List[str]) -> Tuple[bool, str]:
    q = safe_str(q).strip()
    if len(q) < min_len:
        return False, "too_short"
    if len(q) > max_len:
        return False, "too_long"
    qn = norm_lc(q)
    if hit_ban_terms(qn, ban_terms):
        return False, "contains_ban_term"
    if re.search(r"[\\/]{2,}", q):
        return False, "slashes_spam"
    return True, "ok"

def rules_queries(cluster: str, phone_model: str, lex: Lexicon, n: int) -> List[str]:
    pm = safe_str(phone_model).strip()
    base = []
    if cluster == "phone":
        if pm:
            base += [f"чехол {pm}", f"чехол на {pm}", f"кейс {pm}", f"силиконовый чехол {pm}"]
        else:
            base += ["чехол для телефона", "силиконовый чехол", "прозрачный чехол"]
    else:
        if pm:
            base += [
                f"чехол {pm} tpu карман",
                f"чехол {pm} силиконовый с карманом",
                f"чехол {pm} держатель карт",
                f"чехол {pm} cardholder tpu",
            ]
        else:
            base += ["чехол tpu с карманом", "чехол держатель карт силиконовый"]
    out = []
    seen = set()
    for q in base:
        qn = norm_lc(q)
        if qn in seen:
            continue
        seen.add(qn)
        out.append(q)
        if len(out) >= n:
            break
    return out

def build_llm_query_prompt(*, cluster: str, phone_model: str, want_n: int, existing: List[str],
                           must_any_groups: List[dict], ban_terms: List[str], model_tokens: List[str]) -> List[dict]:
    sys_msg = (
        "Ты генерируешь поисковые запросы для Wildberries на русском. "
        "Верни ТОЛЬКО JSON объект формата {\"queries\":[\"...\", ...]}. "
        "Никакого markdown, никакого текста, только JSON."
    )
    crit = []
    if cluster == "phone":
        crit.append("Запросы про чехлы для указанной модели телефона.")
    else:
        crit.append("Запросы про TPU или силиконовый чехол С карманом под карты для указанной модели.")
    payload = {
        "cluster": cluster,
        "phone_model": phone_model,
        "model_tokens": model_tokens[:8],
        "want_n": int(want_n),
        "must_any_groups": must_any_groups[:6],
        "ban_terms": ban_terms[:20],
        "already_have": existing[:12],
        "criteria": crit,
    }
    return [{"role": "system", "content": sys_msg},
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False)}]

def stage_D_queries(out_dir: Path, *, resume: bool, rules_per_cluster: int, llm_extra_per_cluster: int,
                    min_len: int, max_len: int, use_llm: bool,
                    llm_provider: str, llm_model: str, llm_base_url: str, llm_timeout: int, llm_max_tokens: int, llm_temperature: float) -> Path:
    manifest = load_manifest(out_dir)
    run_id = manifest["meta"]["run_id"]
    scope = scope_from_manifest(manifest)
    lex = Lexicon(**((manifest.get("config") or {}).get("lexicon") or dataclasses.asdict(load_lexicon(None))))
    intent_map = map_by_nm(out_dir / "intent.jsonl")

    out_path = out_dir / "queries_raw.jsonl"
    done = read_jsonl_done_ids(out_path) if resume else set()

    api_key = llm_api_key(llm_provider) if use_llm else ""

    for sku in scope:
        nm = nm_id_to_str(sku.get("nm_id"))
        if resume and nm in done:
            continue
        vendor_code = safe_str(sku.get("vendor_code", ""))
        name = safe_str(sku.get("name", ""))

        intent_rec = intent_map.get(nm, {})
        intent = intent_rec.get("intent") or {}
        phone_model = safe_str(intent.get("phone_model") or "")
        model_tokens = intent.get("model_tokens") or []
        clusters = intent.get("clusters") or []

        packs: List[dict] = []
        llm_dbg_all: List[dict] = []

        for c in clusters:
            cluster = safe_str(c.get("cluster"))
            if cluster not in CLUSTERS:
                continue
            must_any_groups = list(c.get("must_any_groups") or [])
            ban_terms = list(c.get("ban_terms") or list(lex.ban_terms_default))

            # Если модель телефона не извлечена — type-market (TPU+карман) без модели превращается в мусор по всем моделям.
            # Поэтому честно отключаем кластер type для этого SKU и не делаем вид, что это валидный рынок.
            if cluster == "type" and not phone_model.strip():
                packs.append({
                    "cluster": cluster,
                    "queries": [],
                    "must_any_groups": must_any_groups,
                    "ban_terms": ban_terms,
                    "disabled": True,
                    "disabled_reason": "NO_PHONE_MODEL",
                    "validation": {"rules": [], "llm": []},
                })
                llm_dbg_all.append({"cluster": cluster, "used": False, "disabled": True, "reason": "NO_PHONE_MODEL"})
                continue

            rule_qs = rules_queries(cluster, phone_model, lex, rules_per_cluster)

            v_rule = []
            for q in rule_qs:
                ok, _ = validate_query(q, min_len=min_len, max_len=max_len, ban_terms=ban_terms)
                if ok:
                    v_rule.append(q)

            extra_qs: List[str] = []
            llm_dbg = {"cluster": cluster, "used": False}
            if use_llm and llm_extra_per_cluster > 0 and phone_model.strip():
                try:
                    msgs = build_llm_query_prompt(
                        cluster=cluster, phone_model=phone_model, want_n=llm_extra_per_cluster,
                        existing=v_rule, must_any_groups=must_any_groups, ban_terms=ban_terms,
                        model_tokens=model_tokens,
                    )
                    parsed, dbg = call_llm_json(
                        provider=llm_provider,
                        model=llm_model,
                        api_key=api_key,
                        base_url=llm_base_url,
                        messages=msgs,
                        timeout_sec=llm_timeout,
                        max_tokens=llm_max_tokens,
                        temperature=llm_temperature,
                        force_json=True,
                    )
                    llm_dbg = {"cluster": cluster, "used": True, **dbg}
                    cand = parsed.get("queries", [])
                    if not isinstance(cand, list):
                        cand = []
                    for q in cand:
                        ok, _ = validate_query(q, min_len=min_len, max_len=max_len, ban_terms=ban_terms)
                        if ok:
                            extra_qs.append(safe_str(q).strip())
                except Exception as e:
                    llm_dbg = {"cluster": cluster, "used": True, "error": safe_str(e)}
            llm_dbg_all.append(llm_dbg)

            all_qs = []
            seen = set()
            for q in v_rule + extra_qs:
                qn = norm_lc(q)
                if qn in seen:
                    continue
                seen.add(qn)
                all_qs.append(q)
            packs.append({
                "cluster": cluster,
                "phone_model": phone_model,
                "queries": all_qs[: max(1, rules_per_cluster + llm_extra_per_cluster)],
                "must_any_groups": must_any_groups,
                "ban_terms": ban_terms,
            })

        rec = {"meta": make_meta(run_id, "D", nm, vendor_code, name), "query_packs": packs, "llm_debug": llm_dbg_all}
        append_jsonl(out_path, rec)

    stage_dbg("D", f"wrote {out_path}")
    return out_path


# =========================
# Stage E: SERP validate + select
# =========================

def stage_E_serp(out_dir: Path, *, timeout: int, sleep_s: float, search_limit: int, resume: bool,
                 min_keep_per_cluster: int, max_keep_per_cluster: int,
                 min_pass_rate_phone: float, min_pass_rate_type: float) -> Path:
    manifest = load_manifest(out_dir)
    run_id = manifest["meta"]["run_id"]
    wb_cfg = (manifest.get("config") or {}).get("wb") or {}
    dests = list(wb_cfg.get("dests") or DEFAULT_DESTS)
    hosts = list(wb_cfg.get("search_hosts") or DEFAULT_SEARCH_HOSTS)
    queries_path = out_dir / "queries_raw.jsonl"
    if not queries_path.exists():
        raise FileNotFoundError("Нет queries_raw.jsonl. Сначала Stage D.")

    out_path = out_dir / "queries_valid.jsonl"
    done = read_jsonl_done_ids(out_path) if resume else set()

    sess = req_session()
    cache_dir = out_dir / ".wb_cache" / "serp"
    ensure_dir(cache_dir)

    rows = iter_jsonl(queries_path)
    pbar = SimpleProgress(total=len(rows), desc="[E] serp_validate")
    for rec in rows:
        meta = rec.get("meta") or {}
        nm = nm_id_to_str(meta.get("nm_id"))
        if not nm or (resume and nm in done):
            pbar.update(1)
            continue
        vendor_code = safe_str(meta.get("vendor_code", ""))
        name = safe_str(meta.get("name", ""))

        packs = rec.get("query_packs") or []
        valid_packs: List[dict] = []

        for pack in packs:
            cluster = safe_str(pack.get("cluster"))
            if cluster not in CLUSTERS:
                continue
            must_any_groups = list(pack.get("must_any_groups") or [])
            ban_terms = list(pack.get("ban_terms") or [])
            queries = [safe_str(x).strip() for x in (pack.get("queries") or []) if safe_str(x).strip()]

            if bool(pack.get("disabled")):
                valid_packs.append({
                    "cluster": cluster,
                    "disabled": True,
                    "disabled_reason": safe_str(pack.get("disabled_reason") or "disabled"),
                    "selected_queries": [],
                    "validation": {"queries": []},
                    "min_pass_rate": 0.0,
                })
                continue

            per_q: List[dict] = []

            for q in queries:
                if hit_ban_terms(q, ban_terms):
                    per_q.append({"query": q, "status": "dropped_ban_term", "rel50": 0, "items_count": 0,
                                  "metrics": {"top_n": 0, "pass_n": 0, "pass_rate": 0.0}, "cache_path": None})
                    continue

                status, code, js, url, dest_used, rel50, host_used = fetch_search_best(
                    sess, q, dests=dests, hosts=hosts, timeout=timeout, limit=search_limit
                )
                items = parse_search_items(js) if (status == "ok" and js) else []

                top_n = min(50, len(items))
                pass_n = 0
                for it in items[:top_n]:
                    title = safe_str(it.get("name") or "")
                    if hit_ban_terms(title, ban_terms):
                        continue
                    if not pass_must_any_groups(title, must_any_groups):
                        continue
                    pass_n += 1
                pass_rate = (pass_n / top_n) if top_n else 0.0

                cache_path = None
                if js:
                    cache_name = f"{nm}_{cluster}_{sha1_short(q, 10)}.json"
                    cache_path = cache_dir / cache_name
                    write_json(cache_path, js)

                per_q.append({
                    "query": q,
                    "status": status,
                    "http": code,
                    "rel50": int(rel50),
                    "items_count": len(items),
                    "metrics": {"top_n": int(top_n), "pass_n": int(pass_n), "pass_rate": round(pass_rate, 4)},
                    "url": url,
                    "dest_used": dest_used,
                    "host_used": host_used,
                    "cache_path": str(cache_path) if cache_path else None,
                })
                time.sleep(max(0.0, float(sleep_s)))

            min_rate = min_pass_rate_phone if cluster == "phone" else min_pass_rate_type
            good = [r for r in per_q if safe_float((r.get("metrics") or {}).get("pass_rate"), 0.0) >= min_rate and (r.get("http") == 200)]
            good_sorted = sorted(good, key=lambda r: (
                safe_float((r.get("metrics") or {}).get("pass_rate"), 0.0),
                safe_int(r.get("rel50"), 0),
                safe_int(r.get("items_count"), 0),
            ), reverse=True)
            keep = good_sorted[:max_keep_per_cluster]
            if len(keep) < min_keep_per_cluster:
                rest_src = [r for r in per_q if r.get("status") == "ok" and r.get("cache_path") and (r.get("http") == 200)]
                rest = sorted(rest_src, key=lambda r: (
                    safe_float((r.get("metrics") or {}).get("pass_rate"), 0.0),
                    safe_int(r.get("rel50"), 0),
                    safe_int(r.get("items_count"), 0),
                ), reverse=True)
                for r in rest:
                    if r in keep:
                        continue
                    keep.append(r)
                    if len(keep) >= min_keep_per_cluster:
                        break

            valid_packs.append({
                "cluster": cluster,
                "selected_queries": [r.get("query") for r in keep if r.get("query")],
                "validation": {"queries": per_q},
                "min_pass_rate": float(min_rate),
            })

        out_rec = {"meta": make_meta(run_id, "E", nm, vendor_code, name), "valid_packs": valid_packs}
        append_jsonl(out_path, out_rec)
        try:
            sp = 0; st = 0
            for pk in valid_packs:
                if pk.get("cluster") == "phone":
                    sp = len(pk.get("selected_queries") or [])
                if pk.get("cluster") == "type":
                    st = len(pk.get("selected_queries") or [])
            stage_sku("E", nm, f"selected_queries phone={sp} type={st}", level=1, err=(sp==0))
        except Exception:
            pass

        pbar.update(1)
    pbar.close()
    stage_dbg("E", f"wrote {out_path}")
    return out_path


# =========================
# Stage F: competitor pool
# =========================

def seller_key(item: dict) -> str:
    sid = item.get("seller_id")
    if sid is not None:
        return f"id:{sid}"
    sn = norm_lc(item.get("seller_name") or "")
    return f"name:{sn}" if sn else "unknown"

def cand_key(item: dict) -> str:
    imt = item.get("imt_id")
    if imt is not None:
        return f"imt:{imt}"
    return f"nm:{item.get('nm_id')}"

def stage_F_pool(out_dir: Path, *, resume: bool, competitors_k: int, per_query_take: int) -> Path:
    manifest = load_manifest(out_dir)
    run_id = manifest["meta"]["run_id"]

    val_path = out_dir / "queries_valid.jsonl"
    if not val_path.exists():
        raise FileNotFoundError("Нет queries_valid.jsonl. Сначала Stage E.")

    out_path = out_dir / "competitor_pool.jsonl"
    done = read_jsonl_done_ids(out_path) if resume else set()

    for rec in iter_jsonl(val_path):
        meta = rec.get("meta") or {}
        nm = nm_id_to_str(meta.get("nm_id"))
        if not nm or (resume and nm in done):
            continue
        vendor_code = safe_str(meta.get("vendor_code", ""))
        name = safe_str(meta.get("name", ""))

        pool_by_cluster: Dict[str, Dict[str, dict]] = {c: {} for c in CLUSTERS}
        stats_by_cluster: Dict[str, Dict[str, Any]] = {c: {"queries": 0, "items": 0} for c in CLUSTERS}

        for pack in rec.get("valid_packs") or []:
            cluster = safe_str(pack.get("cluster"))
            if cluster not in CLUSTERS:
                continue
            if bool(pack.get("disabled")):
                stats_by_cluster[cluster]["disabled"] = True
                continue

            selected = [safe_str(x) for x in (pack.get("selected_queries") or []) if safe_str(x)]
            selected_set = set(selected)
            stats_by_cluster[cluster]["queries"] += len(selected)
            if not selected_set:
                continue

            qrows = ((pack.get("validation") or {}).get("queries")) or []
            for qrow in qrows:
                qtxt = safe_str(qrow.get("query") or "")
                if selected_set and qtxt not in selected_set:
                    continue

                cache_path = qrow.get("cache_path")
                if not cache_path:
                    if v(2) and (not selected_set or qtxt in selected_set):
                        veprint(2, f"[F][WARN] nm={nm} cluster={cluster} selected_query_no_cache: {qtxt[:120]}")
                    continue
                p = Path(cache_path)
                if not p.exists():
                    if v(2) and (not selected_set or qtxt in selected_set):
                        veprint(2, f"[F][WARN] nm={nm} cluster={cluster} cache_missing: {str(p)} for query={qtxt[:120]}")
                    continue
                try:
                    js = read_json(p)
                except Exception:
                    continue
                items = parse_search_items(js)
                stats_by_cluster[cluster]["items"] += len(items)
                for it in items[:max(1, int(per_query_take))]:
                    key = cand_key(it)
                    if key not in pool_by_cluster[cluster]:
                        pool_by_cluster[cluster][key] = {
                            "nm_id": it.get("nm_id"),
                            "imt_id": it.get("imt_id"),
                            "seller_id": it.get("seller_id"),
                            "seller_name": it.get("seller_name"),
                            "name": it.get("name"),
                            "brand": it.get("brand"),
                            "subject_id": it.get("subject_id"),
                            "subject_name": it.get("subject_name"),
                            "rating": it.get("rating"),
                            "feedbacks": it.get("feedbacks"),
                            "price_rub": search_price_rub(it),
                            "best_pos": safe_int(it.get("pos"), 10**9),
                            "appearances": 1,
                        }
                    else:
                        c0 = pool_by_cluster[cluster][key]
                        c0["appearances"] = int(c0.get("appearances") or 0) + 1
                        c0["best_pos"] = min(int(c0.get("best_pos") or 10**9), safe_int(it.get("pos"), 10**9) or 10**9)

        selected_by_cluster: Dict[str, List[dict]] = {}
        for cluster in CLUSTERS:
            cands = list(pool_by_cluster[cluster].values())
            cands = sorted(cands, key=lambda c: (
                -(int(c.get("appearances") or 0)),
                int(c.get("best_pos") or 10**9),
                -(int((safe_float(c.get("rating"), 0.0) or 0.0) * 100)),
                -(int(c.get("feedbacks") or 0)),
            ))
            picked: List[dict] = []
            seller_seen: Dict[str, int] = {}
            for c in cands:
                if len(picked) >= competitors_k:
                    break
                sk = seller_key(c)
                if seller_seen.get(sk, 0) >= 2:
                    continue
                seller_seen[sk] = seller_seen.get(sk, 0) + 1
                picked.append(c)
            selected_by_cluster[cluster] = picked

        out_rec = {"meta": make_meta(run_id, "F", nm, vendor_code, name), "pool": {"by_cluster": selected_by_cluster, "stats": stats_by_cluster}}
        append_jsonl(out_path, out_rec)

    stage_dbg("F", f"wrote {out_path}")
    return out_path


# =========================
# Stage G: competitor lite fetch
# =========================

def stage_G_lite(out_dir: Path, *, timeout: int, sleep_s: float, resume: bool) -> Path:
    manifest = load_manifest(out_dir)
    run_id = manifest["meta"]["run_id"]
    wb_cfg = (manifest.get("config") or {}).get("wb") or {}
    dests = list(wb_cfg.get("dests") or DEFAULT_DESTS)

    pool_path = out_dir / "competitor_pool.jsonl"
    if not pool_path.exists():
        raise FileNotFoundError("Нет competitor_pool.jsonl. Сначала Stage F.")

    out_path = out_dir / "competitor_lite.jsonl"
    done = read_jsonl_done_ids(out_path) if resume else set()

    sess = req_session()
    cache_dir = out_dir / ".wb_cache" / "comp_lite"
    ensure_dir(cache_dir)

    for rec in iter_jsonl(pool_path):
        meta = rec.get("meta") or {}
        nm = nm_id_to_str(meta.get("nm_id"))
        if not nm or (resume and nm in done):
            continue
        vendor_code = safe_str(meta.get("vendor_code", ""))
        name = safe_str(meta.get("name", ""))

        by_cluster = ((rec.get("pool") or {}).get("by_cluster")) or {}
        out_clusters: List[dict] = []

        for cluster in CLUSTERS:
            picked = by_cluster.get(cluster) or []
            lite_items: List[dict] = []
            ok_n = 0
            fail_n = 0
            for c in picked:
                cnm = nm_id_to_str(c.get("nm_id"))
                if not cnm:
                    continue
                cpath = cache_dir / f"{cnm}.json"
                card = None
                fetch = {"cache": False}
                if cpath.exists():
                    try:
                        card = read_json(cpath)
                        fetch = {"cache": True, "path": str(cpath)}
                    except Exception:
                        card = None
                if card is None:
                    status, http_code, js, api_url, dest_used = fetch_card_any(sess, cnm, dests=dests, timeout=timeout)
                    parsed = parse_card_v4(js) if js else {}
                    card = {"status": status, "http": http_code, "api_url": api_url, "dest": dest_used, "v4": parsed}
                    write_json(cpath, card)
                    fetch = {"cache": False, "status": status, "http": http_code}
                    time.sleep(max(0.0, float(sleep_s)))
                try:
                    if safe_str(card.get("status")) == "ok" and int(card.get("http") or 0) == 200:
                        ok_n += 1
                    else:
                        fail_n += 1
                except Exception:
                    fail_n += 1
                lite_items.append({"nm_id": cnm, "seed": c, "card": card, "fetch": fetch})
            out_clusters.append({"cluster": cluster, "items": lite_items, "stats": {"ok_n": int(ok_n), "fail_n": int(fail_n)}})

        out_rec = {"meta": make_meta(run_id, "G", nm, vendor_code, name), "clusters": out_clusters}
        append_jsonl(out_path, out_rec)
        try:
            cstats = {x.get("cluster"): x.get("stats") for x in (out_rec.get("clusters") or []) if isinstance(x, dict)}
            ph = cstats.get("phone") or {}
            ty = cstats.get("type") or {}
            stage_sku("G", nm, f"lite phone_ok={ph.get('ok_n')} fail={ph.get('fail_n')} | type_ok={ty.get('ok_n')} fail={ty.get('fail_n')}", level=1)
        except Exception:
            pass

    stage_dbg("G", f"wrote {out_path}")
    return out_path


# =========================
# Stage H: relevance filter (rules + optional LLM)
# =========================

def build_relevance_prompt(owner_nm: str, phone_model: str, cluster: str, items: List[dict]) -> List[dict]:
    sys_msg = (
        "Ты фильтруешь кандидатов-конкурентов по релевантности. "
        "Верни ТОЛЬКО JSON объект: {\"items\":[{\"nm_id\":\"...\",\"label\":\"KEEP|DROP\",\"reason\":\"...\"}, ...]}. "
        "Без markdown. Не выдумывай факты."
    )
    short_items = []
    for it in items[:25]:
        seed = it.get("seed") or {}
        card = (it.get("card") or {}).get("v4") or {}
        title = safe_str((card.get("content") or {}).get("title") or seed.get("name") or "")
        short_items.append({
            "nm_id": nm_id_to_str(it.get("nm_id")),
            "title": title[:140],
            "brand": safe_str((card.get("content") or {}).get("brand") or seed.get("brand") or "")[:40],
        })
    payload = {"owner_nm_id": owner_nm, "phone_model": phone_model, "cluster": cluster, "candidates": short_items}
    return [{"role": "system", "content": sys_msg},
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False)}]

def stage_H_relevance(out_dir: Path, *, resume: bool, use_llm: bool,
                      min_keep_competitors: int, max_keep_competitors: int,
                      llm_provider: str, llm_model: str, llm_base_url: str, llm_timeout: int, llm_max_tokens: int, llm_temperature: float) -> Path:
    manifest = load_manifest(out_dir)
    run_id = manifest["meta"]["run_id"]
    lex = Lexicon(**((manifest.get("config") or {}).get("lexicon") or dataclasses.asdict(load_lexicon(None))))
    scope = scope_from_manifest(manifest)

    intent_map = map_by_nm(out_dir / "intent.jsonl")
    lite_map = map_by_nm(out_dir / "competitor_lite.jsonl")

    out_path = out_dir / "relevance.jsonl"
    done = read_jsonl_done_ids(out_path) if resume else set()

    api_key = llm_api_key(llm_provider) if use_llm else ""

    for sku in scope:
        nm = nm_id_to_str(sku.get("nm_id"))
        if not nm or (resume and nm in done):
            continue
        vendor_code = safe_str(sku.get("vendor_code", ""))
        name = safe_str(sku.get("name", ""))

        intent = (intent_map.get(nm, {}).get("intent") or {})
        phone_model = safe_str(intent.get("phone_model") or "")
        clusters_cfg = intent.get("clusters") or []

        lite = lite_map.get(nm, {})
        clusters = lite.get("clusters") or []
        out_clusters: List[dict] = []
        llm_debug: List[dict] = []

        cfg_map = {safe_str(c.get("cluster")): c for c in clusters_cfg if isinstance(c, dict)}

        for c in clusters:
            cluster = safe_str(c.get("cluster"))
            if cluster not in CLUSTERS:
                continue
            cfg = cfg_map.get(cluster, {})
            must_any_groups = list(cfg.get("must_any_groups") or [])
            ban_terms = list(cfg.get("ban_terms") or list(lex.ban_terms_default))

            items = c.get("items") or []
            keep: List[dict] = []
            drop: List[dict] = []
            borderline: List[dict] = []

            for it in items:
                cnm = nm_id_to_str(it.get("nm_id"))
                card_v4 = ((it.get("card") or {}).get("v4") or {})
                title = safe_str((card_v4.get("content") or {}).get("title") or ((it.get("seed") or {}).get("name") or ""))
                if hit_ban_terms(title, ban_terms):
                    drop.append({"nm_id": cnm, "reason": "ban_term"})
                    continue
                if not pass_must_any_groups(title, must_any_groups):
                    borderline.append({"nm_id": cnm, "title": title[:180]})
                    continue
                keep.append({"nm_id": cnm, "title": title[:180]})

            llm_dbg = {"cluster": cluster, "used": False}
            if use_llm and borderline:
                try:
                    cand_items = [it for it in items if nm_id_to_str(it.get("nm_id")) in {b["nm_id"] for b in borderline}]
                    msgs = build_relevance_prompt(nm, phone_model, cluster, cand_items)
                    parsed, dbg = call_llm_json(
                        provider=llm_provider,
                        model=llm_model,
                        api_key=api_key,
                        base_url=llm_base_url,
                        messages=msgs,
                        timeout_sec=llm_timeout,
                        max_tokens=llm_max_tokens,
                        temperature=llm_temperature,
                        force_json=True,
                    )
                    llm_dbg = {"cluster": cluster, "used": True, **dbg}
                    rows = parsed.get("items", [])
                    if not isinstance(rows, list):
                        rows = []
                    labels = {}
                    for r in rows:
                        if not isinstance(r, dict):
                            continue
                        cid = nm_id_to_str(r.get("nm_id"))
                        lab = safe_str(r.get("label")).strip().upper()
                        if cid and lab in ("KEEP", "DROP"):
                            labels[cid] = (lab, safe_str(r.get("reason"))[:180])
                    for b in borderline:
                        cid = b["nm_id"]
                        if cid in labels and labels[cid][0] == "KEEP":
                            keep.append({"nm_id": cid, "title": b.get("title",""), "reason": labels[cid][1]})
                        else:
                            drop.append({"nm_id": cid, "reason": labels.get(cid, ("DROP","borderline"))[1] if cid in labels else "borderline"})
                    borderline = []
                except Exception as e:
                    llm_dbg = {"cluster": cluster, "used": True, "error": safe_str(e)}

            # Санити: кластер не должен остаться без keep, иначе дальше ты анализируешь вакуум.
            fallback_added = 0
            trimmed_n = 0
            low_confidence = False
            if int(min_keep_competitors or 0) > 0 and len(keep) < int(min_keep_competitors):
                # 1) сначала добираем из оставшихся borderline (если LLM не использовали/упал/оставил хвост)
                while borderline and len(keep) < int(min_keep_competitors):
                    b = borderline.pop(0)
                    keep.append({
                        "nm_id": nm_id_to_str(b.get("nm_id")),
                        "title": safe_str(b.get("title") or "")[:180],
                        "reason": "fallback_borderline",
                    })
                    fallback_added += 1
                    low_confidence = True

                # 2) если всё ещё мало, добираем из items. Сначала тех, кто проходит must_any_groups. Если вообще нечего брать — добираем любые (но помечаем LOW_CONFIDENCE).
                if len(keep) < int(min_keep_competitors):
                    keep_ids = {nm_id_to_str(x.get("nm_id")) for x in keep}

                    def _title_for(itx: dict) -> str:
                        cv4 = ((itx.get("card") or {}).get("v4") or {})
                        return safe_str((cv4.get("content") or {}).get("title") or ((itx.get("seed") or {}).get("name") or ""))

                    cands_good = []
                    cands_any = []
                    for it2 in items:
                        cid = nm_id_to_str(it2.get("nm_id"))
                        if not cid or cid in keep_ids:
                            continue
                        title2 = _title_for(it2)
                        if hit_ban_terms(title2, ban_terms):
                            continue
                        if pass_must_any_groups(title2, must_any_groups):
                            cands_good.append((cid, title2))
                        else:
                            cands_any.append((cid, title2))

                    # сначала "хорошие"
                    for cid, title2 in cands_good:
                        keep.append({"nm_id": cid, "title": title2[:180], "reason": "fallback_any_must"})
                        keep_ids.add(cid)
                        fallback_added += 1
                        if len(keep) >= int(min_keep_competitors):
                            break

                    # если всё равно мало — берём любые (но честно помечаем низкую уверенность)
                    if len(keep) < int(min_keep_competitors) and cands_any:
                        low_confidence = True
                        for cid, title2 in cands_any:
                            keep.append({"nm_id": cid, "title": title2[:180], "reason": "fallback_any"})
                            keep_ids.add(cid)
                            fallback_added += 1
                            if len(keep) >= int(min_keep_competitors):
                                break

            # Ограничение сверху (чтобы не тащить 100500 конкурентов в метрики, если пул раздуло).
            if int(max_keep_competitors or 0) > 0 and len(keep) > int(max_keep_competitors):
                trimmed_n = len(keep) - int(max_keep_competitors)
                keep = keep[: int(max_keep_competitors)]

            if v(2) and (fallback_added or trimmed_n):
                veprint(2, f"[H][DBG] nm={nm} cluster={cluster} keep={len(keep)} fallback_added={fallback_added} trimmed={trimmed_n} borderline_left={len(borderline)}")
            llm_debug.append(llm_dbg)

            out_clusters.append({
                "cluster": cluster,
                "keep": keep,
                "drop": drop,
                "borderline_left": borderline,
                "stats": {"keep_n": len(keep), "drop_n": len(drop), "borderline_n": len(borderline), "fallback_added": int(fallback_added), "trimmed_n": int(trimmed_n), "low_confidence": bool(low_confidence)},
            })

        out_rec = {"meta": make_meta(run_id, "H", nm, vendor_code, name), "clusters": out_clusters, "llm_debug": llm_debug}
        append_jsonl(out_path, out_rec)

    stage_dbg("H", f"wrote {out_path}")
    return out_path


# =========================
# Stage I: market pulse (reviews velocity)
# =========================

def _utc_now() -> datetime:
    return datetime.now(timezone.utc).replace(microsecond=0)

def _parse_dt(x: Any) -> Optional[datetime]:
    if x is None:
        return None
    if isinstance(x, datetime):
        return x if x.tzinfo else x.replace(tzinfo=timezone.utc)
    s = safe_str(x).strip()
    if not s:
        return None
    s = s.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(s)
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except Exception:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(s[:19], fmt).replace(tzinfo=timezone.utc)
            return dt
        except Exception:
            continue
    return None

def _extract_feedback_list(js: dict) -> List[dict]:
    if not isinstance(js, dict):
        return []
    if isinstance(js.get("feedbacks"), list):
        return [x for x in js["feedbacks"] if isinstance(x, dict)]
    data = js.get("data")
    if isinstance(data, dict) and isinstance(data.get("feedbacks"), list):
        return [x for x in data["feedbacks"] if isinstance(x, dict)]
    return []


def _extract_feedback_count(js: dict) -> int:
    if not isinstance(js, dict):
        return 0
    if js.get("feedbackCount") is not None:
        return safe_int(js.get("feedbackCount"), 0) or 0
    data = js.get("data")
    if isinstance(data, dict) and data.get("feedbackCount") is not None:
        return safe_int(data.get("feedbackCount"), 0) or 0
    if js.get("count") is not None:
        return safe_int(js.get("count"), 0) or 0
    return 0

def _fb_key(it: dict, dt: datetime) -> str:
    # Нужен ключ для дедупа между страницами.
    # Если у отзыва есть id — используем его, иначе склеиваем дату+оценку+кусок текста.
    if isinstance(it, dict):
        for k in ("id", "feedbackId", "feedback_id"):
            if it.get(k) is not None:
                return f"id:{safe_str(it.get(k))}"
        rate = it.get("productValuation") or it.get("valuation") or it.get("rate") or ""
        txt = safe_str(it.get("text") or it.get("comment") or "")
        txt = txt.replace("\n", " ").strip()[:80]
        return f"dt:{dt.isoformat()}|r:{safe_str(rate)}|t:{txt}"
    return f"dt:{dt.isoformat()}"

@dataclass
class ReviewSummary:
    imt_id: int
    recent_30: int
    recent_90: int
    days_since_last: Optional[int]
    newest_dt: Optional[str]
    oldest_dt: Optional[str]
    pages_fetched: int
    reviews_seen: int
    source_url: str
    warnings: List[str]



def fetch_review_summary_for_imt_v1_get(
    sess: requests.Session,
    imt_id: int,
    *,
    v1_urls: List[str],
    timeout: int,
    take: int = 100,
    max_pages: int = 20,
    max_skip: int = 5000,
    early_stop_days: int = 90,
    sleep_s: float = 0.15,
) -> Tuple[Optional[ReviewSummary], Dict[str, Any]]:
    """Публичные отзывы через feedbacks1/2.

    Важно:
    - обычно отдаёт свежий хвост отзывов.
    - пагинация через take/skip может поддерживаться, а может игнорироваться.
      Поэтому мы дедупим и останавливаемся, если страница даёт только дубли.
    """
    now = _utc_now()
    cut_30 = now - timedelta(days=30)
    cut_90 = now - timedelta(days=90)

    warnings: List[str] = []
    all_dts: List[datetime] = []
    recent_30 = 0
    recent_90 = 0
    pages_fetched = 0
    reviews_seen = 0

    used_url = ""
    last_info: Dict[str, Any] = {"imt_id": int(imt_id), "mode": "v1_get", "pages": []}

    for tpl in (v1_urls or []):
        used_url = safe_str(tpl).format(imt_id=int(imt_id), rootId=int(imt_id), root_id=int(imt_id))
        if not used_url:
            continue

        seen: set[str] = set()
        pages_fetched = 0
        reviews_seen = 0
        all_dts = []
        recent_30 = 0
        recent_90 = 0

        ok_any = False
        fb_count = 0

        for page in range(int(max_pages)):
            skip = page * int(take)
            if skip > int(max_skip):
                warnings.append("skip_cap_reached")
                break

            params = {"take": int(take), "skip": int(skip)}
            js, info = wb_get_json(sess, used_url, timeout=timeout, params=params, retries=3)
            info = {**info, "page": int(page), "skip": int(skip)}
            last_info["pages"].append(info)

            if js is None:
                if page == 0:
                    try:
                        st = info.get("status")
                        er = info.get("error") or ""
                        warnings.append(f"unreachable:{st}:{er}" if st is not None else f"unreachable:{er}")
                    except Exception:
                        warnings.append("unreachable:first_page")
                    break
                warnings.append("page_fetch_failed")
                break

            ok_any = True
            if page == 0:
                fb_count = _extract_feedback_count(js)

            items = _extract_feedback_list(js)
            pages_fetched += 1
            if not items:
                break

            added = 0
            page_dts: List[datetime] = []
            for it in items:
                dt = _parse_dt(it.get("createdDate") or it.get("createdDateTime") or it.get("date") or it.get("created"))
                if dt is None:
                    continue
                key = _fb_key(it, dt)
                if key in seen:
                    continue
                seen.add(key)
                added += 1

                page_dts.append(dt)
                all_dts.append(dt)
                reviews_seen += 1
                if dt >= cut_30:
                    recent_30 += 1
                if dt >= cut_90:
                    recent_90 += 1

            # если сервер игнорирует skip и снова отдаёт те же отзывы — выходим
            if added == 0:
                warnings.append("pagination_ignored_or_duplicates")
                break

            if page_dts:
                oldest = min(page_dts)
                if oldest < cut_90:
                    break

            time.sleep(max(0.0, float(sleep_s)))

        if ok_any:
            # Частичная история: отзывов всего много, а мы не смогли уйти глубже 90 дней.
            if fb_count and fb_count > reviews_seen and all_dts:
                if min(all_dts) >= cut_90:
                    warnings.append("partial_window_90")
            break

    if not all_dts:
        warn2 = warnings[:]
        if not used_url:
            warn2.append("no_v1_url")
        if any(w.startswith("unreachable") for w in warn2) or ("unreachable" in warn2):
            if "unreachable" not in warn2:
                warn2.append("unreachable")
        else:
            warn2.append("no_reviews")
        return ReviewSummary(
            imt_id=int(imt_id), recent_30=0, recent_90=0, days_since_last=None,
            newest_dt=None, oldest_dt=None, pages_fetched=int(pages_fetched), reviews_seen=int(reviews_seen),
            source_url=used_url, warnings=warn2,
        ), last_info

    newest = max(all_dts)
    oldest = min(all_dts)
    days_since_last = int((_utc_now() - newest).total_seconds() // 86400)

    return ReviewSummary(
        imt_id=int(imt_id), recent_30=int(recent_30), recent_90=int(recent_90),
        days_since_last=days_since_last, newest_dt=newest.isoformat(), oldest_dt=oldest.isoformat(),
        pages_fetched=int(pages_fetched), reviews_seen=int(reviews_seen),
        source_url=used_url, warnings=warnings,
    ), last_info


def fetch_review_summary_for_imt(
    sess: requests.Session,
    imt_id: int,
    *,
    mode: str,
    v1_urls: List[str],
    v1_take: int,
    v1_max_pages: int,
    v1_max_skip: int,
    base_urls: List[str],
    post_take: int,
    post_max_pages: int,
    post_max_skip: int,
    early_stop_days: int,
    timeout: int,
    sleep_s: float,
    allow_post_fallback: bool = True,
) -> Tuple[Optional[ReviewSummary], Dict[str, Any]]:
    """Unified wrapper: try v1_get first (default), then optional legacy POST fallback."""
    m = safe_str(mode).strip().lower() or "v1_get"

    # По умолчанию предпочитаем v1_get даже если в старом манифесте остались public-feedbacks.
    if m in ("v1", "v1_get", "get", "feedbacks1", "feedbacks_v1"):
        summ, dbg = fetch_review_summary_for_imt_v1_get(
            sess, imt_id,
            v1_urls=v1_urls,
            timeout=timeout,
            take=v1_take,
            max_pages=v1_max_pages,
            max_skip=v1_max_skip,
            early_stop_days=early_stop_days,
            sleep_s=sleep_s,
        )
        # Если недоступно, можно пробовать legacy post (если разрешено)
        if summ and ("unreachable" in (summ.warnings or [])) and allow_post_fallback and base_urls:
            summ2, dbg2 = fetch_review_summary_for_imt_post(
                sess, imt_id,
                base_urls=base_urls,
                timeout=timeout,
                take=post_take,
                max_pages=post_max_pages,
                max_skip=post_max_skip,
                early_stop_days=early_stop_days,
                sleep_s=sleep_s,
            )
            return summ2, {"primary": dbg, "fallback": dbg2}
        return summ, dbg

    # forced legacy
    return fetch_review_summary_for_imt_post(
        sess, imt_id,
        base_urls=base_urls,
        timeout=timeout,
        take=post_take,
        max_pages=post_max_pages,
        max_skip=post_max_skip,
        early_stop_days=early_stop_days,
        sleep_s=sleep_s,
    )

def fetch_review_summary_for_imt_post(
    sess: requests.Session,
    imt_id: int,
    *,
    base_urls: List[str],
    timeout: int,
    take: int = 30,
    max_pages: int = 40,
    max_skip: int = 1000,
    early_stop_days: int = 90,
    sleep_s: float = 0.25,
) -> Tuple[Optional[ReviewSummary], Dict[str, Any]]:
    now = _utc_now()
    cut_30 = now - timedelta(days=30)
    cut_90 = now - timedelta(days=90)

    warnings: List[str] = []
    all_dts: List[datetime] = []
    recent_30 = 0
    recent_90 = 0
    pages_fetched = 0
    reviews_seen = 0

    used_url = ""
    last_info: Dict[str, Any] = {}
    success_any = False

    for base in base_urls:
        ok_any = False
        pages_fetched = 0
        reviews_seen = 0
        all_dts = []
        recent_30 = 0
        recent_90 = 0
        used_url = base
        last_info = {"imt_id": imt_id, "base_url": base, "pages": []}

        for page in range(max_pages):
            skip = page * take
            if skip > max_skip:
                warnings.append("skip_cap_reached")
                break

            payload = {"imtId": int(imt_id), "take": int(take), "skip": int(skip), "order": "dateDesc"}
            js, info = post_json(sess, base, payload, timeout=timeout, retries=3)
            last_info["pages"].append(info)

            if js is None:
                if page == 0:
                    try:
                        st = info.get("status")
                        er = info.get("error") or ""
                        warnings.append(f"unreachable:{st}:{er}" if st is not None else f"unreachable:{er}")
                    except Exception:
                        warnings.append("unreachable:first_page")
                    break
                warnings.append("page_fetch_failed")
                break

            items = _extract_feedback_list(js)
            pages_fetched += 1
            if not items:
                ok_any = True
                break

            page_dts: List[datetime] = []
            for it in items:
                dt = _parse_dt(it.get("createdDate") or it.get("createdDateTime") or it.get("date") or it.get("created"))
                if dt is None:
                    continue
                page_dts.append(dt)
                all_dts.append(dt)
                reviews_seen += 1
                if dt >= cut_30:
                    recent_30 += 1
                if dt >= cut_90:
                    recent_90 += 1

            ok_any = True
            if page_dts:
                oldest = min(page_dts)
                if oldest < cut_90:
                    break

            time.sleep(max(0.0, float(sleep_s)))

        if ok_any:
            success_any = True
            break

    if not all_dts:
        warn2 = warnings[:]
        if not used_url:
            warn2.append("no_base_url_worked")
        if not success_any:
            warn2.append("unreachable")
        else:
            warn2.append("no_reviews")
        return ReviewSummary(
            imt_id=int(imt_id), recent_30=0, recent_90=0, days_since_last=None,
            newest_dt=None, oldest_dt=None, pages_fetched=pages_fetched, reviews_seen=reviews_seen,
            source_url=used_url, warnings=warn2,
        ), last_info

    newest = max(all_dts)
    oldest = min(all_dts)
    days_since_last = int((_utc_now() - newest).total_seconds() // 86400)

    return ReviewSummary(
        imt_id=int(imt_id), recent_30=int(recent_30), recent_90=int(recent_90),
        days_since_last=days_since_last, newest_dt=newest.isoformat(), oldest_dt=oldest.isoformat(),
        pages_fetched=int(pages_fetched), reviews_seen=int(reviews_seen), source_url=used_url,
        warnings=warnings,
    ), last_info

def _median_int(vals: List[int]) -> Optional[float]:
    if not vals:
        return None
    vs = sorted(vals)
    n = len(vs)
    mid = n // 2
    if n % 2 == 1:
        return float(vs[mid])
    return (vs[mid - 1] + vs[mid]) / 2.0

def stage_I_pulse(out_dir: Path, *, timeout: int, sleep_s: float, resume: bool, strict: bool = False) -> Path:
    manifest = load_manifest(out_dir)
    run_id = manifest["meta"]["run_id"]
    cfg = manifest.get("config") or {}
    
    reviews_cfg = cfg.get("reviews") or DEFAULT_REVIEWS_CFG

    # Backward-compat: старые манифесты могут иметь только base_urls (public-feedbacks).
    # Но по умолчанию используем v1_get (feedbacks1/2), иначе снова словим 403 и «рынок мёртв».
    mode = safe_str(reviews_cfg.get("mode") or "").strip().lower() or "v1_get"
    v1_urls = list(reviews_cfg.get("v1_urls") or DEFAULT_REVIEWS_CFG.get("v1_urls") or [])
    allow_post_fallback = bool(reviews_cfg.get("allow_post_fallback", True))
    base_urls = list(reviews_cfg.get("base_urls") or DEFAULT_REVIEWS_CFG.get("base_urls") or [])

    v1_take = int(reviews_cfg.get("v1_take") or 100)
    v1_max_pages = int(reviews_cfg.get("v1_max_pages") or 20)
    v1_max_skip = int(reviews_cfg.get("v1_max_skip") or 5000)

    post_take = int(reviews_cfg.get("post_take") or 30)
    post_max_pages = int(reviews_cfg.get("post_max_pages") or 40)
    post_max_skip = int(reviews_cfg.get("post_max_skip") or 1000)

    rel_path = out_dir / "relevance.jsonl"
    if not rel_path.exists():
        raise FileNotFoundError("Нет relevance.jsonl. Сначала Stage H.")

    out_path = out_dir / "market_pulse.jsonl"
    err_path = out_dir / "market_pulse_errors.jsonl"
    done = read_jsonl_done_ids(out_path) if resume else set()

    cache_dir = out_dir / ".wb_cache" / "reviews"
    ensure_dir(cache_dir)
    sess = req_session()

    def cache_path_for(imt_id: int) -> Path:
        return cache_dir / f"imt_{int(imt_id)}.json"

    def cache_is_fresh(p: Path, ttl_hours: int) -> bool:
        if not p.exists():
            return False
        try:
            js = read_json(p)
            dt = _parse_dt(js.get("fetched_at"))
            if not dt:
                return False
            age = (_utc_now() - dt).total_seconds() / 3600.0
            return age <= float(ttl_hours)
        except Exception:
            return False
    ttl = int(reviews_cfg.get("cache_ttl_hours") or 24)
    early_stop_days = int(reviews_cfg.get("early_stop_days") or 90)

    rel_rows = iter_jsonl(rel_path)
    lite_map = map_by_nm(out_dir / "competitor_lite.jsonl")

    pbar = SimpleProgress(total=len(rel_rows), desc="[I] market_pulse")
    for rec in rel_rows:
        meta = rec.get("meta") or {}
        nm = nm_id_to_str(meta.get("nm_id"))
        if not nm or (resume and nm in done):
            pbar.update(1)
            continue
        vendor_code = safe_str(meta.get("vendor_code", ""))
        name = safe_str(meta.get("name", ""))

        clusters = rec.get("clusters") or []
        out_clusters: List[dict] = []
        errors: List[dict] = []

        lite = lite_map.get(nm, {})
        lite_clusters = {safe_str(x.get("cluster")): x for x in (lite.get("clusters") or []) if isinstance(x, dict)}

        for c in clusters:
            cluster = safe_str(c.get("cluster"))
            if cluster not in CLUSTERS:
                continue
            keep_ids = {nm_id_to_str(x.get("nm_id")) for x in (c.get("keep") or [])}
            items = (lite_clusters.get(cluster, {}).get("items") or [])
            imt_ids = []
            for it in items:
                cid = nm_id_to_str(it.get("nm_id"))
                if cid and cid in keep_ids:
                    v4 = ((it.get("card") or {}).get("v4") or {})
                    imt = safe_int((v4.get("ids") or {}).get("imt_id")) or safe_int(((it.get("seed") or {}).get("imt_id")))
                    if imt:
                        imt_ids.append(imt)
            imt_ids = sorted(set(imt_ids))

            summaries_all: List[ReviewSummary] = []
            summaries_ok: List[ReviewSummary] = []
            unreachable_n = 0
            no_reviews_n = 0

            for imt_id in imt_ids:
                cp = cache_path_for(imt_id)
                summ: Optional[ReviewSummary] = None
                if cache_is_fresh(cp, ttl):
                    try:
                        js = read_json(cp)
                        summ = ReviewSummary(
                            imt_id=int(imt_id),
                            recent_30=int(js.get("recent_30") or 0),
                            recent_90=int(js.get("recent_90") or 0),
                            days_since_last=js.get("days_since_last"),
                            newest_dt=js.get("newest_dt"),
                            oldest_dt=js.get("oldest_dt"),
                            pages_fetched=int(js.get("pages_fetched") or 0),
                            reviews_seen=int(js.get("reviews_seen") or 0),
                            source_url=safe_str(js.get("source_url") or ""),
                            warnings=list(js.get("warnings") or []),
                        )
                        if summ and ("no_reviews_or_unreachable" in (summ.warnings or [])):
                            summ = None
                    except Exception:
                        summ = None

                if summ is None:
                    try:
                        summ, _debug = fetch_review_summary_for_imt(
                            sess, imt_id,
                            mode=mode,
                            v1_urls=v1_urls,
                            v1_take=v1_take,
                            v1_max_pages=v1_max_pages,
                            v1_max_skip=v1_max_skip,
                            base_urls=base_urls,
                            post_take=post_take,
                            post_max_pages=post_max_pages,
                            post_max_skip=post_max_skip,
                            early_stop_days=early_stop_days,
                            timeout=timeout,
                            sleep_s=sleep_s,
                            allow_post_fallback=allow_post_fallback,
                        )
                        write_json(cp, {
                            "fetched_at": utc_now_iso(),
                            "imt_id": int(imt_id),
                            "recent_30": summ.recent_30 if summ else 0,
                            "recent_90": summ.recent_90 if summ else 0,
                            "days_since_last": summ.days_since_last if summ else None,
                            "newest_dt": summ.newest_dt if summ else None,
                            "oldest_dt": summ.oldest_dt if summ else None,
                            "pages_fetched": summ.pages_fetched if summ else 0,
                            "reviews_seen": summ.reviews_seen if summ else 0,
                            "source_url": summ.source_url if summ else "",
                            "warnings": summ.warnings if summ else ["unreachable"],
                        })
                    except Exception as e:
                        errors.append({"cluster": cluster, "imt_id": int(imt_id), "error": safe_str(e)})
                        summ = None

                if summ:
                    summaries_all.append(summ)
                    if v(2):
                        veprint(2, f"[I][imt] nm={nm} cluster={cluster} imt={summ.imt_id} r30={summ.recent_30} r90={summ.recent_90} dsl={summ.days_since_last} warn={(summ.warnings or [])} src={summ.source_url}")
                    wset = set((summ.warnings or []))
                    if "unreachable" in wset or any(str(x).startswith("unreachable:") for x in wset):
                        unreachable_n += 1
                    elif "no_reviews" in wset:
                        no_reviews_n += 1
                        summaries_ok.append(summ)
                    else:
                        summaries_ok.append(summ)

            r30 = [s.recent_30 for s in summaries_ok]
            r90 = [s.recent_90 for s in summaries_ok]
            dsl = [s.days_since_last for s in summaries_ok if s.days_since_last is not None]
            ok_n = len(summaries_ok)

            agg = {
                "imt_n": len(imt_ids),
                "imt_with_data_n": ok_n,
                "imt_unreachable_n": int(unreachable_n),
                "imt_no_reviews_n": int(no_reviews_n),
                "unreachable_share": round((unreachable_n / len(imt_ids)) if imt_ids else 0.0, 4),
                "recent_30_median": _median_int(r30),
                "recent_90_median": _median_int(r90),
                "days_since_last_median": _median_int([int(x) for x in dsl]) if dsl else None,
                "days_since_last_min": min(dsl) if dsl else None,
            }

            out_clusters.append({
                "cluster": cluster,
                "imt_summaries": [{"imt_id": s.imt_id, "recent_30": s.recent_30, "recent_90": s.recent_90, "days_since_last": s.days_since_last, "warnings": s.warnings} for s in summaries_all],
                "agg": agg,
            })

            stage_sku("I", nm, f"cluster={cluster} imt={len(imt_ids)} ok={ok_n} unreachable={unreachable_n} no_reviews={no_reviews_n}", level=1, err=(ok_n==0 and unreachable_n>0))

            if strict and imt_ids and ok_n == 0 and unreachable_n > 0:
                raise RuntimeError(f"[I] Reviews unreachable for nm={nm} cluster={cluster} (imt={len(imt_ids)}). \nПохоже WB режет IP/сеть. Для WB стадий держи VPN OFF, а для LLM включай только на LLM стадиях.\n")

        out_rec = {"meta": make_meta(run_id, "I", nm, vendor_code, name), "clusters": out_clusters}
        append_jsonl(out_path, out_rec)
        if errors:
            append_jsonl(err_path, {"meta": out_rec["meta"], "errors": errors})

        pbar.update(1)
    pbar.close()
    stage_dbg("I", f"wrote {out_path}")
    return out_path


# =========================
# Stage J: supply/structure
# =========================

def quantile(sorted_vals: List[float], q: float) -> Optional[float]:
    if not sorted_vals:
        return None
    n = len(sorted_vals)
    if n == 1:
        return float(sorted_vals[0])
    q = max(0.0, min(1.0, float(q)))
    pos = q * (n - 1)
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return float(sorted_vals[lo])
    w = pos - lo
    return float(sorted_vals[lo] * (1 - w) + sorted_vals[hi] * w)

def median(vals: List[float]) -> Optional[float]:
    if not vals:
        return None
    s = sorted(vals)
    return quantile(s, 0.5)

def trim(vals: List[float], trim_ratio: float = 0.10) -> List[float]:
    if not vals:
        return []
    s = sorted(vals)
    n = len(s)
    k = int(math.floor(n * float(trim_ratio)))
    if 2 * k >= n:
        return s
    return s[k:n-k]

def hhi(shares: List[float]) -> float:
    return float(sum((x * x) for x in shares))

def stage_J_supply(out_dir: Path, *, resume: bool = False) -> Path:
    manifest = load_manifest(out_dir)
    run_id = manifest["meta"]["run_id"]
    thr = ((manifest.get("config") or {}).get("supply_thresholds") or DEFAULT_SUPPLY_THRESHOLDS)

    rel_map = map_by_nm(out_dir / "relevance.jsonl")
    lite_map = map_by_nm(out_dir / "competitor_lite.jsonl")

    out_path = out_dir / "supply_structure.jsonl"
    done = read_jsonl_done_ids(out_path) if resume else set()

    scope = scope_from_manifest(manifest)

    for sku in scope:
        nm = nm_id_to_str(sku.get("nm_id"))
        if resume and nm in done:
            continue
        vendor_code = safe_str(sku.get("vendor_code", ""))
        name = safe_str(sku.get("name", ""))

        rel = rel_map.get(nm, {})
        lite = lite_map.get(nm, {})

        lite_clusters = {safe_str(x.get("cluster")): x for x in (lite.get("clusters") or []) if isinstance(x, dict)}
        rel_clusters = {safe_str(x.get("cluster")): x for x in (rel.get("clusters") or []) if isinstance(x, dict)}

        out_clusters: List[dict] = []

        for cluster in CLUSTERS:
            keep_ids = {nm_id_to_str(x.get("nm_id")) for x in (rel_clusters.get(cluster, {}).get("keep") or [])}
            items = lite_clusters.get(cluster, {}).get("items") or []

            prices: List[float] = []
            sellers: List[str] = []
            qtys: List[float] = []

            for it in items:
                cid = nm_id_to_str(it.get("nm_id"))
                if cid not in keep_ids:
                    continue
                v4 = ((it.get("card") or {}).get("v4") or {})
                seed = it.get("seed") or {}
                pr_u = (v4.get("pricing") or {}).get("salePriceU") or (v4.get("pricing") or {}).get("priceU")
                pr = safe_int(pr_u, None)
                if pr is not None:
                    pr_rub = pr / 100.0 if pr > 1000 else float(pr)
                    prices.append(float(pr_rub))
                else:
                    pr2 = seed.get("price_rub")
                    if pr2 is not None:
                        prices.append(float(pr2))
                sellers.append(seller_key(seed))
                tq = safe_float((v4.get("metrics") or {}).get("total_quantity"), None)
                if tq is not None:
                    qtys.append(float(tq))

            prices_sorted = sorted(prices)
            p10 = quantile(prices_sorted, 0.10) if prices_sorted else None
            p50 = quantile(prices_sorted, 0.50) if prices_sorted else None
            p90 = quantile(prices_sorted, 0.90) if prices_sorted else None
            trimmed = trim(prices, 0.10)
            p_trim = median(trimmed)
            outlier_rate = 0.0
            if p50 and prices:
                far = sum(1 for x in prices if x < 0.6 * p50 or x > 1.8 * p50)
                outlier_rate = far / len(prices)

            seller_counts: Dict[str, int] = {}
            for s in sellers:
                seller_counts[s] = seller_counts.get(s, 0) + 1
            total = sum(seller_counts.values()) or 1
            shares = [c / total for c in seller_counts.values()]
            top1_share = max(shares) if shares else 0.0
            hhi_val = hhi(shares) if shares else 0.0
            unique_sellers = len(seller_counts)

            flags: List[str] = []

            dump_cfg = thr.get("dumping") or {}
            min_unique = int(dump_cfg.get("min_unique_sellers", 10))
            p10_ratio_lt = float(dump_cfg.get("p10_ratio_lt", 0.75))
            out_rate_gte = float(dump_cfg.get("outlier_rate_gte", 0.10))
            min_price_n = int(dump_cfg.get("min_price_n_or", 12))
            if len(prices) >= min_price_n and p10 is not None and p50 is not None and unique_sellers >= min_unique:
                if (p10 / p50) < p10_ratio_lt or outlier_rate >= out_rate_gte:
                    flags.append("DUMPING_PRESSURE")

            mono_cfg = thr.get("monopoly") or {}
            top1_gte = float(mono_cfg.get("top1_share_gte", 0.35))
            hhi_gte = float(mono_cfg.get("hhi_gte", 0.25))
            if unique_sellers >= 1:
                if top1_share >= top1_gte or hhi_val >= hhi_gte:
                    flags.append("MONOPOLY_DANGER")

            rel_stats = (rel_clusters.get(cluster, {}).get("stats") or {})
            if bool(rel_stats.get("low_confidence")):
                flags.append("LOW_CONFIDENCE")

            if len(prices) < int(thr.get("min_n", 8)):
                flags.append("LOW_CONFIDENCE")

            out_clusters.append({
                "cluster": cluster,
                "n": len(prices),
                "price": {"p10": p10, "p50": p50, "p90": p90, "trimmed_median": p_trim, "outlier_rate": round(outlier_rate, 4)},
                "sellers": {"unique": unique_sellers, "top1_share": round(top1_share, 4), "hhi": round(hhi_val, 4)},
                "stock": {"qty_median": median(qtys), "qty_n": len(qtys)},
                "flags": flags,
            })

        out_rec = {"meta": make_meta(run_id, "J", nm, vendor_code, name), "clusters": out_clusters}
        append_jsonl(out_path, out_rec)

    stage_dbg("J", f"wrote {out_path}")
    return out_path


# =========================
# Stage K: cluster verdicts
# =========================

def pulse_to_status(cluster_row: dict, rules: Dict[str, Any]) -> Tuple[str, str, List[str]]:
    flags: List[str] = []
    rules = rules or DEFAULT_PULSE_RULES

    alive_r30_med_gte = float(rules.get("alive_r30_med_gte", 1.0))
    alive_days_since_last_min_lte = int(rules.get("alive_days_since_last_min_lte", 21))
    alive_r90_med_gte = float(rules.get("alive_r90_med_gte", 3.0))
    alive_days_since_last_med_lte = int(rules.get("alive_days_since_last_med_lte", 45))

    slow_r90_med_gte = float(rules.get("slow_r90_med_gte", 1.0))
    slow_days_since_last_med_lte = int(rules.get("slow_days_since_last_med_lte", 90))
    slow_days_since_last_min_lte = int(rules.get("slow_days_since_last_min_lte", 120))

    imt_n = int(cluster_row.get("imt_n") or 0)
    imt_ok = int(cluster_row.get("imt_with_data_n") or 0)
    unreach_n = int(cluster_row.get("imt_unreachable_n") or 0)
    unreach_share = float(cluster_row.get("unreachable_share") or 0.0)

    if imt_ok >= 10:
        conf = "HIGH"
    elif imt_ok >= 6:
        conf = "MEDIUM"
    else:
        conf = "LOW"
        flags.append("LOW_CONFIDENCE")

    if imt_ok == 0:
        if imt_n > 0 and (unreach_n > 0 or unreach_share >= 0.25):
            flags.append("REVIEWS_UNREACHABLE")
            return "UNKNOWN", "LOW", flags
        return "DEAD", conf, flags + ["NO_REVIEW_DATA"]

    if unreach_share >= 0.5:
        flags.append("REVIEWS_UNREACHABLE")
        flags.append("LOW_CONFIDENCE")
        return "UNKNOWN", "LOW", flags

    r30_med = cluster_row.get("recent_30_median")
    r90_med = cluster_row.get("recent_90_median")
    dsl_med = cluster_row.get("days_since_last_median")
    dsl_min = cluster_row.get("days_since_last_min")

    if (r30_med is not None and float(r30_med) >= alive_r30_med_gte and dsl_min is not None and int(dsl_min) <= alive_days_since_last_min_lte) or \
       (r90_med is not None and float(r90_med) >= alive_r90_med_gte and dsl_med is not None and float(dsl_med) <= alive_days_since_last_med_lte):
        return "ALIVE", conf, flags

    if (r90_med is not None and float(r90_med) >= slow_r90_med_gte) and \
       (dsl_med is not None and float(dsl_med) <= slow_days_since_last_med_lte) and \
       (dsl_min is not None and int(dsl_min) <= slow_days_since_last_min_lte):
        return "SLOW", conf, flags

    if dsl_min is None and dsl_med is None and (r90_med is None or float(r90_med) == 0.0) and (r30_med is None or float(r30_med) == 0.0):
        flags.append("LOW_CONFIDENCE")
        return "UNKNOWN", "LOW", flags

    return "DEAD", conf, flags


def stage_K_cluster_verdicts(out_dir: Path, *, resume: bool = False) -> Path:
    manifest = load_manifest(out_dir)
    run_id = manifest["meta"]["run_id"]
    rules = (manifest.get("config") or {}).get("pulse_rules") or DEFAULT_PULSE_RULES

    pulse_map = map_by_nm(out_dir / "market_pulse.jsonl")
    supply_map = map_by_nm(out_dir / "supply_structure.jsonl")

    out_path = out_dir / "cluster_verdicts.jsonl"
    done = read_jsonl_done_ids(out_path) if resume else set()

    scope = scope_from_manifest(manifest)

    for sku in scope:
        nm = nm_id_to_str(sku.get("nm_id"))
        if resume and nm in done:
            continue
        vendor_code = safe_str(sku.get("vendor_code", ""))
        name = safe_str(sku.get("name", ""))

        pulse = pulse_map.get(nm, {})
        supply = supply_map.get(nm, {})

        pulse_clusters = {safe_str(x.get("cluster")): x for x in (pulse.get("clusters") or []) if isinstance(x, dict)}
        supply_clusters = {safe_str(x.get("cluster")): x for x in (supply.get("clusters") or []) if isinstance(x, dict)}

        verdicts: List[dict] = []
        for cluster in CLUSTERS:
            p_row = (pulse_clusters.get(cluster, {}).get("agg") or {})
            s_row = supply_clusters.get(cluster, {})
            status, conf, flags = pulse_to_status(p_row, rules)
            stage_sku("K", nm, f"cluster={cluster} status={status} conf={conf} flags={flags}", level=1)

            issues: List[str] = []
            for fl in (s_row.get("flags") or []):
                if fl in ("DUMPING_PRESSURE", "MONOPOLY_DANGER", "LOW_CONFIDENCE"):
                    issues.append(fl)

            verdicts.append({
                "cluster": cluster,
                "market_status": status,
                "confidence": conf,
                "pulse": p_row,
                "supply": s_row,
                "issues": sorted(set(issues + flags)),
            })

        out_rec = {"meta": make_meta(run_id, "K", nm, vendor_code, name), "cluster_verdicts": verdicts}
        append_jsonl(out_path, out_rec)

    stage_dbg("K", f"wrote {out_path}")
    return out_path


# =========================
# Stage L: final decision
# =========================

def karma_is_toxic(rating: Optional[float], feedbacks: Optional[int], *, min_fb: int = 30, min_rating: float = 3.7) -> bool:
    if rating is None or feedbacks is None:
        return False
    return int(feedbacks) >= int(min_fb) and float(rating) < float(min_rating)

def stage_L_decisions(out_dir: Path, *, resume: bool = False) -> Path:
    manifest = load_manifest(out_dir)
    run_id = manifest["meta"]["run_id"]
    decision_cfg = ((manifest.get("config") or {}).get("decision") or DEFAULT_DECISION_CFG)
    intent_map = map_by_nm(out_dir / "intent.jsonl")
    verdict_map = map_by_nm(out_dir / "cluster_verdicts.jsonl")

    out_path = out_dir / "decisions.jsonl"
    done = read_jsonl_done_ids(out_path) if resume else set()

    scope = scope_from_manifest(manifest)

    for sku in scope:
        nm = nm_id_to_str(sku.get("nm_id"))
        if resume and nm in done:
            continue
        vendor_code = safe_str(sku.get("vendor_code", ""))
        name = safe_str(sku.get("name", ""))

        intent = (intent_map.get(nm, {}).get("intent") or {})
        karma = (intent.get("karma") or {})
        rating = safe_float(karma.get("rating"), None)
        feedbacks = safe_int(karma.get("feedbacks"), None)

        vrec = verdict_map.get(nm, {})
        rows = vrec.get("cluster_verdicts") or []
        by = {safe_str(r.get("cluster")): r for r in rows if isinstance(r, dict)}

        phone_status = safe_str((by.get("phone") or {}).get("market_status") or "DEAD")
        type_status = safe_str((by.get("type") or {}).get("market_status") or "DEAD")

        intent = (intent_map.get(nm, {}).get("intent") or {})
        phone_model = safe_str(intent.get("phone_model") or "")

        risk_flags: List[str] = []
        backlog: List[str] = []
        rationale: List[str] = []
        force_rework = False

        if phone_status == "DEAD":
            verdict = "DROP"
            risk_flags.append("PHONE_MARKET_DEAD")
            rationale.append("Общий рынок чехлов по модели выглядит мёртвым по пульсу отзывов конкурентов.")
        elif phone_status == "UNKNOWN":
            verdict = "REVIVE_REWORK"
            risk_flags.append("PHONE_MARKET_UNKNOWN")
            risk_flags.append("LOW_CONFIDENCE")
            risk_flags.append("REVIEWS_UNREACHABLE")
            rationale.append("По рынку модели нет надёжных данных по отзывам (эндпоинт/сеть недоступны). Нельзя честно сказать ALIVE/DEAD.")
            backlog.append("Пере-запустить Stage I (reviews) с корректной сетью (VPN OFF/РФ IP) или обновить эндпоинт отзывов WB.")
        else:
            allow_type_check = True
            if not phone_model.strip():
                allow_type_check = False
                risk_flags.append("NO_PHONE_MODEL")
                risk_flags.append("LOW_CONFIDENCE")
                rationale.append("Модель телефона не извлечена из своей карточки. Оценка type-market (TPU+карман) пропущена, чтобы не ловить мусор по всем моделям.")

            if allow_type_check and type_status == "UNKNOWN":
                risk_flags.append("TYPE_MARKET_UNKNOWN")
                risk_flags.append("LOW_CONFIDENCE")
                risk_flags.append("REVIEWS_UNREACHABLE")
                rationale.append("Type-market (TPU+карман) не удалось надёжно оценить по отзывам конкурентов. Дальше решение будет с пониженной уверенностью.")
                force_rework = True

            if allow_type_check and type_status == "DEAD":
                verdict = "DROP"
                risk_flags.append("TYPE_MARKET_DEAD")
                rationale.append("Рынок TPU+карман для этой модели выглядит мёртвым. Оживлять именно этот тип бессмысленно.")
                if bool(decision_cfg.get("alt_strategy_on_type_dead", True)):
                    risk_flags.append("ALT_STRATEGY_OTHER_TYPE")
                    backlog.append("Рассмотреть другой тип чехла для этой модели (без кармана, другой материал/форм‑фактор) вместо TPU+карман.")
                    rationale.append("При живом общем рынке модели можно попробовать альтернативный тип, а не хоронить SKU целиком.")
            else:
                if karma_is_toxic(
                    rating, feedbacks,
                    min_fb=int(decision_cfg.get("karma_min_feedbacks", 30) or 30),
                    min_rating=float(decision_cfg.get("karma_min_rating", 3.7) or 3.7)
                ):
                    verdict = "CLONE_NEW_CARD"
                    risk_flags.append("TOXIC_KARMA")
                    rationale.append("У текущей карточки токсичная карма (низкий рейтинг при достаточных отзывах). Лучше клонировать в новую карточку.")
                else:
                    issues = set((by.get("phone") or {}).get("issues") or [])
                    if allow_type_check:
                        issues |= set((by.get("type") or {}).get("issues") or [])
                    else:
                        issues.add("LOW_CONFIDENCE")

                    if "DUMPING_PRESSURE" in issues or "MONOPOLY_DANGER" in issues:
                        verdict = "REVIVE_REWORK"
                        risk_flags.extend(sorted(issues))
                        rationale.append("Рынок живой, но есть риски структуры (демпинг/монополия). Нужна аккуратная стратегия.")
                    else:
                        verdict = "REVIVE_FAST"
                        rationale.append("Рынок живой и без жёстких флагов. Можно оживлять быстро (SEO+контент+прайс).")
                        if force_rework:
                            verdict = "REVIVE_REWORK"
                            risk_flags.append("LOW_CONFIDENCE")
                            rationale.append("Из-за UNKNOWN по части рынков понижаю до REVIVE_REWORK (нужна ручная проверка/повтор Stage I).")
        if verdict in ("REVIVE_FAST", "REVIVE_REWORK", "CLONE_NEW_CARD"):
            backlog += [
                "Собрать семантику под модель телефона (общий рынок) и под TPU+карман (тип).",
                "Проверить конкурентов: фото, заголовок, характеристики, УТП, наличие видео.",
                "Сделать новый заголовок + описание без переспама, но с ключами.",
                "Заполнить характеристики под фильтры WB (материал TPU, карман, совместимость).",
            ]
        if verdict == "REVIVE_REWORK":
            backlog += [
                "Проверить ценовой коридор (p10/p50/p90) и выставить цену без демпинговой истерики.",
                "Продумать дифференциацию (цвета/комплектация/фото), если рынок забит.",
            ]
        if verdict == "CLONE_NEW_CARD":
            backlog += ["Создать новую карточку и перенести лучшее, но с чистой кармой."]

        out_rec = {
            "meta": make_meta(run_id, "L", nm, vendor_code, name),
            "decision": {
                "verdict": verdict,
                "risk_flags": sorted(set(risk_flags)),
                "backlog": backlog[:18],
                "rationale": rationale[:8],
                "market": {"phone": phone_status, "type": type_status},
                "karma": {"rating": rating, "feedbacks": feedbacks},
            },
        }
        append_jsonl(out_path, out_rec)
        stage_sku("L", nm, f"verdict={verdict} flags={sorted(set(risk_flags))}", level=1, err=(verdict=="DROP"))

    stage_dbg("L", f"wrote {out_path}")
    return out_path


# =========================
# Stage M: reports (XLSX + HTML)
# =========================

def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col[:2000]:
            try:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(60, max(10, max_len + 2))

def write_reports(out_dir: Path, args: Optional[argparse.Namespace] = None, *, title: str = "WB Necromancer v2 Report") -> Tuple[Path, Path]:
    manifest = load_manifest(out_dir)
    run_id = manifest["meta"]["run_id"]
    scope = scope_from_manifest(manifest)

    intent_map = map_by_nm(out_dir / "intent.jsonl")
    cluster_map = map_by_nm(out_dir / "cluster_verdicts.jsonl")
    decision_map = map_by_nm(out_dir / "decisions.jsonl")

    rows: List[Dict[str, Any]] = []
    for sku in scope:
        nm = nm_id_to_str(sku.get("nm_id"))
        vc = safe_str(sku.get("vendor_code", ""))
        name = safe_str(sku.get("name", ""))

        intent = (intent_map.get(nm, {}).get("intent") or {})
        pm = safe_str(intent.get("phone_model") or "")
        karma = intent.get("karma") or {}
        rating = karma.get("rating")
        feedbacks = karma.get("feedbacks")

        cv = cluster_map.get(nm, {}).get("cluster_verdicts") or []
        cv_by = {safe_str(r.get("cluster")): r for r in cv if isinstance(r, dict)}

        dec = (decision_map.get(nm, {}).get("decision") or {})
        verdict = safe_str(dec.get("verdict") or "")
        flags = ", ".join(dec.get("risk_flags") or [])
        rationale = " | ".join(dec.get("rationale") or [])
        backlog = "\n".join([f"• {x}" for x in (dec.get("backlog") or [])])

        def _ms(cluster: str) -> str:
            return safe_str((cv_by.get(cluster) or {}).get("market_status") or "")
        def _issues(cluster: str) -> str:
            return ", ".join((cv_by.get(cluster) or {}).get("issues") or [])

        rows.append({
            "nm_id": nm,
            "vendor_code": vc,
            "name": name,
            "phone_model": pm,
            "karma_rating": rating,
            "karma_feedbacks": feedbacks,
            "market_phone": _ms("phone"),
            "market_type": _ms("type"),
            "issues_phone": _issues("phone"),
            "issues_type": _issues("type"),
            "verdict": verdict,
            "risk_flags": flags,
            "rationale": rationale,
            "backlog": backlog,
            "url": wb_product_url(nm),
        })


    # ---------- опц LLM: делаем человеческую выжимку, чтобы отчёт был не только для робота ----------
    exec_summary: Optional[dict] = None
    exec_path = out_dir / "exec_summary.json"
    if args is not None and getattr(args, "use_llm_m", False):
        try:
            from collections import Counter
            vc = Counter()
            mp = Counter()
            mt = Counter()
            flags_c = Counter()
            for r in rows:
                vc[safe_str(r.get("verdict") or "UNKNOWN")] += 1
                mp[safe_str(r.get("market_phone") or "UNKNOWN")] += 1
                mt[safe_str(r.get("market_type") or "UNKNOWN")] += 1
                fl = safe_str(r.get("risk_flags") or "")
                for f in [x.strip() for x in fl.split(",") if x.strip()]:
                    flags_c[f] += 1

            facts = {
                "run_id": run_id,
                "total": len(rows),
                "counts_by_verdict": dict(vc),
                "market_phone": dict(mp),
                "market_type": dict(mt),
                "top_risk_flags": [{"flag": k, "count": v} for k, v in flags_c.most_common(12)],
            }

            prompt = (
                "Сделай короткую и понятную выжимку отчёта для человека. "
                "ВАЖНО: не выдумывай числа и факты, используй только JSON ниже. "
                "Вывод строго JSON без текста вокруг.\n\n"
                "Схема вывода:\n"
                "{\n"
                "  \"title\": str,\n"
                "  \"summary_md\": str,\n"
                "  \"key_points\": [str],\n"
                "  \"risks\": [str],\n"
                "  \"next_steps\": [str]\n"
                "}\n\n"
                "Факты (JSON):\n" + json.dumps(facts, ensure_ascii=False)
            )

            messages = [
                {"role": "system", "content": "Ты аналитик маркетплейса WB. Пиши по делу, без воды. Числа только из фактов."},
                {"role": "user", "content": prompt},
            ]

            key = llm_api_key(args.llm_provider)
            resp, meta = call_llm_json(
                provider=args.llm_provider,
                model=args.llm_model,
                api_key=key,
                messages=messages,
                base_url=args.llm_base_url,
                timeout_sec=int(args.llm_timeout),
                max_tokens=int(getattr(args, "llm_max_tokens_m", 1100) or 1100),
                temperature=float(args.llm_temperature),
                force_json=True,
            )
            if isinstance(resp, dict) and resp:
                exec_summary = resp
                exec_summary["_meta"] = meta
                exec_path.write_text(json.dumps(exec_summary, ensure_ascii=False, indent=2), encoding="utf-8")
                stage_dbg("M", f"LLM exec summary -> {exec_path}")
        except Exception as e:
            stage_warn("M", f"LLM exec summary failed: {safe_str(e)}")

    out_xlsx = out_dir / "WB_NECROMANCER_REPORT.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "REPORT"
    headers = list(rows[0].keys()) if rows else ["nm_id"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    autosize_columns(ws)

    ws2 = wb.create_sheet("RUN")
    ws2["A1"] = "run_id"; ws2["B1"] = run_id
    ws2["A2"] = "created_at"; ws2["B2"] = (manifest.get("meta") or {}).get("created_at")
    ws2["A3"] = "script_version"; ws2["B3"] = (manifest.get("meta") or {}).get("script_version")
    autosize_columns(ws2)

    if exec_summary:
        ws3 = wb.create_sheet("SUMMARY")
        ws3["A1"] = safe_str(exec_summary.get("title") or "Executive Summary")
        ws3["A3"] = "summary_md"
        ws3["A4"] = safe_str(exec_summary.get("summary_md") or "")
        ws3["A6"] = "key_points"
        kp = exec_summary.get("key_points") or []
        for i, x in enumerate(kp, start=7):
            ws3[f"A{i}"] = f"- {safe_str(x)}"
        rr = exec_summary.get("risks") or []
        base = 7 + len(kp) + 1
        ws3[f"A{base}"] = "risks"
        for j, x in enumerate(rr, start=base+1):
            ws3[f"A{j}"] = f"- {safe_str(x)}"
        ns = exec_summary.get("next_steps") or []
        base2 = base + 1 + len(rr) + 1
        ws3[f"A{base2}"] = "next_steps"
        for k, x in enumerate(ns, start=base2+1):
            ws3[f"A{k}"] = f"- {safe_str(x)}"
        autosize_columns(ws3)

    wb.save(str(out_xlsx))

    out_html = out_dir / "WB_NECROMANCER_REPORT.html"
    html_rows = []
    for r in rows:
        rid = _html.escape(safe_str(r.get("nm_id")))
        v = _html.escape(safe_str(r.get("verdict")))
        flags = _html.escape(safe_str(r.get("risk_flags")))
        nmname = _html.escape(safe_str(r.get("name")))
        url = _html.escape(safe_str(r.get("url")))
        html_rows.append(f"""
<tr class="row" data-nm="{rid}" data-verdict="{v}">
  <td class="mono">{rid}</td>
  <td class="mono">{_html.escape(safe_str(r.get("vendor_code")))}</td>
  <td><a href="{url}" target="_blank" rel="noreferrer">{nmname}</a></td>
  <td>{_html.escape(safe_str(r.get("phone_model")))}</td>
  <td>{_html.escape(safe_str(r.get("market_phone")))} / {_html.escape(safe_str(r.get("market_type")))}</td>
  <td>{v}</td>
  <td class="mono">{flags}</td>
  <td><details><summary>rationale</summary><pre>{_html.escape(safe_str(r.get("rationale")))}</pre></details></td>
  <td><details><summary>backlog</summary><pre>{_html.escape(safe_str(r.get("backlog")))}</pre></details></td>
</tr>
""")
    css = """
body{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;margin:0;background:#0b1020;color:#e6e6e6}
header{position:sticky;top:0;background:#0f1220;border-bottom:1px solid #24283b;padding:12px 16px;z-index:5}
h1{margin:0;font-size:18px}
.container{padding:14px 16px}
.controls{display:flex;gap:10px;flex-wrap:wrap;margin:10px 0}
input,select{background:#0b1020;color:#e6e6e6;border:1px solid #24283b;border-radius:8px;padding:8px 10px}
table{width:100%;border-collapse:collapse}
th,td{border-bottom:1px solid #24283b;padding:8px 10px;vertical-align:top;font-size:12px}
th{position:sticky;top:74px;background:#0f1220;text-align:left}
a{color:#9ad1ff}
.mono{font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace}
pre{white-space:pre-wrap;margin:0}
.summary{border:1px solid #24283b;border-radius:12px;padding:12px 14px;margin:14px 0;background:#0f1220}
.summary h2{margin:0 0 8px 0;font-size:16px}
.summary .sumtext{font-size:12px;opacity:.95;margin-bottom:10px}
.summary .grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px}
.summary h3{margin:6px 0;font-size:13px}
.summary ul{margin:6px 0 0 18px;padding:0}
.details{max-width:480px}
details summary{cursor:pointer;opacity:.9}
.badge{display:inline-block;padding:2px 8px;border:1px solid #24283b;border-radius:999px;font-size:11px;margin-right:6px}
"""
    js = """
const rows = Array.from(document.querySelectorAll('tr.row'));
function applyFilters(){
  const term = (document.getElementById('q').value||'').trim().toLowerCase();
  const v = document.getElementById('verdict').value;
  let shown=0;
  for(const r of rows){
    let ok=true;
    if(term){
      const hay=(r.dataset.nm+' '+r.innerText).toLowerCase();
      ok = hay.includes(term);
    }
    if(ok && v!=='ALL') ok = (r.dataset.verdict===v);
    r.style.display = ok?'':'none';
    if(ok) shown++;
  }
  document.getElementById('shown').textContent = shown;
}
document.getElementById('q').addEventListener('input', applyFilters);
document.getElementById('verdict').addEventListener('change', applyFilters);
applyFilters();
"""

    summary_block = ""
    if exec_summary:
        t = _html.escape(safe_str(exec_summary.get("title") or "Executive Summary"))
        sm = _html.escape(safe_str(exec_summary.get("summary_md") or "")).replace("\n", "<br>")
        def _list(key: str) -> str:
            arr = exec_summary.get(key) or []
            if not isinstance(arr, list) or not arr:
                return ""
            items = "".join([f"<li>{_html.escape(safe_str(x))}</li>" for x in arr[:20]])
            return f"<ul>{items}</ul>"
        summary_block = f"""<section class='summary'>
  <h2>{t}</h2>
  <div class='sumtext'>{sm}</div>
  <div class='grid'>
    <div><h3>Ключевые пункты</h3>{_list('key_points')}</div>
    <div><h3>Риски</h3>{_list('risks')}</div>
    <div><h3>Следующие шаги</h3>{_list('next_steps')}</div>
  </div>
</section>"""

    html_doc = f"""<!doctype html>
<html lang="ru"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{_html.escape(title)}</title>
<style>{css}</style>
</head><body>
<header>
  <h1>{_html.escape(title)}</h1>
  <div class="mono">run_id={_html.escape(run_id)} • shown=<span id="shown">0</span> / {len(rows)}</div>
  <div class="controls">
    <input id="q" type="text" placeholder="Поиск nm_id, vendor_code, название..." />
    <select id="verdict">
      <option value="ALL">Все решения</option>
      <option value="REVIVE_FAST">REVIVE_FAST</option>
      <option value="REVIVE_REWORK">REVIVE_REWORK</option>
      <option value="CLONE_NEW_CARD">CLONE_NEW_CARD</option>
      <option value="DROP">DROP</option>
    </select>
  </div>
</header>
<div class="container">
{summary_block}
<table>
<thead><tr>
<th>nm_id</th><th>vendor_code</th><th>name</th><th>phone_model</th><th>market</th><th>verdict</th><th>flags</th><th>rationale</th><th>backlog</th>
</tr></thead>
<tbody>
{''.join(html_rows)}
</tbody>
</table>
</div>
<script>{js}</script>
</body></html>
"""
    out_html.write_text(html_doc, encoding="utf-8")

    stage_dbg("M", f"wrote {out_xlsx} + {out_html}")
    return out_xlsx, out_html


# =========================
# Runner + меню
# =========================

def stage_banner(code: str, args: argparse.Namespace) -> None:
    if not v(1):
        return
    title = _stage_title(code)
    net = _stage_effective_network(code, args)
    vpn = _stage_vpn_hint(code, args)

    llm_note = ""
    if _stage_llm_flag(code):
        llm_note = " | LLM=ON" if _stage_llm_enabled(code, args) else " | LLM=OFF"

    print("\n" + "=" * 78)
    print(f"[{code}] {title}")
    print(f"     net={net}{llm_note} | {vpn}")
    io = STAGE_IO.get(code) or {}
    if io:
        print(f"     in:  {io.get('in')}")
        print(f"     out: {io.get('out')}")
    print("=" * 78)



def run_stage(code: str, out_dir: Path, args: argparse.Namespace) -> None:
    """Одна стадия без пауз/баннеров. Здесь только бизнес-логика."""
    if code == "A":
        stage_A_manifest(
            out_dir,
            input_xlsx=Path(args.input).resolve(),
            sheet=args.sheet,
            expect_count=args.expect_count,
            dedupe=args.dedupe,
            dests=args.dests,
            search_hosts=args.search_hosts,
            search_limit=args.search_limit,
            lexicon_path=args.lexicon,
        )
    elif code == "B":
        stage_B_own_fetch(out_dir, timeout=args.wb_timeout, sleep_s=args.sleep, resume=args.resume, deep_card=(not args.no_deep_card))
    elif code == "C":
        stage_C_intent(out_dir, resume=args.resume)
    elif code == "D":
        stage_D_queries(
            out_dir,
            resume=args.resume,
            rules_per_cluster=args.rules_per_cluster,
            llm_extra_per_cluster=args.llm_extra_per_cluster,
            min_len=args.query_min_len,
            max_len=args.query_max_len,
            use_llm=args.use_llm_d,
            llm_provider=args.llm_provider,
            llm_model=args.llm_model,
            llm_base_url=args.llm_base_url,
            llm_timeout=args.llm_timeout,
            llm_max_tokens=args.llm_max_tokens_d,
            llm_temperature=args.llm_temperature,
        )
    elif code == "E":
        stage_E_serp(
            out_dir,
            timeout=args.wb_timeout,
            sleep_s=args.sleep,
            search_limit=args.search_limit,
            resume=args.resume,
            min_keep_per_cluster=args.min_keep_per_cluster,
            max_keep_per_cluster=args.max_keep_per_cluster,
            min_pass_rate_phone=args.min_pass_rate_phone,
            min_pass_rate_type=args.min_pass_rate_type,
        )
    elif code == "F":
        stage_F_pool(out_dir, resume=args.resume, competitors_k=args.competitors_k, per_query_take=args.per_query_take)
    elif code == "G":
        stage_G_lite(out_dir, timeout=args.wb_timeout, sleep_s=args.sleep, resume=args.resume)
    elif code == "H":
        stage_H_relevance(
            out_dir,
            resume=args.resume,
            use_llm=args.use_llm_h,
            min_keep_competitors=args.min_keep_competitors,
            max_keep_competitors=args.max_keep_competitors,
            llm_provider=args.llm_provider,
            llm_model=args.llm_model,
            llm_base_url=args.llm_base_url,
            llm_timeout=args.llm_timeout,
            llm_max_tokens=args.llm_max_tokens_h,
            llm_temperature=args.llm_temperature,
        )
    elif code == "I":
        stage_I_pulse(out_dir, timeout=args.wb_timeout, sleep_s=args.sleep_reviews, resume=args.resume, strict=args.strict)
    elif code == "J":
        stage_J_supply(out_dir, resume=args.resume)
    elif code == "K":
        stage_K_cluster_verdicts(out_dir, resume=args.resume)
    elif code == "L":
        stage_L_decisions(out_dir, resume=args.resume)
    elif code == "M":
        write_reports(out_dir, args=args, title=args.report_title)
    else:
        raise ValueError(f"Unknown stage: {code}")


def run_pipeline(args: argparse.Namespace) -> None:
    out_dir = Path(args.out).resolve()
    ensure_dir(out_dir)
    init_run_logger(out_dir)

    # Список стадий, которые реально будут выполняться
    try:
        si = STAGE_ORDER.index(args.start_stage)
        ei = STAGE_ORDER.index(args.end_stage)
    except Exception:
        raise ValueError(f"Bad stage bounds: {args.start_stage}..{args.end_stage}")
    if si > ei:
        raise ValueError(f"start-stage ({args.start_stage}) позже end-stage ({args.end_stage}). Это не квест.")
    plan = STAGE_ORDER[si:ei+1]

    # Какие стадии считаются LLM-стадиями в этом прогоне
    llm_stages: set[str] = set()
    if getattr(args, "use_llm_d", False):
        llm_stages.add("D")
    if getattr(args, "use_llm_h", False):
        llm_stages.add("H")
    if getattr(args, "use_llm_m", False):
        llm_stages.add("M")

    for idx, code in enumerate(plan):
        # ВАЖНО: баннер печатается ДО паузы, чтобы человек знал, что сейчас будет и нужен ли VPN.
        stage_banner(code, args)

        need_pause_between = bool(getattr(args, "pause_between_stages", False) and idx > 0)
        need_pause_before_llm = bool(
            getattr(args, "pause_before_llm", False)
            and (code in llm_stages)
            and (not getattr(args, "_did_pause_before_llm", False))
        )

        if need_pause_between or need_pause_before_llm:
            if getattr(args, "yes", False) or (not sys.stdin.isatty()):
                # headless: просто логируем, что пауза пропущена
                if need_pause_before_llm:
                    stage_line("", "[pause-before-llm] Пропускаю паузу (yes/headless).", level=1)
                    setattr(args, "_did_pause_before_llm", True)
                else:
                    stage_line("", "[pause] Пропускаю паузу (yes/headless).", level=1)
            else:
                if need_pause_before_llm:
                    input("Пауза перед LLM. Включи VPN (если надо) и жми Enter... ")
                    setattr(args, "_did_pause_before_llm", True)
                else:
                    input("Пауза. Нажми Enter, когда готов продолжать (например, включил VPN для LLM). ")


        t0 = time.time()
        try:
            run_stage(code, out_dir, args)
        except Exception as e:
            stage_fail(code, f"{type(e).__name__}: {safe_str(e)}")
            if v(2):
                traceback.print_exc()
            raise
        dt = time.time() - t0
        # Короткая строка что сделано и где артефакты
        stage_ok(code, f"done in {dt:.1f}s | {stage_artifacts_summary(code, out_dir)}", level=1)
        # Дебаг: сколько кэша насыпали
        if v(2):
            cs = stage_caches_summary(code, out_dir)
            if cs:
                stage_dbg(code, f"caches: {cs}")


def print_models() -> None:
    print("\nДоступные дешёвые модели (пресеты):")
    for prov, cfg in MODEL_PRESETS.items():
        print(f"- {prov}: default={cfg['default']}")
        for m in cfg["cheap"]:
            print(f"    • {m}")

def _ask_yes(prompt: str, default: bool) -> bool:
    # headless/CI/без stdin: молча берём дефолт, не устраиваем театр
    if not sys.stdin.isatty():
        return default
    s = input(prompt).strip().lower()
    if not s:
        return default
    return s in ("y", "yes", "да", "д", "1", "true", "ага")

def _ask_tokens_triplet(prompt: str, d0: int, h0: int, m0: int) -> Tuple[int, int, int]:
    """Парсер ввода вида:
    - пусто -> дефолты
    - одно число -> применить ко всем
    - три числа через запятую/пробел -> D,H,M
    """
    if not sys.stdin.isatty():
        return d0, h0, m0
    s = input(prompt).strip()
    if not s:
        return d0, h0, m0
    parts = [p for p in re.split(r"[\s,;]+", s) if p]
    nums: List[int] = []
    for p in parts[:3]:
        try:
            nums.append(int(p))
        except Exception:
            pass
    if not nums:
        return d0, h0, m0
    if len(nums) == 1:
        v = nums[0]
        if v <= 0:
            return d0, h0, m0
        return v, v, v
    # len 2/3: добиваем дефолтами
    while len(nums) < 3:
        nums.append(0)
    d, h, m = nums[0], nums[1], nums[2]
    d = d if d > 0 else d0
    h = h if h > 0 else h0
    m = m if m > 0 else m0
    return d, h, m


@dataclass
class MenuConfig:
    # Базовые пути
    input: str = "WB_INPUT_64_FROM_POCKETS_POD.xlsx"
    sheet: str = "INPUT_64"
    out: str = "WB_NECROMANCER_RUN"

    # Поведение раннера
    resume: bool = True
    pause_between: bool = True
    pause_before_llm: bool = True
    yes: bool = False

    # Вывод
    verbosity: int = 1  # 0/1/2

    # LLM
    use_llm_d: bool = False
    use_llm_h: bool = False
    use_llm_m: bool = False

    # Лимиты токенов (None = использовать дефолты пайплайна/аргпарсера)
    llm_max_tokens: Optional[int] = None
    llm_max_tokens_d: Optional[int] = None
    llm_max_tokens_h: Optional[int] = None
    llm_max_tokens_m: Optional[int] = None

    # Relevance sanity (Stage H)
    min_keep_competitors: int = 6
    max_keep_competitors: int = 18

# --- Menu config persistence (профиль меню) ---
# По умолчанию сохраняем настройки меню рядом со скриптом: ./profiles/necromancer_menu_last.json
MENU_PROFILES_DIR = Path(__file__).resolve().parent / "profiles"
MENU_PROFILE_DEFAULT = MENU_PROFILES_DIR / "necromancer_menu_last.json"

def menu_profile_path() -> Path:
    """Путь к профилю меню. Можно переопределить env NECROMANCER_MENU_PROFILE."""
    env = (os.getenv("NECROMANCER_MENU_PROFILE") or "").strip()
    if env:
        try:
            return Path(env).expanduser().resolve()
        except Exception:
            return Path(env).expanduser()
    return MENU_PROFILE_DEFAULT

def cfg_save(cfg: "MenuConfig", path: Optional[Path] = None) -> Optional[Path]:
    """Сохранить настройки меню в json. Возвращает путь или None если не вышло."""
    path = path or menu_profile_path()
    cfg_sanitize(cfg)
    try:
        ensure_dir(path.parent)
        tmp = path.with_suffix(path.suffix + ".tmp")
        tmp.write_text(json.dumps(asdict(cfg), ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(path)
        return path
    except Exception as e:
        veprint(2, f"[CFG][DBG] save failed: {e}")
        return None

def cfg_load(cfg: "MenuConfig", path: Optional[Path] = None) -> Tuple["MenuConfig", bool]:
    """Загрузить настройки меню из json. Возвращает (cfg, loaded_ok)."""
    path = path or menu_profile_path()
    try:
        raw = path.read_text(encoding="utf-8")
        data = json.loads(raw) if raw.strip() else {}
    except FileNotFoundError:
        return cfg, False
    except Exception as e:
        veprint(2, f"[CFG][DBG] load failed: {e}")
        return cfg, False

    if not isinstance(data, dict):
        return cfg, False

    allowed = {f.name for f in fields(MenuConfig)}
    for k, v0 in data.items():
        if k not in allowed:
            continue
        try:
            setattr(cfg, k, v0)
        except Exception:
            pass
    cfg_sanitize(cfg)
    return cfg, True

def cfg_reset(cfg: "MenuConfig") -> None:
    """Сбросить настройки меню к дефолтам."""
    fresh = MenuConfig()
    for f in fields(MenuConfig):
        setattr(cfg, f.name, getattr(fresh, f.name))
    cfg_sanitize(cfg)


def cfg_sanitize(cfg: "MenuConfig") -> None:
    """Привести cfg к валидному виду (после load/ручной порчи/настроек)."""
    # строки
    cfg.input = safe_str(cfg.input).strip() or "WB_INPUT_64_FROM_POCKETS_POD.xlsx"
    cfg.sheet = safe_str(cfg.sheet).strip() or "INPUT_64"
    cfg.out = safe_str(cfg.out).strip() or "WB_NECROMANCER_RUN"

    # bool
    cfg.resume = bool(cfg.resume)
    cfg.pause_between = bool(cfg.pause_between)
    cfg.pause_before_llm = bool(cfg.pause_before_llm)
    cfg.yes = bool(cfg.yes)
    cfg.use_llm_d = bool(cfg.use_llm_d)
    cfg.use_llm_h = bool(cfg.use_llm_h)
    cfg.use_llm_m = bool(cfg.use_llm_m)

    # verbosity
    try:
        cfg.verbosity = int(cfg.verbosity)
    except Exception:
        cfg.verbosity = 1
    cfg.verbosity = 0 if cfg.verbosity < 0 else (2 if cfg.verbosity > 2 else cfg.verbosity)

    # thresholds
    try:
        cfg.min_keep_competitors = int(cfg.min_keep_competitors)
    except Exception:
        cfg.min_keep_competitors = 6
    try:
        cfg.max_keep_competitors = int(cfg.max_keep_competitors)
    except Exception:
        cfg.max_keep_competitors = 18
    if cfg.min_keep_competitors < 1:
        cfg.min_keep_competitors = 1
    if cfg.max_keep_competitors < cfg.min_keep_competitors:
        cfg.max_keep_competitors = cfg.min_keep_competitors

    # tokens: <=0 -> None, иначе int
    def _tok(x: Optional[int]) -> Optional[int]:
        if x is None:
            return None
        try:
            v0 = int(x)
        except Exception:
            return None
        return v0 if v0 > 0 else None

    cfg.llm_max_tokens = _tok(cfg.llm_max_tokens)
    cfg.llm_max_tokens_d = _tok(cfg.llm_max_tokens_d)
    cfg.llm_max_tokens_h = _tok(cfg.llm_max_tokens_h)
    cfg.llm_max_tokens_m = _tok(cfg.llm_max_tokens_m)


def _ask_int(prompt: str, default: int, min_v: int = 0, max_v: int = 1000000) -> int:
    if not sys.stdin.isatty():
        return default
    s = input(prompt).strip()
    if not s:
        return default
    try:
        v = int(s)
    except Exception:
        return default
    if v < min_v:
        return default
    if v > max_v:
        return default
    return v


def _cfg_summary(cfg: MenuConfig) -> str:
    llm = []
    if cfg.use_llm_d: llm.append("D")
    if cfg.use_llm_h: llm.append("H")
    if cfg.use_llm_m: llm.append("M")
    llm_s = ",".join(llm) if llm else "OFF"
    toks = []
    if cfg.llm_max_tokens is not None:
        toks.append(f"all={cfg.llm_max_tokens}")
    if cfg.llm_max_tokens_d is not None:
        toks.append(f"D={cfg.llm_max_tokens_d}")
    if cfg.llm_max_tokens_h is not None:
        toks.append(f"H={cfg.llm_max_tokens_h}")
    if cfg.llm_max_tokens_m is not None:
        toks.append(f"M={cfg.llm_max_tokens_m}")
    toks_s = " ".join(toks) if toks else "defaults"
    return (
        f"input={cfg.input} sheet={cfg.sheet} out={cfg.out}\n"
        f"resume={'Y' if cfg.resume else 'N'} pauses={'Y' if cfg.pause_between else 'N'} "
        f"pause_before_llm={'Y' if cfg.pause_before_llm else 'N'} yes={'Y' if cfg.yes else 'N'}\n"
        f"verbosity={cfg.verbosity} | LLM={llm_s} | max_tokens={toks_s} | keep={cfg.min_keep_competitors}..{cfg.max_keep_competitors}"
    )


def _settings_menu(cfg: MenuConfig) -> None:
    while True:
        print("\nНастройки (меняй что хочешь, только потом не удивляйся):")
        print(f"Профиль меню: {menu_profile_path()}")
        print(_cfg_summary(cfg))
        print("\nПункты:")
        print("  1) input xlsx")
        print("  2) sheet")
        print("  3) out folder")
        print("  4) resume (toggle)")
        print("  5) pause between stages (toggle)")
        print("  6) pause before first LLM (toggle)")
        print("  7) yes (auto-continue) (toggle)")
        print("  8) verbosity (0/1/2)")
        print("  9) LLM toggles (D/H/M)")
        print(" 10) max_tokens (global + per-stage, можно очистить)")
        print(" 11) keep thresholds for Stage H (min/max)")
        print(" 12) save settings (profile)")
        print(" 13) load settings (profile)")
        print(" 14) reset settings to defaults")
        print("  0) назад")

        ch = (input("Выбор: ").strip() if sys.stdin.isatty() else "0")
        if ch == "0":
            cfg_save(cfg)
            return

        if ch == "1":
            if sys.stdin.isatty():
                s = input(f"input xlsx (Enter оставить {cfg.input}): ").strip()
                if s:
                    cfg.input = s

        elif ch == "2":
            if sys.stdin.isatty():
                s = input(f"sheet (Enter оставить {cfg.sheet}): ").strip()
                if s:
                    cfg.sheet = s

        elif ch == "3":
            if sys.stdin.isatty():
                s = input(f"out folder (Enter оставить {cfg.out}): ").strip()
                if s:
                    cfg.out = s

        elif ch == "4":
            cfg.resume = not cfg.resume

        elif ch == "5":
            cfg.pause_between = not cfg.pause_between

        elif ch == "6":
            cfg.pause_before_llm = not cfg.pause_before_llm

        elif ch == "7":
            cfg.yes = not cfg.yes

        elif ch == "8":
            cfg.verbosity = _ask_int("verbosity 0/1/2 (Enter=оставить): ", cfg.verbosity, 0, 2)

        elif ch == "9":
            cfg.use_llm_d = _ask_yes(f"LLM на D? [{'Y' if cfg.use_llm_d else 'N'}]: ", cfg.use_llm_d)
            cfg.use_llm_h = _ask_yes(f"LLM на H? [{'Y' if cfg.use_llm_h else 'N'}]: ", cfg.use_llm_h)
            cfg.use_llm_m = _ask_yes(f"LLM на M? [{'Y' if cfg.use_llm_m else 'N'}]: ", cfg.use_llm_m)

        elif ch == "10":
            print("\nmax_tokens:")
            print("  a) global (all stages)")
            print("  d) stage D")
            print("  h) stage H")
            print("  m) stage M")
            print("  c) clear (сбросить всё)")
            sub = (input("Выбор: ").strip().lower() if sys.stdin.isatty() else "")
            if sub == "c":
                cfg.llm_max_tokens = None
                cfg.llm_max_tokens_d = None
                cfg.llm_max_tokens_h = None
                cfg.llm_max_tokens_m = None
            elif sub == "a":
                cfg.llm_max_tokens = _ask_int("global max_tokens (0=сброс): ", cfg.llm_max_tokens or 0, 0, 100000)
                if cfg.llm_max_tokens <= 0:
                    cfg.llm_max_tokens = None
            elif sub == "d":
                cfg.llm_max_tokens_d = _ask_int("D max_tokens (0=сброс): ", cfg.llm_max_tokens_d or 0, 0, 100000)
                if cfg.llm_max_tokens_d <= 0:
                    cfg.llm_max_tokens_d = None
            elif sub == "h":
                cfg.llm_max_tokens_h = _ask_int("H max_tokens (0=сброс): ", cfg.llm_max_tokens_h or 0, 0, 100000)
                if cfg.llm_max_tokens_h <= 0:
                    cfg.llm_max_tokens_h = None
            elif sub == "m":
                cfg.llm_max_tokens_m = _ask_int("M max_tokens (0=сброс): ", cfg.llm_max_tokens_m or 0, 0, 100000)
                if cfg.llm_max_tokens_m <= 0:
                    cfg.llm_max_tokens_m = None

        elif ch == "11":
            cfg.min_keep_competitors = _ask_int("min_keep_competitors: ", cfg.min_keep_competitors, 1, 200)
            cfg.max_keep_competitors = _ask_int("max_keep_competitors: ", cfg.max_keep_competitors, cfg.min_keep_competitors, 500)

        elif ch == "12":
            pth = cfg_save(cfg)
            if pth:
                print(f"Ок. Сохранил профиль: {pth}")
            else:
                print("Не смог сохранить профиль (права/путь).")

        elif ch == "13":
            _, ok = cfg_load(cfg)
            if ok:
                print(f"Ок. Загрузил профиль: {menu_profile_path()}")
            else:
                print("Профиль не найден или битый. Нечего грузить.")

        elif ch == "14":
            cfg_reset(cfg)
            print("Сбросил настройки к дефолтам.")

        else:
            print("Мимо.")

def _menu_run(cfg: MenuConfig) -> Tuple[str, str, bool, bool, bool]:
    """Возвращает start,end и фактические флаги LLM для этого прогона."""
    print("\nРежимы запуска:")
    print("  1) Полный прогон A..M (rules-only)")
    print("  2) Только WB-часть A..G (rules-only)  [собрать данные без VPN]")
    print("  3) Только хвост H..M (после A..G)     [VPN для LLM, если включишь]")
    print("  4) Полный прогон A..M (LLM на D+H, опц LLM в M)")
    print("  5) Кастом: выбрать start/end + LLM-флаги")
    ch = (input("Выбор: ").strip() if sys.stdin.isatty() else "1")

    use_llm_d = cfg.use_llm_d
    use_llm_h = cfg.use_llm_h
    use_llm_m = cfg.use_llm_m

    if ch == "1":
        return "A", "M", False, False, False
    if ch == "2":
        return "A", "G", False, False, False
    if ch == "3":
        # по умолчанию для хвоста включаем H (иначе он бессмысленный)
        use_llm_h = True
        use_llm_m = _ask_yes("В M делать LLM-выжимку (exec summary)? [y/N]: ", False)
        return "H", "M", False, use_llm_h, use_llm_m
    if ch == "4":
        use_llm_d = True
        use_llm_h = True
        use_llm_m = _ask_yes("В M делать LLM-выжимку (exec summary)? [Y/n]: ", True)
        return "A", "M", use_llm_d, use_llm_h, use_llm_m
    # кастом
    st = (input("start-stage (A..M): ").strip().upper() if sys.stdin.isatty() else "A") or "A"
    en = (input("end-stage (A..M): ").strip().upper() if sys.stdin.isatty() else "M") or "M"
    if st not in STAGE_ORDER: st = "A"
    if en not in STAGE_ORDER: en = "M"
    use_llm_d = _ask_yes(f"LLM на D? [{'Y' if use_llm_d else 'N'}]: ", use_llm_d)
    use_llm_h = _ask_yes(f"LLM на H? [{'Y' if use_llm_h else 'N'}]: ", use_llm_h)
    use_llm_m = _ask_yes(f"LLM на M? [{'Y' if use_llm_m else 'N'}]: ", use_llm_m)
    return st, en, use_llm_d, use_llm_h, use_llm_m


def _prompt_tokens_per_stage(cfg: MenuConfig, use_llm_d: bool, use_llm_h: bool, use_llm_m: bool) -> None:
    """Просим max_tokens отдельно для каждого этапа (как ты и хотел, без этих 'триплетов')."""
    # дефолты как в плане
    d0, h0, m0 = 900, 1200, 1100

    def _def(stage: str, local: Optional[int], base: int) -> int:
        if local is not None and local > 0:
            return local
        if cfg.llm_max_tokens is not None and cfg.llm_max_tokens > 0:
            return cfg.llm_max_tokens
        return base

    if not sys.stdin.isatty():
        return

    if use_llm_d:
        dd = _def("D", cfg.llm_max_tokens_d, d0)
        cfg.llm_max_tokens_d = _ask_int(f"max_tokens для LLM D (Enter={dd}): ", dd, 64, 100000)

    if use_llm_h:
        hh = _def("H", cfg.llm_max_tokens_h, h0)
        cfg.llm_max_tokens_h = _ask_int(f"max_tokens для LLM H (Enter={hh}): ", hh, 64, 100000)

    if use_llm_m:
        mm = _def("M", cfg.llm_max_tokens_m, m0)
        cfg.llm_max_tokens_m = _ask_int(f"max_tokens для LLM M (Enter={mm}): ", mm, 64, 100000)


def _cfg_to_argv(cfg: MenuConfig, start_stage: str, end_stage: str,
                use_llm_d: bool, use_llm_h: bool, use_llm_m: bool) -> List[str]:
    argv: List[str] = []
    argv += ["--input", cfg.input, "--sheet", cfg.sheet, "--out", cfg.out]
    argv += ["--start-stage", start_stage, "--end-stage", end_stage]
    argv += ["--verbosity", str(cfg.verbosity)]
    argv += ["--min-keep-competitors", str(cfg.min_keep_competitors), "--max-keep-competitors", str(cfg.max_keep_competitors)]

    if cfg.resume:
        argv.append("--resume")

    if cfg.pause_between:
        argv.append("--pause-between-stages")

    if (use_llm_d or use_llm_h or use_llm_m) and cfg.pause_before_llm:
        argv.append("--pause-before-llm")

    if cfg.yes:
        argv.append("--yes")

    if use_llm_d:
        argv.append("--use-llm-d")
    if use_llm_h:
        argv.append("--use-llm-h")
    if use_llm_m:
        argv.append("--use-llm-m")
    # max_tokens:
    # - global задаёт общий потолок
    # - per-stage может переопределять global (в main global применяется ТОЛЬКО если stage=None)
    if (use_llm_d or use_llm_h or use_llm_m) and cfg.llm_max_tokens is not None and cfg.llm_max_tokens > 0:
        argv += ["--llm-max-tokens", str(cfg.llm_max_tokens)]

    # по стадиям: добавляем только те, что реально используются
    if use_llm_d and cfg.llm_max_tokens_d is not None and cfg.llm_max_tokens_d > 0:
        argv += ["--llm-max-tokens-d", str(cfg.llm_max_tokens_d)]
    if use_llm_h and cfg.llm_max_tokens_h is not None and cfg.llm_max_tokens_h > 0:
        argv += ["--llm-max-tokens-h", str(cfg.llm_max_tokens_h)]
    if use_llm_m and cfg.llm_max_tokens_m is not None and cfg.llm_max_tokens_m > 0:
        argv += ["--llm-max-tokens-m", str(cfg.llm_max_tokens_m)]

    return argv


def interactive_menu() -> List[str]:
    cfg = MenuConfig()
    cfg_load(cfg)  # авто-подхват last profile

    while True:
        print("\nWB Некромант v2 (rewrite) — меню.")
        print(_cfg_summary(cfg))
        print(f"(профиль: {menu_profile_path()})")

        print("\nГлавное меню:")
        print("  1) Запуск (выбрать режим)")
        print("  2) Настройка (всё настраивается тут)")
        print("  3) Показать стадии")
        print("  4) Показать список моделей (пресеты)")
        print("  0) Выход")

        ch = (input("Выбор: ").strip() if sys.stdin.isatty() else "0")
        if ch == "0":
            cfg_save(cfg)
            raise SystemExit(0)

        if ch == "3":
            print_stage_table()
            continue

        if ch == "4":
            print_models()
            continue

        if ch == "2":
            _settings_menu(cfg)
            cfg_save(cfg)
            continue

        if ch == "1":
            st, en, use_llm_d, use_llm_h, use_llm_m = _menu_run(cfg)

            if use_llm_d or use_llm_h or use_llm_m:
                _prompt_tokens_per_stage(cfg, use_llm_d, use_llm_h, use_llm_m)

            cfg_save(cfg)
            return _cfg_to_argv(cfg, st, en, use_llm_d, use_llm_h, use_llm_m)

        print("Мимо.")

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog=SCRIPT_NAME,
        description="WB Revival v2 Necromancer — rewrite: staged pipeline A..M with JSONL artifacts + XLSX+HTML report.",
    )
    p.add_argument("--menu", action="store_true", help="Интерактивное меню (если без аргументов — тоже стартует).")

    # verbosity по умолчанию = 1 (нормально). 2 = дебаг. 0 = тихо.
    p.add_argument("--verbosity", type=int, choices=[0, 1, 2], default=1,
                   help="Уровень болтливости: 0 тихо (без прогресс-баров), 1 нормально (дефолт), 2 дебаг.")
    p.add_argument("-v", "--verbose", dest="verbosity", action="store_const", const=2,
                   help="Синоним: --verbosity 2 (чуть больше инфы).")
    p.add_argument("--quiet", "--no-verbose", dest="verbosity", action="store_const", const=0,
                   help="Синоним: --verbosity 0 (без прогресс-баров).")

    p.add_argument("--input", default="WB_INPUT_64_FROM_POCKETS_POD.xlsx", help="Input XLSX")
    p.add_argument("--sheet", default="INPUT_64", help="Sheet name")
    p.add_argument("--expect-count", type=int, default=None, help="Опционально: требовать точное число строк")
    p.add_argument("--dedupe", action="store_true", help="Молчаливо выкинуть дубли nm_id вместо падения")
    p.add_argument("--out", default="WB_NECROMANCER_RUN", help="Папка вывода")

    p.add_argument("--start-stage", default="A", choices=STAGE_ORDER)
    p.add_argument("--end-stage", default="M", choices=STAGE_ORDER)
    p.add_argument("--resume", action="store_true", help="Продолжить (skip уже записанные nm_id в jsonl)")

    p.add_argument("--dests", nargs="+", type=int, default=DEFAULT_DESTS)
    p.add_argument("--search-hosts", nargs="+", default=DEFAULT_SEARCH_HOSTS)
    p.add_argument("--search-limit", type=int, default=100)
    p.add_argument("--wb-timeout", type=int, default=30)
    p.add_argument("--sleep", type=float, default=0.35, help="Sleep между WB запросами (сек)")
    p.add_argument("--no-deep-card", action="store_true", help="Не тянуть wbbasket deep card.json")

    p.add_argument("--lexicon", default="", help="Путь к lexicon.json (опционально)")

    p.add_argument("--rules-per-cluster", type=int, default=8)
    p.add_argument("--llm-extra-per-cluster", type=int, default=3)
    p.add_argument("--query-min-len", type=int, default=8)
    p.add_argument("--query-max-len", type=int, default=64)

    p.add_argument("--min-keep-per-cluster", type=int, default=2)
    p.add_argument("--max-keep-per-cluster", type=int, default=5)
    p.add_argument("--min-pass-rate-phone", type=float, default=0.12)
    p.add_argument("--min-pass-rate-type", type=float, default=0.08)

    p.add_argument("--competitors-k", type=int, default=18, help="Сколько конкурентов брать на кластер")
    p.add_argument("--per-query-take", type=int, default=40, help="Сколько верхних позиций из каждой выдачи брать в пул")

    p.add_argument("--min-keep-competitors", type=int, default=6, help="Stage H: минимум KEEP конкурентов на кластер (fallback если меньше)")
    p.add_argument("--max-keep-competitors", type=int, default=18, help="Stage H: максимум KEEP конкурентов на кластер (обрезка если больше)")

    p.add_argument("--sleep-reviews", type=float, default=0.25)
    p.add_argument("--strict", action="store_true", help="Строгий режим (падать чаще).")

    p.add_argument("--use-llm-d", action="store_true", help="LLM обогащение запросов (Stage D)")
    p.add_argument("--use-llm-h", action="store_true", help="LLM для пограничной релевантности (Stage H)")
    p.add_argument("--use-llm-m", action="store_true", help="LLM выжимка для отчёта (Stage M: exec summary)")
    p.add_argument("--llm-provider", choices=["openai", "openrouter"], default=os.environ.get("LLM_PROVIDER","openrouter"))
    p.add_argument("--llm-model", default=os.environ.get("LLM_MODEL", MODEL_PRESETS.get(os.environ.get("LLM_PROVIDER","openrouter"), MODEL_PRESETS["openrouter"])["default"]))
    p.add_argument("--llm-base-url", default=os.environ.get("LLM_BASE_URL",""))
    p.add_argument("--llm-timeout", type=int, default=60)
    p.add_argument("--llm-temperature", type=float, default=0.2)
    p.add_argument("--llm-max-tokens", type=int, default=None, help="Глобальный max_tokens для всех LLM стадий (если не заданы --llm-max-tokens-d/h/m)")
    p.add_argument("--llm-max-tokens-d", type=int, default=None)
    p.add_argument("--llm-max-tokens-h", type=int, default=None)
    p.add_argument("--llm-max-tokens-m", type=int, default=None)

    p.add_argument("--pause-between-stages", action="store_true", help="Пауза между стадиями (удобно для VPN)")
    p.add_argument("--pause-before-llm", action="store_true", help="Одна пауза перед первой LLM-стадией (включи VPN)")
    p.add_argument("--yes", action="store_true", help="Не задавать вопросы, авто-продолжение (для headless)")
    p.add_argument("--list-stages", action="store_true", help="Показать стадии и выйти")
    p.add_argument("--list-models", action="store_true", help="Показать модели и выйти")

    p.add_argument("--report-title", default="WB Necromancer v2 Report")

    return p

def main(argv: Optional[List[str]] = None) -> None:
    argv = list(argv) if argv is not None else sys.argv[1:]
    if not argv:
        argv = ["--menu"]

    if "--menu" in argv:
        extra = interactive_menu()
        if extra == ["--menu"]:
            return
        argv = [a for a in argv if a != "--menu"] + extra

    p = build_parser()
    args = p.parse_args(argv)

    # Влияет на прогресс-бары и часть болтовни.
    set_global_verbosity(args.verbosity)

    # max_tokens: глобальный -> стадийные -> дефолты
    DEFAULT_TOKENS_D, DEFAULT_TOKENS_H, DEFAULT_TOKENS_M = 900, 1200, 1100
    gtok = getattr(args, "llm_max_tokens", None)
    if isinstance(gtok, int) and gtok > 0:
        if getattr(args, "llm_max_tokens_d", None) is None:
            args.llm_max_tokens_d = gtok
        if getattr(args, "llm_max_tokens_h", None) is None:
            args.llm_max_tokens_h = gtok
        if getattr(args, "llm_max_tokens_m", None) is None:
            args.llm_max_tokens_m = gtok

    if not isinstance(getattr(args, "llm_max_tokens_d", None), int) or args.llm_max_tokens_d <= 0:
        args.llm_max_tokens_d = DEFAULT_TOKENS_D
    if not isinstance(getattr(args, "llm_max_tokens_h", None), int) or args.llm_max_tokens_h <= 0:
        args.llm_max_tokens_h = DEFAULT_TOKENS_H
    if not isinstance(getattr(args, "llm_max_tokens_m", None), int) or args.llm_max_tokens_m <= 0:
        args.llm_max_tokens_m = DEFAULT_TOKENS_M

    if v(2):
        veprint(2, "[CFG][DBG] "
                f"input={args.input} sheet={args.sheet} out={args.out} "
                f"stages={args.start_stage}..{args.end_stage} resume={args.resume} "
                f"pauses=between:{args.pause_between_stages} before_llm:{args.pause_before_llm} yes={args.yes} "
                f"LLM=D:{args.use_llm_d} H:{args.use_llm_h} M:{args.use_llm_m} "
                f"max_tokens=D:{args.llm_max_tokens_d} H:{args.llm_max_tokens_h} M:{args.llm_max_tokens_m} "
                f"keep={args.min_keep_competitors}..{args.max_keep_competitors}")

    if args.list_stages:
        print_stage_table(args)
        return

    if args.list_models:
        print_models()
        return

    if args.use_llm_d or args.use_llm_h or args.use_llm_m:
        key = llm_api_key(args.llm_provider)
        if not key:
            env_path = Path(args.out).resolve() / ".env"
            ensure_dir(env_path.parent)
            print(f"Нету API ключа для {args.llm_provider}. Сейчас спросим.")
            val = getpass.getpass("Введи ключ (не будет видно): ").strip()
            if not val:
                raise SystemExit("Ключ пустой. Без ключа LLM не работает.")
            if args.llm_provider == "openrouter":
                os.environ["OPENROUTER_API_KEY"] = val
                write_dotenv(env_path, {"OPENROUTER_API_KEY": val, "LLM_PROVIDER": "openrouter", "LLM_MODEL": args.llm_model})
            else:
                os.environ["OPENAI_API_KEY"] = val
                write_dotenv(env_path, {"OPENAI_API_KEY": val, "LLM_PROVIDER": "openai", "LLM_MODEL": args.llm_model})
            print(f"Ок, сохранил ключ в {env_path} (plain text, да).")

    run_pipeline(args)

if __name__ == "__main__":
    main()
